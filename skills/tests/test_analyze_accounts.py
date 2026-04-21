# tests/test_analyze_accounts.py
import json
import time
import urllib.request
import pytest
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from analyze_accounts import (
    strip_json_fences,
    compute_stats,
    format_image_preference,
    compute_interaction_rate,
    compute_stickiness,
    filter_notes_by_recent_month,
    is_valid_detailed_note,
    is_complete_analysis_row,
    _extract_cover_url,
    _is_safe_http_url,
    _parse_analysis_output,
    _cookie_str_to_list,
    analyze_with_claude,
    process_account,
    get_analysis_workbook_path,
    HEADERS,
    XhsFetcher,
)

# ── strip_json_fences ──────────────────────────────────────────────────────────

def test_strip_json_fences_removes_markdown_block():
    text = '```json\n{"a": 1}\n```'
    assert strip_json_fences(text) == '{"a": 1}'

def test_strip_json_fences_plain_json_unchanged():
    text = '{"a": 1}'
    assert strip_json_fences(text) == '{"a": 1}'

def test_strip_json_fences_strips_whitespace():
    text = '  ```json\n{"a": 1}\n```  '
    assert json.loads(strip_json_fences(text)) == {"a": 1}

# ── compute_stats ──────────────────────────────────────────────────────────────

SAMPLE_NOTES = [
    {"image_count": 3, "word_count": 150, "liked_count": 200, "collected_count": 80, "comment_count": 12},
    {"image_count": 1, "word_count": 80,  "liked_count": 100, "collected_count": 20, "comment_count": 5},
    {"image_count": 5, "word_count": 200, "liked_count": 300, "collected_count": 120, "comment_count": 20},
]

def test_compute_stats_averages():
    stats = compute_stats(SAMPLE_NOTES)
    assert stats["avg_images"] == pytest.approx(3.0)
    assert stats["avg_words"] == 143       # (150+80+200)//3
    assert stats["avg_liked"] == 200       # (200+100+300)//3
    assert stats["avg_collected"] == 73    # (80+20+120)//3
    assert stats["avg_comments"] == 12     # (12+5+20)//3

def test_compute_stats_empty_notes():
    stats = compute_stats([])
    assert stats["avg_images"] == 0
    assert stats["avg_liked"] == 0


def test_compute_stats_tolerates_none_for_image_and_comment_counts():
    stats = compute_stats([{
        "image_count": None,
        "word_count": 10,
        "liked_count": 20,
        "collected_count": 5,
        "comment_count": None,
    }])
    assert stats["avg_images"] == 0
    assert stats["avg_comments"] == 0


def test_filter_notes_by_recent_month_marks_missing_create_time():
    notes = [{"create_time": None}]
    out = filter_notes_by_recent_month(notes)
    assert out == [{"create_time": None, "create_time_missing": True}]


def test_filter_notes_by_recent_month_drops_old_notes():
    notes = [{"create_time": "2020-01-01"}]
    out = filter_notes_by_recent_month(notes)
    assert out == []


def test_filter_notes_by_recent_month_mixed_missing_and_old():
    notes = [{"create_time": None}, {"create_time": "2020-01-01"}]
    out = filter_notes_by_recent_month(notes)
    assert len(out) == 1
    assert out[0]["create_time_missing"] is True

# ── format_image_preference ────────────────────────────────────────────────────

def test_format_image_preference_single():
    notes = [{"image_count": 1}] * 5
    result = format_image_preference(notes)
    assert "单图" in result

def test_format_image_preference_multi():
    notes = [{"image_count": 6}] * 5
    result = format_image_preference(notes)
    assert "多图" in result

def test_format_image_preference_mixed():
    notes = [{"image_count": 1}] * 3 + [{"image_count": 5}] * 3
    result = format_image_preference(notes)
    assert "混合" in result

# ── compute_interaction_rate ───────────────────────────────────────────────────

def test_compute_interaction_rate_normal():
    result = compute_interaction_rate(avg_liked=100, avg_collected=50, fans_count=1000)
    assert result == "15.0%"

def test_compute_interaction_rate_zero_fans():
    result = compute_interaction_rate(avg_liked=100, avg_collected=50, fans_count=0)
    assert result == "N/A"

# ── compute_stickiness ─────────────────────────────────────────────────────────

def test_compute_stickiness_normal():
    result = compute_stickiness(avg_collected=40, avg_liked=100)
    assert result == "40.0%"

def test_compute_stickiness_zero_liked():
    result = compute_stickiness(avg_collected=40, avg_liked=0)
    assert result == "N/A"


# ── write_excel ────────────────────────────────────────────────────────────────

import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
from analyze_accounts import write_excel
from openpyxl import load_workbook


def test_write_excel_creates_file(tmp_path):
    out = str(tmp_path / "test.xlsx")
    write_excel([], out)
    assert Path(out).exists()


def test_write_excel_correct_headers(tmp_path):
    out = str(tmp_path / "test.xlsx")
    write_excel([], out)
    wb = load_workbook(out)
    ws = wb.active
    actual = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
    assert actual == HEADERS


def test_write_excel_data_row(tmp_path):
    out = str(tmp_path / "test.xlsx")
    row = {h: f"v{i}" for i, h in enumerate(HEADERS)}
    write_excel([row], out)
    wb = load_workbook(out)
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(HEADERS) + 1)]
    rating_col = headers.index("账号权重评级") + 1
    assert ws.cell(row=2, column=1).value == row["账号昵称"]
    assert ws.cell(row=2, column=rating_col).value == row["账号权重评级"]


def test_get_analysis_workbook_path_points_to_cumulative_file():
    path = get_analysis_workbook_path()
    assert path.name == "xhs_account_analysis.xlsx"
    assert path.parent.name == "爆款分析"


def test_is_valid_detailed_note_requires_explicit_success_flag():
    assert is_valid_detailed_note({"detail_fetched": True}) is True
    assert is_valid_detailed_note({"detail_fetched": False}) is False
    assert is_valid_detailed_note({"title": "只有标题"}) is False


def test_is_complete_analysis_row_filters_incomplete_ratings():
    assert is_complete_analysis_row({"账号权重评级": "A 内容稳定"}) is True
    assert is_complete_analysis_row({"账号权重评级": "分析失败"}) is False
    assert is_complete_analysis_row({"账号权重评级": "低活跃度-无法获取详情"}) is False
    assert is_complete_analysis_row({"账号权重评级": "N/A"}) is False


def test_extract_cover_url_prefers_cover_and_falls_back_to_image_list():
    note1 = {"cover": {"url_default": "https://img.example.com/a.jpg"}}
    note2 = {"cover": {}, "image_list": [{"url_pre": "https://img.example.com/b.jpg"}]}
    note3 = {"cover": {}, "image_list": []}
    assert _extract_cover_url(note1) == "https://img.example.com/a.jpg"
    assert _extract_cover_url(note2) == "https://img.example.com/b.jpg"
    assert _extract_cover_url(note3) == ""


def test_is_safe_http_url_blocks_local_and_non_http():
    assert _is_safe_http_url("https://sns-img-qc.xhscdn.com/abc.jpg") is True
    assert _is_safe_http_url("http://img.example.com/a.jpg") is True
    assert _is_safe_http_url("ftp://img.example.com/a.jpg") is False
    assert _is_safe_http_url("http://localhost:8080/a.jpg") is False
    assert _is_safe_http_url("http://127.0.0.1/a.jpg") is False
    assert _is_safe_http_url("javascript:alert(1)") is False
    assert _is_safe_http_url("") is False


def test_parse_analysis_output_supports_pipe_and_json():
    text_pipe = "图文风格描述|简洁信息流\n账号权重评级|A 稳定"
    parsed_pipe = _parse_analysis_output(text_pipe)
    assert parsed_pipe["图文风格描述"] == "简洁信息流"
    assert parsed_pipe["账号权重评级"] == "A 稳定"

    text_json = """```json
{"图文风格描述":"视觉统一","账号权重评级":"B 成长中"}
```"""
    parsed_json = _parse_analysis_output(text_json)
    assert parsed_json["图文风格描述"] == "视觉统一"
    assert parsed_json["账号权重评级"] == "B 成长中"


def test_cookie_str_to_list_applies_known_and_default_attrs():
    cookies = _cookie_str_to_list("web_session=abc; a1=xyz")
    assert cookies[0]["name"] == "web_session"
    assert cookies[0]["httpOnly"] is True
    assert cookies[0]["sameSite"] == "Lax"
    assert cookies[0]["secure"] is True
    assert cookies[1]["name"] == "a1"
    assert cookies[1]["httpOnly"] is False
    assert cookies[1]["sameSite"] == "None"

    only = _cookie_str_to_list("unknown_cookie=val")
    assert only[0]["httpOnly"] is False
    assert only[0]["sameSite"] == "Lax"
    assert only[0]["secure"] is True


class _FakeEl:
    def __init__(self, text):
        self._text = text

    def inner_text(self):
        return self._text


class _FakePageForReferer:
    def __init__(self):
        self.url = "https://www.xiaohongshu.com/explore"
        self.headers_calls = []
        self.goto_calls = []
        self.inpage_nav_urls = []

    def set_extra_http_headers(self, headers):
        self.headers_calls.append(dict(headers))

    def goto(self, url, **kwargs):
        self.goto_calls.append({"url": url, **kwargs})
        self.url = url

    def evaluate(self, script, *args):
        if "window.location.href = targetUrl" in script and args:
            self.inpage_nav_urls.append(args[0])
            self.url = args[0]
            return None
        if "document.querySelectorAll('a[href*=\"/explore/\"]')" in script:
            return False
        return ""

    def wait_for_selector(self, _selector, timeout=0):
        return None

    def wait_for_url(self, _matcher, timeout=0):
        return None

    def query_selector(self, _selector):
        return _FakeEl("正文内容")


class _FakePageForExtractDesc:
    def __init__(self):
        self.wait_calls = []
        self.query_calls = []

    def wait_for_selector(self, selector, timeout=0):
        self.wait_calls.append((selector, timeout))
        return None

    def query_selector(self, selector):
        self.query_calls.append(selector)
        if selector == ".note-content .desc":
            return _FakeEl("这是正文")
        return None

    def evaluate(self, _script):
        return ""


def test_extract_desc_uses_single_combined_wait_and_returns_text():
    fetcher = XhsFetcher("web_session=abc; a1=xyz")
    fake_page = _FakePageForExtractDesc()
    fetcher._page = fake_page

    desc = fetcher._extract_desc()

    assert desc == "这是正文"
    assert fake_page.wait_calls == [
        (".note-content .desc, #detail-desc, .desc.expand, #app [class*='desc']", 4000)
    ]
    assert fake_page.query_calls[0] == ".note-content .desc"


def test_fetch_note_detail_sets_and_clears_referer_header_with_profile_url():
    fetcher = XhsFetcher("web_session=abc; a1=xyz")
    fake_page = _FakePageForReferer()
    fetcher._page = fake_page
    fetcher._current_profile_url = "https://www.xiaohongshu.com/user/profile/abc123"
    fetcher._needs_captcha = lambda: False
    fetcher._extract_desc = lambda: "正文内容"

    note = {"note_id": "note1", "xsec_token": "token1", "title": "标题"}
    result = fetcher.fetch_note_detail(note)

    assert result["detail_fetched"] is True
    assert fake_page.headers_calls[0]["Referer"] == "https://www.xiaohongshu.com/user/profile/abc123"
    assert fake_page.headers_calls[-1] == {}
    assert fake_page.inpage_nav_urls
    assert "xsec_token=token1" in fake_page.inpage_nav_urls[0]


def test_fetch_note_detail_uses_explore_referer_when_profile_url_missing():
    fetcher = XhsFetcher("web_session=abc; a1=xyz")
    fake_page = _FakePageForReferer()
    fetcher._page = fake_page
    fetcher._current_profile_url = None
    fetcher._needs_captcha = lambda: False
    fetcher._extract_desc = lambda: "正文内容"

    note = {"note_id": "note2", "xsec_token": "token2", "title": "标题"}
    result = fetcher.fetch_note_detail(note)

    assert result["detail_fetched"] is True
    assert fake_page.headers_calls[0]["Referer"] == "https://www.xiaohongshu.com/explore"
    assert fake_page.headers_calls[-1] == {}
    assert fake_page.inpage_nav_urls
    assert "xsec_token=token2" in fake_page.inpage_nav_urls[0]


class _FakeTabForConcurrent:
    def __init__(self):
        self.url = "https://www.xiaohongshu.com/explore"
        self.wait_calls = []
        self.headers_calls = []

    def wait_for_timeout(self, ms):
        self.wait_calls.append(ms)
        time.sleep(0.005)

    def set_extra_http_headers(self, headers):
        self.headers_calls.append(dict(headers))

    def goto(self, url, **kwargs):
        self.url = url
        time.sleep(0.005)

    def evaluate(self, script, *args):
        if "window.location.href = targetUrl" in script and args:
            self.url = args[0]
            return None
        if "document.querySelectorAll('a[href*=\"/explore/\"]')" in script:
            return False
        return ""

    def wait_for_selector(self, _selector, timeout=0):
        return None

    def wait_for_url(self, _matcher, timeout=0):
        return None

    def query_selector(self, _selector):
        return _FakeEl("正文内容")

    def close(self):
        return None


class _FakeContextForConcurrent:
    def __init__(self, tabs):
        self._tabs = tabs
        self._idx = 0

    def new_page(self):
        tab = self._tabs[self._idx]
        self._idx += 1
        return tab


def test_fetch_note_details_concurrent_uses_staggered_tab_delay(monkeypatch):
    monkeypatch.setattr("random.uniform", lambda _a, _b: 2000)

    tabs = [_FakeTabForConcurrent(), _FakeTabForConcurrent()]
    fetcher = XhsFetcher("web_session=abc; a1=xyz")
    fetcher._context = _FakeContextForConcurrent(tabs)
    fetcher._current_profile_url = "https://www.xiaohongshu.com/user/profile/abc123"

    notes = [
        {"note_id": "n1", "xsec_token": "t1", "title": "T1"},
        {"note_id": "n2", "xsec_token": "t2", "title": "T2"},
    ]
    out = fetcher.fetch_note_details_concurrent(notes, max_tabs=2)

    assert len(out) == 2
    assert tabs[0].wait_calls, "tab0 应有 wait_for_timeout 调用"
    assert tabs[1].wait_calls, "tab1 应有 wait_for_timeout 调用"
    assert tabs[1].wait_calls[0] >= tabs[0].wait_calls[0] + 1500
    assert tabs[0].headers_calls[0]["Referer"] == "https://www.xiaohongshu.com/user/profile/abc123"
    assert tabs[1].headers_calls[0]["Referer"] == "https://www.xiaohongshu.com/user/profile/abc123"
    assert tabs[0].headers_calls[-1] == {}
    assert tabs[1].headers_calls[-1] == {}


class _FakeProfilePageNoSleep:
    def __init__(self):
        self.url = "https://www.xiaohongshu.com/explore"

    def on(self, _event, _cb):
        return None

    def remove_listener(self, _event, _cb):
        return None

    def goto(self, url, **kwargs):
        self.url = url
        return None

    def wait_for_load_state(self, *_args, **_kwargs):
        return None

    def query_selector(self, _selector):
        return None

    def query_selector_all(self, _selector):
        return []


class _FakeLink:
    def __init__(self, href):
        self._href = href

    def get_attribute(self, name):
        if name == "href":
            return self._href
        return ""


class _FakeDetailPageNoSleep:
    def __init__(self, note_id):
        self.note_id = note_id
        self.url = "https://www.xiaohongshu.com/user/profile/u1"
        self.wait_timeout_calls = []

    def on(self, _event, _cb):
        return None

    def remove_listener(self, _event, _cb):
        return None

    def set_extra_http_headers(self, _headers):
        return None

    def goto(self, url, **kwargs):
        self.url = url
        return None

    def wait_for_load_state(self, *_args, **_kwargs):
        return None

    def wait_for_timeout(self, ms):
        self.wait_timeout_calls.append(ms)
        return None

    def evaluate(self, script, *args):
        if "window.__pinia" in script:
            return None
        if "a.click(); return true;" in script:
            clicked_note_id = args[0]
            self.url = f"https://www.xiaohongshu.com/explore/{clicked_note_id}"
            return True
        if "document.body?.innerText" in script:
            return ""
        return ""

    def query_selector(self, selector):
        if selector.startswith("a[href*='/explore/"):
            return _FakeLink(f"/explore/{self.note_id}")
        return None


def test_fetch_profile_and_notes_does_not_call_sleep_one_second(monkeypatch):
    sleep_calls = []
    monkeypatch.setattr("time.sleep", lambda s: sleep_calls.append(s))

    fetcher = XhsFetcher("web_session=abc; a1=xyz")
    fetcher._page = _FakeProfilePageNoSleep()
    fetcher._needs_captcha = lambda: False

    fetcher.fetch_profile_and_notes("u1", limit=1)
    assert 1 not in sleep_calls


def test_fetch_note_detail_no_token_path_does_not_call_sleep_one_or_half(monkeypatch):
    sleep_calls = []
    monkeypatch.setattr("time.sleep", lambda s: sleep_calls.append(s))

    note_id = "n100"
    fetcher = XhsFetcher("web_session=abc; a1=xyz")
    fake_page = _FakeDetailPageNoSleep(note_id)
    fetcher._page = fake_page
    fetcher._current_profile_url = "https://www.xiaohongshu.com/user/profile/u1"
    fetcher._token_refetch_done = False
    fetcher._note_tokens = {}
    fetcher._needs_captcha = lambda: False
    fetcher._extract_desc = lambda: "正文内容"

    note = {"note_id": note_id, "xsec_token": "", "title": "标题"}
    out = fetcher.fetch_note_detail(note)

    assert out["detail_fetched"] is False
    assert out["detail_skip_reason"] == "no_token_click_failed"
    assert 1 not in sleep_calls
    assert 0.5 not in sleep_calls
    assert fake_page.wait_timeout_calls == []


def test_analyze_with_claude_skips_unsafe_cover_image_url(monkeypatch):
    captured = {}

    class _DummyResp:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return json.dumps({
                "choices": [{
                    "message": {
                        "content": "图文风格描述|简洁\n账号权重评级|A 稳定"
                    }
                }]
            }).encode("utf-8")

    def _fake_urlopen(req, timeout=90):
        captured["payload"] = json.loads(req.data.decode("utf-8"))
        return _DummyResp()

    monkeypatch.setattr("analyze_accounts._load_dashscope_key", lambda: ("test-key", "https://example.com/v1/chat/completions", "qwen-vl-max"))
    monkeypatch.setattr(urllib.request, "urlopen", _fake_urlopen)

    notes = [{
        "title": "标题",
        "desc": "正文",
        "cover_url": "http://127.0.0.1/x.jpg",
    }]
    analyze_with_claude(notes, "测试账号")

    content_blocks = captured["payload"]["messages"][0]["content"]
    image_blocks = [b for b in content_blocks if isinstance(b, dict) and b.get("type") == "image_url"]
    assert image_blocks == []


def test_analyze_with_claude_caps_cover_images_to_three(monkeypatch):
    captured = {}

    class _DummyResp:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return json.dumps({
                "choices": [{
                    "message": {
                        "content": "图文风格描述|简洁\n账号权重评级|A 稳定"
                    }
                }]
            }).encode("utf-8")

    def _fake_urlopen(req, timeout=90):
        captured["payload"] = json.loads(req.data.decode("utf-8"))
        return _DummyResp()

    monkeypatch.setattr("analyze_accounts._load_dashscope_key", lambda: ("test-key", "https://example.com/v1/chat/completions", "qwen-vl-max"))
    monkeypatch.setattr(urllib.request, "urlopen", _fake_urlopen)

    notes = [
        {"title": f"标题{i}", "desc": "正文", "cover_url": f"https://sns-img-qc.xhscdn.com/{i}.jpg"}
        for i in range(6)
    ]
    analyze_with_claude(notes, "测试账号")

    content_blocks = captured["payload"]["messages"][0]["content"]
    image_blocks = [b for b in content_blocks if isinstance(b, dict) and b.get("type") == "image_url"]
    assert len(image_blocks) <= 3


class DummyFetcher:
    def __init__(self, detailed_notes):
        self._token_refetch_done = False
        self._note_tokens = {"n1": "token"}
        self._detailed_notes = detailed_notes

    def search_user_id(self, nickname, xhs_id=""):
        return "user_1"

    def fetch_profile_and_notes(self, user_id, limit=6):
        profile = {"fans_count": 1000, "notes_count": len(self._detailed_notes)}
        notes = [{"note_id": "n1", "title": "标题", "xsec_token": "token"}]
        return profile, notes

    def fetch_note_details_concurrent(self, notes, max_tabs=2):
        return self._detailed_notes

    def fetch_note_detail(self, note):
        return note


def test_process_account_skips_notes_without_detail_success(monkeypatch):
    monkeypatch.setattr("analyze_accounts.filter_notes_by_recent_month", lambda notes, days=30: notes)
    monkeypatch.setattr("analyze_accounts.analyze_with_claude", lambda notes, nickname: pytest.fail("不应进入 Claude 分析"))

    fetcher = DummyFetcher([{"note_id": "n1", "title": "只有标题", "desc": "", "detail_fetched": False}])
    account = {"name": "测试账号", "status": "爆款账号"}

    row = process_account(fetcher, account)

    assert row["分析笔记篇数"] == 0
    assert row["账号权重评级"] == "低活跃度-无法获取详情"


def test_process_account_reuses_refetched_tokens_and_skips_still_missing(monkeypatch):
    class TokenPromotingFetcher:
        def __init__(self):
            self._token_refetch_done = False
            self._note_tokens = {}
            self.concurrent_calls = []
            self.detail_calls = []

        def search_user_id(self, nickname, xhs_id=""):
            return "user_1"

        def fetch_profile_and_notes(self, user_id, limit=6):
            profile = {"fans_count": 1000, "notes_count": 3}
            notes = [
                {"note_id": "n1", "title": "有token", "xsec_token": "t1", "image_count": 1, "liked_count": 1, "collected_count": 1, "comment_count": 1},
                {"note_id": "n2", "title": "后续补token", "xsec_token": "", "image_count": 1, "liked_count": 1, "collected_count": 1, "comment_count": 1},
                {"note_id": "n3", "title": "仍无token", "xsec_token": "", "image_count": 1, "liked_count": 1, "collected_count": 1, "comment_count": 1},
            ]
            return profile, notes

        def fetch_note_details_concurrent(self, notes, max_tabs=2):
            ids = [n.get("note_id") for n in notes]
            self.concurrent_calls.append(ids)
            # 第一轮并发后，模拟缓存中补到 n2 的 token
            if ids == ["n1"]:
                self._note_tokens["n2"] = "tok2"
            out = []
            for n in notes:
                enriched = dict(n)
                enriched["desc"] = f"正文-{n.get('note_id')}"
                enriched["word_count"] = len(enriched.get("title", "") + enriched["desc"])
                enriched["detail_fetched"] = True
                out.append(enriched)
            return out

        def fetch_note_detail(self, note):
            self.detail_calls.append(note.get("note_id"))
            enriched = dict(note)
            enriched["desc"] = "fallback"
            enriched["word_count"] = len(enriched.get("title", "") + enriched["desc"])
            enriched["detail_fetched"] = True
            return enriched

    monkeypatch.setattr("analyze_accounts.filter_notes_by_recent_month", lambda notes, days=30: notes)
    monkeypatch.setattr("analyze_accounts.filter_high_performance_notes", lambda notes: notes)
    monkeypatch.setattr("analyze_accounts.save_style_model", lambda account_name, style_model: "")
    monkeypatch.setattr("analyze_accounts.analyze_with_claude", lambda notes, nickname: {
        "图文风格描述": "稳定",
        "内容偏好": "求职",
        "受众分析": "留学生",
        "目标对象关联点": "投递痛点",
        "账号权重评级": "A 稳定",
        "Hook模板": "h1 | h2",
        "内容结构": "s1 | s2",
        "CTA模板": "c1 | c2",
    })

    fetcher = TokenPromotingFetcher()
    account = {"name": "测试账号", "status": "爆款账号"}
    row = process_account(fetcher, account)

    assert fetcher.concurrent_calls == [["n1"], ["n2"]]
    assert fetcher.detail_calls == []
    assert row["分析笔记篇数"] == 2
    assert row["账号权重评级"] == "A 稳定"


def test_process_account_calls_fetch_profile_even_with_prefetched_notes(monkeypatch):
    import analyze_accounts as aa

    fetch_called = {"called": False}

    class FakeFetcher:
        _note_tokens = {}
        _token_refetch_done = False

        def search_user_id(self, nickname, xhs_id=""):
            return "uid123"

        def fetch_profile_and_notes(self, user_id, limit=6):
            fetch_called["called"] = True
            return {"fans_count": 5000, "notes_count": 20}, []

        def fetch_note_details_concurrent(self, notes, max_tabs=2):
            return []

    explore_data = {
        "prefetched_notes": [
            {"title": "标题1", "content": "正文内容一", "likes": 100, "saves": 50},
        ]
    }

    row = aa.process_account(FakeFetcher(), {"name": "测试号", "status": "爆款账号"}, explore_data=explore_data)

    assert fetch_called["called"] is True
    assert row["粉丝数"] == 5000
    assert row["笔记总数"] == 20


def test_process_account_uses_full_detail_when_available(monkeypatch):
    import analyze_accounts as aa

    class FakeFetcher:
        _note_tokens = {}
        _token_refetch_done = False

        def search_user_id(self, nickname, xhs_id=""):
            return "uid456"

        def fetch_profile_and_notes(self, user_id, limit=6):
            return {"fans_count": 8000, "notes_count": 30}, [
                {"note_id": "n1", "xsec_token": "tok1", "create_time": 9999999999}
            ]

        def fetch_note_details_concurrent(self, notes, max_tabs=2):
            return [{
                "note_id": "n1", "title": "标题", "desc": "正文",
                "word_count": 100, "image_count": 3,
                "liked_count": 200, "collected_count": 80, "comment_count": 15,
                "detail_fetched": True, "cover_url": "",
            }]

    monkeypatch.setattr(aa, "analyze_with_claude", lambda notes, nick: {
        "图文风格描述": "ok", "内容偏好": "ok", "受众分析": "ok",
        "目标对象关联点": "ok", "账号权重评级": "A", "Hook模板": "ok",
        "内容结构": "ok", "CTA模板": "ok",
    })
    monkeypatch.setattr(aa, "save_style_model", lambda *a: None)
    monkeypatch.setattr(aa, "build_style_model", lambda notes: {"title_templates": ["t"], "tone": "积极", "image_style": "单图"})
    monkeypatch.setattr(aa, "filter_notes_by_recent_month", lambda notes, days=30: notes)
    monkeypatch.setattr(aa, "filter_high_performance_notes", lambda notes: notes)

    row = aa.process_account(
        FakeFetcher(),
        {"name": "测试号", "status": "爆款账号"},
        explore_data={"prefetched_notes": [{"title": "标题", "content": "正文", "likes": 999, "saves": 999}]},
    )

    assert row["粉丝数"] == 8000
    assert row["平均图片张数"] != "N/A"
    assert row["平均评论数"] != "N/A"
    assert row["互动率"] != "N/A"


def test_process_account_falls_back_to_prefetched_with_na_fields(monkeypatch):
    import analyze_accounts as aa

    class FakeFetcher:
        _note_tokens = {}
        _token_refetch_done = False

        def search_user_id(self, nickname, xhs_id=""):
            return "uid789"

        def fetch_profile_and_notes(self, user_id, limit=6):
            return {"fans_count": 3000, "notes_count": 10}, [
                {"note_id": "n1", "xsec_token": "tok1", "create_time": 9999999999}
            ]

        def fetch_note_details_concurrent(self, notes, max_tabs=2):
            return []

    monkeypatch.setattr(aa, "analyze_with_claude", lambda notes, nick: {
        "图文风格描述": "ok", "内容偏好": "ok", "受众分析": "ok",
        "目标对象关联点": "ok", "账号权重评级": "B", "Hook模板": "ok",
        "内容结构": "ok", "CTA模板": "ok",
    })
    monkeypatch.setattr(aa, "save_style_model", lambda *a: None)
    monkeypatch.setattr(aa, "build_style_model", lambda notes: {"title_templates": ["t"], "tone": "积极", "image_style": "单图"})
    monkeypatch.setattr(aa, "filter_notes_by_recent_month", lambda notes, days=30: notes)
    monkeypatch.setattr(aa, "filter_high_performance_notes", lambda notes: notes)

    row = aa.process_account(
        FakeFetcher(),
        {"name": "测试号", "status": "爆款账号"},
        explore_data={
            "prefetched_notes": [
                {"title": "兜底标题", "content": "兜底正文内容", "likes": 50, "saves": 20},
            ]
        },
    )

    assert row["粉丝数"] == 3000
    assert row["平均图片张数"] == "N/A"
    assert row["平均评论数"] == "N/A"
    assert row["互动率"] != "N/A"
    assert row["账号权重评级"] == "B"


def test_main_without_target_count_keeps_incomplete_rows(monkeypatch, tmp_path):
    import analyze_accounts

    captured = {}
    saved_style_accounts = []

    def fake_process_account(fetcher, account, explore_data=None):
        row = {h: "N/A" for h in HEADERS}
        row["账号昵称"] = account["name"]
        row["状态"] = account["status"]
        row["账号权重评级"] = "分析失败" if account["name"] == "bad" else "A 稳定"
        return row

    def fake_write_excel(rows, output_path):
        captured["rows"] = rows
        captured["output_path"] = output_path

    monkeypatch.setattr(analyze_accounts, "_load_cookie", lambda *args: ("cookie", "__env__"))
    monkeypatch.setattr(analyze_accounts, "process_account", fake_process_account)
    monkeypatch.setattr(analyze_accounts, "write_excel", fake_write_excel)
    monkeypatch.setattr(analyze_accounts, "get_analysis_workbook_path", lambda: tmp_path / "爆款分析" / "xhs_account_analysis.xlsx")
    monkeypatch.setattr(analyze_accounts, "load_existing_excel", lambda path: [{"账号昵称": "old", "账号权重评级": "A 稳定"}])
    def fake_save_account_style(account_name, style_params, source="analysis"):
        saved_style_accounts.append(account_name)

    monkeypatch.setattr(analyze_accounts.publish_manager, "save_account_style", fake_save_account_style)
    monkeypatch.setattr(analyze_accounts.random, "uniform", lambda a, b: 0)
    monkeypatch.setattr(analyze_accounts.time, "sleep", lambda s: None)

    monkeypatch.setattr(analyze_accounts.XhsFetcher, "start", lambda self: None)
    monkeypatch.setattr(analyze_accounts.XhsFetcher, "stop", lambda self: None)

    workbook_path = tmp_path / "爆款分析" / "xhs_account_analysis.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path.write_text("stub", encoding="utf-8")

    accounts_file = tmp_path / "accounts.json"
    accounts_file.write_text(
        json.dumps([
            {"name": "good", "status": "爆款账号"},
            {"name": "bad", "status": "爆款账号"},
        ], ensure_ascii=False),
        encoding="utf-8",
    )

    monkeypatch.setattr(sys, "argv", [
        "analyze_accounts.py",
        "--accounts-file", str(accounts_file),
    ])

    analyze_accounts.main()

    assert [row["账号昵称"] for row in captured["rows"]] == ["old", "bad", "good"]
    bad_row = next(r for r in captured["rows"] if r.get("账号昵称") == "bad")
    assert bad_row["账号权重评级"] == "分析失败"
    assert captured["output_path"].endswith("爆款分析/xhs_account_analysis.xlsx")
    assert saved_style_accounts == ["good"]


def test_main_skips_duplicate_accounts_already_in_cumulative_workbook(monkeypatch, tmp_path):
    import analyze_accounts

    captured = {}

    def fake_process_account(fetcher, account, explore_data=None):
        row = {h: "N/A" for h in HEADERS}
        row["账号昵称"] = account["name"]
        row["状态"] = account["status"]
        row["账号权重评级"] = "A 稳定"
        return row

    def fake_write_excel(rows, output_path):
        captured["rows"] = rows
        captured["output_path"] = output_path

    monkeypatch.setattr(analyze_accounts, "_load_cookie", lambda *args: ("cookie", "__env__"))
    monkeypatch.setattr(analyze_accounts, "process_account", fake_process_account)
    monkeypatch.setattr(analyze_accounts, "write_excel", fake_write_excel)
    monkeypatch.setattr(analyze_accounts, "get_analysis_workbook_path", lambda: tmp_path / "爆款分析" / "xhs_account_analysis.xlsx")
    monkeypatch.setattr(analyze_accounts, "load_existing_excel", lambda path: [{"账号昵称": "good", "账号权重评级": "A 稳定"}])
    monkeypatch.setattr(analyze_accounts.publish_manager, "save_account_style", lambda *args, **kwargs: None)
    monkeypatch.setattr(analyze_accounts.random, "uniform", lambda a, b: 0)
    monkeypatch.setattr(analyze_accounts.time, "sleep", lambda s: None)
    monkeypatch.setattr(analyze_accounts.XhsFetcher, "start", lambda self: None)
    monkeypatch.setattr(analyze_accounts.XhsFetcher, "stop", lambda self: None)

    workbook_path = tmp_path / "爆款分析" / "xhs_account_analysis.xlsx"
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook_path.write_text("stub", encoding="utf-8")

    accounts_file = tmp_path / "accounts.json"
    accounts_file.write_text(
        json.dumps([{"name": "good", "status": "爆款账号"}], ensure_ascii=False),
        encoding="utf-8",
    )

    monkeypatch.setattr(sys, "argv", [
        "analyze_accounts.py",
        "--accounts-file", str(accounts_file),
    ])

    analyze_accounts.main()

    assert captured["rows"] == [{"账号昵称": "good", "账号权重评级": "A 稳定"}]
    assert captured["output_path"].endswith("爆款分析/xhs_account_analysis.xlsx")
