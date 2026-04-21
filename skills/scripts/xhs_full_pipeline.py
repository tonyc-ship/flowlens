#!/usr/bin/env python3
"""
小红书全自动 Pipeline

流程：
  1. xhs_explore  — 关键词抓取爆款账号
  2. analyze_accounts — 分析 5 个账号
  3. generate_content — 每账号生成 2 篇文案
  4. info-card-generator — 每篇文案生成图文卡片（5张/篇）
  5. 结果汇总写入 output/小红书备选/

用法：
  python3.13 scripts/xhs_full_pipeline.py --keyword "英国留学求职"
  python3.13 scripts/xhs_full_pipeline.py --keyword "英国留学求职" --accounts 5 --copies 2
"""

import argparse
import json
import os
import re
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from flowlens_runtime import get_flowlens_python
from flowlens_xhs_backend import discover_accounts_via_flowlens

BASE_DIR = Path(__file__).parent.parent
SCRIPTS_DIR = BASE_DIR / "scripts"
EXPLORE_DIR = BASE_DIR / "mcp" / "explore"
CARD_TOOL_DIR = BASE_DIR / "tools" / "info-card-generator"
OUTPUT_DIR = BASE_DIR / "output"
ANALYSIS_DIR = OUTPUT_DIR / "爆款分析"
ANALYSIS_XLSX = ANALYSIS_DIR / "xhs_account_analysis.xlsx"
DRAFT_DIR = OUTPUT_DIR / "小红书备选"


def log(msg: str, level: str = "info"):
    icons = {"info": "ℹ️ ", "ok": "✅", "warn": "⚠️ ", "err": "❌", "step": "🔷"}
    print(f"{icons.get(level, '')} {msg}", flush=True)


# ──────────────────────────────────────────────
# Step 1: Explore
# ──────────────────────────────────────────────

def run_explore(keyword: str, author_limit: int = 20) -> list[dict]:
    """Use FlowLens-backed retrieval to discover candidate XHS authors."""
    log(f"Step 1 — xhs_explore: 关键词「{keyword}」", "step")

    try:
        data = discover_accounts_via_flowlens(
            keyword,
            search_limit=30,
            viral_threshold=30,
            author_limit=author_limit,
        )
    except Exception as exc:
        log(f"FlowLens explore 失败: {exc}", "err")
        sys.exit(1)

    authors = data.get("authors", [])
    log(f"获取 {len(authors)} 个候选账号", "ok")
    return authors


# ──────────────────────────────────────────────
# Step 2: Analyze accounts
# ──────────────────────────────────────────────

def _select_recent_analysis_rows(rows: list[dict], limit: int = 0) -> list[dict]:
    """从累计分析结果里取最近追加的若干条。"""
    if limit <= 0 or len(rows) <= limit:
        return rows
    return rows[-limit:]


def run_analyze(authors: list[dict], keyword: str, target_count: int = 5) -> tuple[Path, list[dict]]:
    """调用 analyze_accounts.py，返回生成的 xlsx 路径和本次新增的分析行。"""
    log(f"Step 2 — analyze_accounts: 目标 {target_count} 个有效账号", "step")

    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)

    # 去重：跳过累计分析表中已分析过的账号
    import openpyxl as _openpyxl
    already_analyzed: set[str] = set()
    existing_rows: list[dict] = []
    if ANALYSIS_XLSX.exists():
        try:
            existing_rows = load_analysis_xlsx(ANALYSIS_XLSX)
            wb0 = _openpyxl.load_workbook(ANALYSIS_XLSX)
            ws0 = wb0.active
            hdrs0 = [c.value for c in ws0[1]]
            if "账号昵称" in hdrs0:
                idx = hdrs0.index("账号昵称")
                for r in ws0.iter_rows(min_row=2, values_only=True):
                    if r[idx]:
                        already_analyzed.add(str(r[idx]))
            log(f"去重：累计分析表中已有 {len(already_analyzed)} 个账号，跳过", "info")
        except Exception:
            pass

    # 传入全部候选，由 --target-count 决定有效数量
    accounts = []
    for a in authors:
        name = a.get("name", "")
        if name in already_analyzed:
            continue
        accounts.append({
            "name": name,
            "xhs_id": a.get("xhs_id"),
            "author_hex_id": a.get("author_hex_id"),
            "profile_url": a.get("profile_url", ""),
            "status": "爆款账号",
            "explore_data": a.get("explore_data", {}),
        })
    log(f"候选账号（去重后）：{len(accounts)} 个", "info")

    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    ) as f:
        json.dump(accounts, f, ensure_ascii=False, indent=2)
        accounts_file = f.name

    try:
        result = subprocess.run(
            [
                get_flowlens_python(),
                str(SCRIPTS_DIR / "analyze_accounts.py"),
                "--accounts-file", accounts_file,
                "--target-count", str(target_count),
            ],
            cwd=str(BASE_DIR),
            text=True,
            timeout=1800,  # 30分钟
        )
        if result.returncode != 0:
            log("analyze_accounts 返回非零退出码", "warn")
    finally:
        os.unlink(accounts_file)

    if not ANALYSIS_XLSX.exists():
        log("未找到分析结果 xlsx", "err")
        sys.exit(1)
    latest = ANALYSIS_XLSX
    all_rows = load_analysis_xlsx(latest)
    appended_rows = all_rows[len(existing_rows):] if len(all_rows) >= len(existing_rows) else all_rows

    log(f"分析完成: {latest.name}", "ok")
    return latest, appended_rows


# ──────────────────────────────────────────────
# Step 3: 读取分析结果
# ──────────────────────────────────────────────

def load_analysis_xlsx(xlsx_path: Path) -> list[dict]:
    """读取 xlsx，返回账号数据列表"""
    try:
        import openpyxl
    except ImportError:
        log("缺少 openpyxl，安装中...", "warn")
        subprocess.run(["pip3", "install", "openpyxl", "--break-system-packages"], check=True)
        import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("账号昵称"):
            rows.append(d)
    return rows


# ──────────────────────────────────────────────
# Step 4: 生成文案
# ──────────────────────────────────────────────

COPY_TEMPLATE = """# {title}

**账号参考**：{name}
**内容结构**：{structure}
**Hook**：{hook}
**情绪**：{emotion}
**图片风格**：{img_style}

---

{hook_text}

{body}

{cta_text}

{hashtags}
"""

CONTENT_TOPICS = {
    "面试": ["面试3大致命错误", "offer拿到手前你要做这几件事"],
    "简历": ["0经验简历这样写不被秒拒", "CV到底差在哪里学姐帮你看"],
    "工签": ["PSW转工签最全时间线", "sponsor公司怎么找一文说清"],
    "留学": ["英国留学生求职最容易踩的坑", "读研期间这样积累经验上岸不难"],
    "默认": ["留英求职3个月我学到了什么", "刷到就是缘分这些干货别错过"],
}


def guess_topics(account: dict) -> list[str]:
    """根据账号内容偏好猜测选题"""
    pref = str(account.get("内容偏好", "") or "")
    for key, topics in CONTENT_TOPICS.items():
        if key in pref:
            return topics
    return CONTENT_TOPICS["默认"]


def generate_copy(account: dict, topic: str, copy_index: int) -> str:
    """为单个账号生成一篇文案"""
    name = account.get("账号昵称", "未知账号")
    structure = account.get("内容结构", "痛点引入 | 解决方案 | 总结")
    hooks_raw = account.get("Hook模板", "") or ""
    hooks = [h.strip() for h in re.split(r"[|｜]", hooks_raw) if h.strip()]
    hook = hooks[copy_index % len(hooks)] if hooks else "刷到就是缘分！"

    ctas_raw = account.get("CTA模板", "") or ""
    ctas = [c.strip() for c in re.split(r"[|｜]", ctas_raw) if c.strip()]
    cta = ctas[copy_index % len(ctas)] if ctas else "收藏备用 💌"

    emotion = account.get("情绪标签", "口语化")
    img_style = account.get("图片风格", "单图大字信息流")

    # 根据结构环节生成正文段落
    sections = [s.strip() for s in re.split(r"[|｜]", structure) if s.strip()]
    body_parts = []
    for s in sections[:4]:
        body_parts.append(f"**{s}**\n围绕「{topic}」的{s}展开，给读者提炼最可执行的一个信息点。")
    body = "\n\n".join(body_parts)

    # 标签
    keyword_tags = ["#英国留学", "#英国求职", "#留英上岸", f"#{topic.replace(' ', '')}"]
    hashtags = " ".join(keyword_tags)

    title_raw = account.get("标题模板", "") or ""
    titles = [t.strip() for t in re.split(r"[|｜]", title_raw) if t.strip()]
    title = titles[copy_index % len(titles)] if titles else f"关于{topic}，学姐来说真话"

    return COPY_TEMPLATE.format(
        title=title,
        name=name,
        structure=structure,
        hook=hook,
        emotion=emotion,
        img_style=img_style,
        hook_text=hook,
        body=body,
        cta_text=cta,
        hashtags=hashtags,
    )


def run_generate_copies(accounts: list[dict], copies_per_account: int = 2) -> list[Path]:
    """为所有账号生成文案，保存到 小红书备选/，返回文件路径列表"""
    log(f"Step 3 — 生成文案: {len(accounts)} 账号 × {copies_per_account} 篇", "step")

    DRAFT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    draft_paths = []

    for account in accounts:
        name = account.get("账号昵称", "未知账号")
        topics = guess_topics(account)

        for i in range(copies_per_account):
            topic = topics[i % len(topics)]
            copy_text = generate_copy(account, topic, i)

            safe_name = re.sub(r'[\\/:*?"<>|]', "_", name)
            filename = DRAFT_DIR / f"{safe_name}_文案{i+1}_{ts}.md"
            filename.write_text(copy_text, encoding="utf-8")
            draft_paths.append(filename)
            log(f"  [{name}] 篇{i+1} → {filename.name}", "ok")

    log(f"文案生成完成，共 {len(draft_paths)} 篇", "ok")
    return draft_paths


# ──────────────────────────────────────────────
# Step 5: 生成图文卡片
# ──────────────────────────────────────────────

CARD_STYLES = {
    "干货": ("notion", "list"),
    "面试": ("bold", "balanced"),
    "简历": ("study-notes", "dense"),
    "工签": ("minimal", "balanced"),
    "情感": ("warm", "balanced"),
    "默认": ("fresh", "balanced"),
}


def pick_style(account: dict) -> tuple[str, str]:
    pref = str(account.get("内容偏好", "") or "")
    for key, (style, layout) in CARD_STYLES.items():
        if key in pref:
            return style, layout
    return CARD_STYLES["默认"]


def generate_card_html(draft_path: Path, account: dict, card_out_dir: Path) -> list[Path]:
    """
    用 info-card-generator 为一篇文案生成 5 张 PNG 卡片。
    返回生成的 PNG 路径列表。
    """
    card_out_dir.mkdir(parents=True, exist_ok=True)
    style, layout = pick_style(account)

    # 读取文案内容
    content = draft_path.read_text(encoding="utf-8")

    # 解析标题、Hook、正文环节、CTA
    title_m = re.search(r'^# (.+)$', content, re.MULTILINE)
    hook_m = re.search(r'\*\*Hook\*\*：(.+)', content)
    cta_m = re.search(r'\*\*CTA\*\*：(.+)', content) or re.search(r'\n([^\n]{5,40}[💌📌])\n', content)

    title = title_m.group(1) if title_m else draft_path.stem
    hook = hook_m.group(1).strip() if hook_m else "刷到就是缘分！"
    cta = cta_m.group(1).strip() if cta_m else "收藏备用 💌"

    # 解析正文环节（**XXX**\n内容）
    sections = re.findall(r'\*\*(.+?)\*\*\n(.+?)(?=\n\n|\Z)', content, re.DOTALL)
    content_sections = [(h, b.strip()) for h, b in sections if h not in ("账号参考", "内容结构", "Hook", "情绪", "图片风格")]

    account_name = account.get("账号昵称", "")
    slug = re.sub(r'[^\w\-]', '-', title)[:30]

    png_paths = []

    # ── 卡片定义 ──────────────────────────────
    cards = []

    # P1: 封面
    cards.append({
        "name": "01-cover",
        "bg_color": "#fff8f2",
        "accent": "#ff6b35",
        "content_html": f"""
          <div class="cover-emoji">😤</div>
          <h1 class="cover-title">{title}</h1>
          <div class="cover-hook">{hook}</div>
        """,
    })

    # P2-P4: 正文环节（最多3个）
    for idx, (sec_title, sec_body) in enumerate(content_sections[:3], 2):
        cards.append({
            "name": f"0{idx}-content",
            "bg_color": "#fdfaf6",
            "accent": "#ff6b35",
            "content_html": f"""
              <div class="sec-num">0{idx-1}</div>
              <h2 class="sec-title">{sec_title}</h2>
              <p class="sec-body">{sec_body[:120]}...</p>
            """,
        })

    # P5: CTA
    cards.append({
        "name": "05-cta",
        "bg_color": "#fff0e8",
        "accent": "#ff6b35",
        "content_html": f"""
          <div class="cta-emoji">💌</div>
          <p class="cta-text">{cta}</p>
          <p class="cta-brand">{account_name}</p>
        """,
    })

    # 生成每张卡片 HTML → PNG
    capture_js = CARD_TOOL_DIR / "scripts" / "capture_card.js"

    for card in cards:
        html_path = card_out_dir / f"{card['name']}-{slug}.html"
        png_path = card_out_dir / f"{card['name']}-{slug}.png"

        progress_dots = "".join(
            f'<div class="dot {"active" if i == cards.index(card) else ""}"></div>'
            for i in range(len(cards))
        )

        html = f"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@400;700;900&family=Caveat:wght@600;700&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#ede9e0; display:flex; justify-content:center; padding:40px 20px; }}
.card-container {{
  width:900px; height:1600px;
  background:{card['bg_color']};
  border-radius:28px; overflow:hidden;
  font-family:'Noto Sans SC',sans-serif;
  box-shadow:0 12px 60px rgba(0,0,0,0.07);
  display:flex; flex-direction:column;
  align-items:center; justify-content:center;
  padding:72px 64px; text-align:center;
  position:relative;
}}
.brand {{ position:absolute; top:36px; right:48px; font-family:'Caveat',cursive; font-size:18px; color:#c8a888; }}
.progress {{ position:absolute; bottom:44px; left:50%; transform:translateX(-50%); display:flex; gap:10px; }}
.dot {{ width:10px;height:10px;border-radius:50%;background:#f0d0b0; }}
.dot.active {{ width:28px;border-radius:6px;background:{card['accent']}; }}
/* Cover */
.cover-emoji {{ font-size:96px; margin-bottom:36px; }}
.cover-title {{ font-size:48px;font-weight:900;color:#2a1a08;line-height:1.2;margin-bottom:20px; }}
.cover-hook {{ display:inline-block;background:{card['accent']};color:#fff;font-size:22px;font-weight:700;padding:10px 28px;border-radius:40px; }}
/* Content */
.sec-num {{ font-family:'Caveat',cursive;font-size:80px;font-weight:700;color:#f0d4c0;line-height:1;margin-bottom:20px; }}
.sec-title {{ font-size:40px;font-weight:900;color:#2a1a08;margin-bottom:24px; }}
.sec-body {{ font-size:20px;color:#7a5a40;line-height:1.8; }}
/* CTA */
.cta-emoji {{ font-size:80px;margin-bottom:28px; }}
.cta-text {{ font-size:26px;font-weight:700;color:#2a1a08;line-height:1.6;margin-bottom:16px; }}
.cta-brand {{ font-family:'Caveat',cursive;font-size:22px;color:#c8a888; }}
</style>
</head>
<body>
<div class="card-container">
  <span class="brand">{account_name}</span>
  {card['content_html']}
  <div class="progress">{progress_dots}</div>
</div>
</body>
</html>"""

        html_path.write_text(html, encoding="utf-8")

        # 截图
        r = subprocess.run(
            ["node", str(capture_js), str(html_path), str(png_path)],
            capture_output=True, text=True, timeout=30
        )
        html_path.unlink(missing_ok=True)  # 截图后删除 HTML

        if png_path.exists():
            png_paths.append(png_path)
            log(f"    {png_path.name}", "ok")
        else:
            log(f"    卡片生成失败: {card['name']}", "warn")

    return png_paths


def run_generate_cards(draft_paths: list[Path], accounts: list[dict]) -> dict:
    """为所有文案生成卡片，返回 {draft_path: [png_path, ...]}"""
    log(f"Step 4 — 生成图文卡片: {len(draft_paths)} 篇文案", "step")

    results = {}
    # 建立账号名 → account 的映射
    acc_map = {a.get("账号昵称", ""): a for a in accounts}

    for draft_path in draft_paths:
        # 从文件名解析账号名（格式：{账号名}_文案{N}_{ts}.md）
        stem = draft_path.stem  # e.g. "Helicopter_文案1_20260330_120000"
        acc_name_guess = re.sub(r'_文案\d+_.+$', '', stem)
        account = acc_map.get(acc_name_guess) or {}

        card_dir = DRAFT_DIR / stem
        log(f"  处理: {draft_path.name}", "info")
        pngs = generate_card_html(draft_path, account, card_dir)
        results[draft_path] = pngs

    log(f"图文卡片生成完成", "ok")
    return results


# ──────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="小红书全自动 Pipeline")
    parser.add_argument("--keyword", required=True, help="搜索关键词")
    parser.add_argument("--accounts", type=int, default=5, help="分析账号数（默认5）")
    parser.add_argument("--copies", type=int, default=2, help="每账号文案数（默认2）")
    parser.add_argument("--skip-explore", action="store_true", help="跳过 explore，使用最新分析 xlsx")
    parser.add_argument("--skip-cards", action="store_true", help="只生成文案，不生成卡片")
    parser.add_argument("--use-agent", action="store_true",
                        help="使用 FlowLens agent 完整调研（替代 explore+analyze 两步，推荐）")
    parser.add_argument("--notes", type=int, default=3,
                        help="每账号深度读取篇数，写入 FlowLens prompt（默认 3，仅 --use-agent 生效）")
    args = parser.parse_args()

    # ── FlowLens agent 模式 ──────────────────────────────────
    if args.use_agent:
        from flowlens_agent_pipeline import run_pipeline as run_agent_pipeline
        run_agent_pipeline(
            keyword=args.keyword,
            accounts=args.accounts,
            notes_per_account=args.notes,
            copies=args.copies,
            skip_generate=args.skip_cards,
        )
        return

    print(f"\n{'='*60}")
    print(f" 小红书全自动 Pipeline")
    print(f" 关键词: {args.keyword} | 账号: {args.accounts} | 文案: {args.copies}篇/账号")
    print(f"{'='*60}\n")

    # Step 1
    if args.skip_explore:
        log("跳过 explore，使用累计分析文件", "warn")
        if not ANALYSIS_XLSX.exists():
            log("未找到分析 xlsx，请先运行 explore", "err")
            sys.exit(1)
        xlsx_path = ANALYSIS_XLSX
        accounts = _select_recent_analysis_rows(load_analysis_xlsx(xlsx_path), args.accounts)
    else:
        authors = run_explore(args.keyword, author_limit=args.accounts * 4)
        xlsx_path, accounts = run_analyze(authors, args.keyword, target_count=args.accounts)

    # Step 2: 读分析结果
    log(f"读取 {len(accounts)} 个账号数据", "ok")

    # Step 3: 生成文案
    draft_paths = run_generate_copies(accounts, copies_per_account=args.copies)

    # Step 4: 生成卡片
    if not args.skip_cards:
        run_generate_cards(draft_paths, accounts)
    else:
        log("跳过图文卡片生成（--skip-cards）", "warn")

    # 汇总
    print(f"\n{'='*60}")
    print(f" ✅ Pipeline 完成！")
    print(f" 分析报告: {xlsx_path.name}")
    print(f" 文案目录: output/小红书备选/")
    print(f" 文案数量: {len(draft_paths)} 篇")
    if not args.skip_cards:
        print(f" 卡片目录: output/小红书备选/<文案名>/")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
