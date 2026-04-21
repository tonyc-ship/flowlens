#!/usr/bin/env python3
"""
小红书全自动运营 — 可视化控制台

流程：
  1. 输入关键词
  2. xhs_explore 抓取爆款账号
  3. analyze_accounts 深度分析 5 个账号
  4. 生成每账号 2 篇备选文案
  5. info-card-generator 生成图文卡片

访问: http://localhost:8888
"""

import html as _html
import json
import os
import re
import subprocess
import sys
import tempfile
import random
import time
import uuid
import webbrowser
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from threading import Thread
from typing import List, Optional
from urllib.parse import parse_qs, unquote, urlparse

sys.path.insert(0, str(Path(__file__).parent))
import account_manager
from flowlens_runtime import get_flowlens_python
import publish_manager
import persona_manager
publish_manager.start_scheduler()

# ── Session 自动保活 ───────────────────────────────────────────────────────────
def _start_keepalive_scheduler():
    """后台线程：storage.json 超过 2 天未刷新时自动保活，之后每 2 天再检查一次"""
    import threading, time as _time, json as _json

    KEEPALIVE_INTERVAL = 2 * 24 * 3600   # 2 天
    ACCOUNTS_PATH = Path(__file__).parent.parent / "accounts.json"

    def _needs_keepalive() -> bool:
        if not ACCOUNTS_PATH.exists():
            return False
        try:
            data = _json.loads(ACCOUNTS_PATH.read_text(encoding="utf-8"))
        except Exception:
            return False
        for acc in data.get("accounts", []):
            if not acc.get("active"):
                continue
            storage = Path(acc.get("storage_path", "")).expanduser()
            if storage.exists():
                age = _time.time() - storage.stat().st_mtime
                if age > KEEPALIVE_INTERVAL:
                    return True
        return False

    def _run():
        # 启动时立即检查一次，无需等待
        while True:
            try:
                if _needs_keepalive():
                    print("[keepalive] storage.json 超过 2 天，自动保活...", flush=True)
                    from xhs_keepalive import run_keepalive
                    results = run_keepalive(headless=True)
                    ok = sum(1 for r in results if r["ok"])
                    print(f"[keepalive] 完成 {ok}/{len(results)} 个账号", flush=True)
                else:
                    print("[keepalive] session 新鲜，跳过", flush=True)
            except Exception as e:
                print(f"[keepalive] 异常: {e}", flush=True)
            _time.sleep(KEEPALIVE_INTERVAL)   # 等 2 天再检查

    t = threading.Thread(target=_run, daemon=True, name="keepalive")
    t.start()

_start_keepalive_scheduler()


BASE_DIR = Path(__file__).parent.parent
SCRIPTS_DIR = BASE_DIR / "scripts"
EXPLORE_DIR = BASE_DIR / "mcp" / "explore"
CARD_TOOL_DIR = BASE_DIR / "tools" / "info-card-generator"
OUTPUT_DIR = BASE_DIR / "output"
ANALYSIS_DIR = OUTPUT_DIR / "爆款分析"
LOG_DIR = BASE_DIR / "output" / "logs"
ANALYSIS_XLSX = ANALYSIS_DIR / "xhs_account_analysis.xlsx"
DRAFT_DIR = OUTPUT_DIR / "小红书备选"

# ── 卡片生成任务池（card_task_id -> {status, cards, reply, error}）──────────────
_card_tasks: dict = {}
_login_tasks: dict = {}
# ── 文件级写锁，防止并发 _bg_gen 竞态 ──────────────────────────────────────────
import threading as _threading
_draft_file_locks: dict = {}
_draft_locks_mutex = _threading.Lock()

def _get_draft_lock(path: Path) -> _threading.Lock:
    key = str(path.resolve())
    with _draft_locks_mutex:
        if key not in _draft_file_locks:
            _draft_file_locks[key] = _threading.Lock()
        return _draft_file_locks[key]


# ── 全局状态 ───────────────────────────────────────────────────────────────────
state = {
    "status": "idle",       # idle | running | completed | error
    "keyword": "",
    "mode": "deep_report",
    "mode_label": "通用深度分析",
    "retrieval_account": "",
    "step": 0,              # 当前步骤 0-8
    "logs": [],
    "accounts": [],         # 分析结果
    "drafts": [],           # 生成的文案列表
    "cards": [],            # 生成的卡片路径
    "stats": {},
    "error": "",
}


_MODE_OPTIONS = {
    "deep_report": {
        "label": "通用深度分析",
        "summary": "完整拆解热门内容，兼顾结构、观点和可复用表达。",
        "instruction": """【本次生成模式】通用深度分析
- 目标：产出一篇信息完整、分析型更强的小红书笔记
- 写法：先快速建立问题背景，再总结核心观点、方法、注意事项和行动建议
- 输出重点：信息密度高，结构清楚，适合收藏
- 标题：偏“总结 / 拆解 / 攻略 / 方法论”风格
- 正文：尽量包含 3-5 个有价值的信息点，不要只写情绪感受""",
    },
    "topic_ideas": {
        "label": "选题研究",
        "summary": "更偏选题库和内容方向提炼，适合做爆款灵感池。",
        "instruction": """【本次生成模式】选题研究
- 目标：把素材转成强选题感的小红书内容
- 写法：突出人群标签、痛点、场景、反差和具体收益
- 输出重点：题目要有点击欲，正文要能快速看出切口
- 标题：尽量体现“什么人 + 什么问题 + 什么结果”
- 正文：适合沉淀为后续选题库，不必过度展开大段分析""",
    },
    "ops_breakdown": {
        "label": "运营拆解",
        "summary": "更强调 Hook、结构、CTA 和互动设计，适合运营复盘。",
        "instruction": """【本次生成模式】运营拆解
- 目标：产出一篇适合运营复盘和模仿的小红书笔记
- 写法：强化开头钩子、信息层次、情绪递进和结尾 CTA
- 输出重点：让读者容易被开头吸引，并愿意点赞收藏评论
- 标题：更强调冲突感、结果感、场景感
- 正文：要有明显的 Hook、主体信息块、结尾引导动作""",
    },
    "experience_summary": {
        "label": "经验总结",
        "summary": "更偏真实经历、决策过程和踩坑复盘，适合经验帖。",
        "instruction": """【本次生成模式】经验总结
- 目标：产出一篇真实经历导向的小红书经验帖
- 写法：强调个人处境、决策原因、过程变化、踩坑和结论
- 输出重点：让读者感受到“这是亲历总结”，不是空泛建议
- 标题：更适合“我怎么做 / 我踩过什么坑 / 我的经验总结”
- 正文：尽量有时间线、真实细节和可执行建议""",
    },
}


def _normalize_mode(mode: str) -> str:
    normalized = str(mode or "").strip()
    return normalized if normalized in _MODE_OPTIONS else "deep_report"


def _mode_label(mode: str) -> str:
    return _MODE_OPTIONS[_normalize_mode(mode)]["label"]


def _load_account_style(account: str) -> Optional[dict]:
    """读取 output/account_styles.json 中指定账号风格。不存在返回 None。"""
    return publish_manager.load_account_style(account)


def _normalize_draft_asset_url(raw_path: str, draft_path: Path) -> str:
    """将保存的图片路径归一化为前端可访问的 URL。"""
    raw = (raw_path or "").strip()
    if not raw:
        return ""
    if raw.startswith("/output/draft_cards/"):
        return raw
    try:
        candidate = Path(raw).expanduser()
        if not candidate.is_absolute():
            candidate = (draft_path.parent / candidate).resolve()
        else:
            candidate = candidate.resolve()
        draft_base = DRAFT_DIR.resolve()
        if candidate.is_relative_to(draft_base):
            rel = candidate.relative_to(draft_base).as_posix()
            return f"/output/draft_cards/{rel}"
    except Exception:
        return raw
    return raw


def _get_saved_images_for_draft(draft_path: Path) -> list[str]:
    """返回指定文案最后确认保存的图片 URL 列表。"""
    pub_path = draft_path.with_suffix(".pub.json")
    if pub_path.exists():
        try:
            pub = json.loads(pub_path.read_text(encoding="utf-8"))
            raw_paths = pub.get("image_paths", [])
            if not raw_paths and pub.get("image_path"):
                raw_paths = [pub.get("image_path")]
            urls = []
            seen = set()
            for raw_path in raw_paths:
                image_url = _normalize_draft_asset_url(str(raw_path), draft_path)
                if image_url and image_url not in seen:
                    urls.append(image_url)
                    seen.add(image_url)
            return urls
        except Exception:
            pass
    return []


def _build_draft_summary(draft_path: Path) -> dict:
    """构建前端使用的文案摘要，包含已生成图片信息。"""
    try:
        text = draft_path.read_text(encoding="utf-8")
        tm = re.search(r'^#\s+(.+)$', text, re.MULTILINE)
        title = tm.group(1).strip() if tm else draft_path.stem
    except Exception:
        title = draft_path.stem
    saved_images = _get_saved_images_for_draft(draft_path)
    return {
        "name": draft_path.name,
        "path": str(draft_path),
        "title": title,
        "cards": saved_images,
        "card_count": len(saved_images),
        "preview_image": saved_images[0] if saved_images else "",
    }


def _extract_draft_meta(content: str) -> dict:
    """从 markdown 元数据块提取账号、风格、主题等字段。"""
    patterns = {
        "account": r"\*\*账号参考\*\*[：:][ \t]*(.+)",
        "topic": r"\*\*主题\*\*[：:][ \t]*(.+)",
        "image_style": r"\*\*图片风格\*\*[：:][ \t]*(.+)",
        "writing_style": r"\*\*图文风格描述\*\*[：:][ \t]*(.+)",
    }
    meta = {}
    for key, pattern in patterns.items():
        m = re.search(pattern, content)
        meta[key] = m.group(1).strip() if m else ""
    return meta


def _strip_card_hashtags(text: str) -> str:
    """卡片渲染前移除话题标签，避免 #标签 进入图片内容。"""
    if not text:
        return ""

    hashtag_re = re.compile(r"(?<![A-Za-z0-9_])#[\w\u4e00-\u9fff-]+")
    cleaned_lines = []
    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        stripped = line.strip()
        if not stripped:
            cleaned_lines.append("")
            continue

        without_tags = hashtag_re.sub("", line)
        without_tags = re.sub(r"\s{2,}", " ", without_tags).strip()
        cleaned_lines.append(without_tags)

    return re.sub(r"\n{3,}", "\n\n", "\n".join(cleaned_lines)).strip()


def setup_file_logging() -> Path:
    """创建带时间戳的日志文件，同时接管 stdout/stderr，所有 print() 自动落盘。"""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    import glob as _glob, time as _time
    for old in _glob.glob(str(LOG_DIR / "server_*.log")):
        if _time.time() - os.path.getmtime(old) > 7 * 86400:
            try:
                os.unlink(old)
            except Exception:
                pass

    log_path = LOG_DIR / f"server_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    class _Tee:
        def __init__(self, *streams):
            self.streams = streams

        def write(self, data):
            for s in self.streams:
                try:
                    s.write(data)
                    s.flush()
                except Exception:
                    pass

        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

        def fileno(self):
            # 注意：此方法返回原始 stdout 的 fd，但 subprocess.run(..., stdout=sys.stdout)
            # 等直接使用 fd 的调用会绕过 _Tee.write()，导致输出不落日志文件。
            # 当前代码库所有子进程均用 capture_output=True，不受影响。
            # 未来若需传递 sys.stdout 给子进程，请改用 log_file 的 fd 或 io.UnsupportedOperation。
            import io
            raise io.UnsupportedOperation(
                "_Tee 不支持 fileno()：子进程请用 capture_output=True，手动 print() 转发输出"
            )

    log_file = open(log_path, "a", encoding="utf-8", buffering=1)
    sys.stdout = _Tee(sys.__stdout__, log_file)
    sys.stderr = _Tee(sys.__stderr__, log_file)
    print(f"[LOG] 日志文件：{log_path}", flush=True)
    return log_path


def add_log(msg: str, level: str = "info"):
    state["logs"].append({
        "time": datetime.now().strftime("%H:%M:%S"),
        "msg": msg,
        "level": level,
    })
    print(f"  [{level.upper()}] {msg}", flush=True)


def _list_managed_accounts() -> list:
    """返回账号管理视图所需的账号明细。"""
    return account_manager.list_accounts_detailed(include_inactive=True)


def _pick_account_task_message(lines: list[str], fallback: str = "登录完成") -> str:
    """从脚本输出中挑选最适合作为登录结果提示的那一行。"""
    if not lines:
        return fallback

    def _is_noise(line: str) -> bool:
        stripped = line.strip()
        if not stripped:
            return True
        compact = re.sub(r"\s+", "", stripped)
        if compact and all(ch in "=-_*~:|+." for ch in compact):
            return True
        return False

    preferred = [
        line for line in lines
        if not _is_noise(line) and "已识别账号" in line
    ]
    if preferred:
        return preferred[-1]

    preferred = [
        line for line in lines
        if not _is_noise(line) and any(token in line for token in ("登录成功", "session", "有效期", "已保存", "已更新"))
    ]
    if preferred:
        return preferred[-1]

    meaningful = [line for line in lines if not _is_noise(line)]
    return meaningful[-1] if meaningful else fallback


def _start_account_login_task(name: str, mode: str = "scan") -> dict:
    """启动账号扫码登录/刷新任务，返回 task 信息。"""
    account_name = (name or "").strip()
    if mode == "refresh" and not account_name:
        raise ValueError("账号名不能为空")
    if mode not in {"scan", "refresh"}:
        raise ValueError("mode 必须为 scan 或 refresh")

    task_id = uuid.uuid4().hex
    task = {
        "id": task_id,
        "name": account_name or "新账号",
        "mode": mode,
        "status": "running",
        "message": "浏览器即将打开，请在 120 秒内完成扫码登录",
        "started_at": datetime.now().isoformat(timespec="seconds"),
    }
    _login_tasks[task_id] = task

    def _run():
        cmd = [
            sys.executable,
            str(SCRIPTS_DIR / "account_manager.py"),
            mode,
        ]
        if account_name:
            cmd.extend(["--name", account_name])
        try:
            proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(BASE_DIR))
            logs = [line.strip() for line in (proc.stdout or "").splitlines() if line.strip()]
            err_logs = [line.strip() for line in (proc.stderr or "").splitlines() if line.strip()]
            if proc.stdout and proc.stdout.strip():
                print(f"[subprocess:{mode}:stdout]\n" + proc.stdout.rstrip(), flush=True)
            if proc.stderr and proc.stderr.strip():
                print(f"[subprocess:{mode}:stderr]\n" + proc.stderr.rstrip(), flush=True)
            if proc.returncode != 0:
                detail = err_logs[-1] if err_logs else (logs[-1] if logs else f"任务退出码 {proc.returncode}")
                raise RuntimeError(detail)
            done_msg = _pick_account_task_message(logs, "登录完成")
            _login_tasks[task_id] = {
                **task,
                "status": "done",
                "message": done_msg,
                "finished_at": datetime.now().isoformat(timespec="seconds"),
            }
        except Exception as e:
            _login_tasks[task_id] = {
                **task,
                "status": "error",
                "message": str(e),
                "finished_at": datetime.now().isoformat(timespec="seconds"),
            }

    Thread(target=_run, daemon=True).start()
    return task


# ── Pipeline ──────────────────────────────────────────────────────────────────

def _level_from_line(line: str) -> str:
    if "[ERROR]" in line or "[error]" in line:
        return "error"
    if "[WARN]" in line or "[warn]" in line:
        return "warn"
    if "[OK]" in line or "[ok]" in line:
        return "ok"
    return "debug"


def _start_stream_threads(
    proc: subprocess.Popen,
    *,
    log_stdout: bool = False,
    stdout_level: str = "debug",
    log_prefix: str = "",
):
    """后台读取 stdout/stderr，避免管道阻塞；按需实时写入前端日志。"""
    import threading as _threading

    stdout_chunks: List[str] = []
    stderr_chunks: List[str] = []

    def _read_stream(stream, collector: List[str], is_stderr: bool):
        for line in stream:
            line = line.rstrip()
            if not line or line.startswith("Debugger"):
                continue
            collector.append(line)
            if is_stderr or log_stdout:
                level = _level_from_line(line) if is_stderr else stdout_level
                add_log(f"{log_prefix}{line}", level)

    t_err = _threading.Thread(
        target=_read_stream, args=(proc.stderr, stderr_chunks, True), daemon=True
    )
    t_out = _threading.Thread(
        target=_read_stream, args=(proc.stdout, stdout_chunks, False), daemon=True
    )
    t_err.start()
    t_out.start()
    return stdout_chunks, stderr_chunks, t_out, t_err

_EXPLORE_PID_FILE = Path("/tmp/xhs_explore_node.pid")
_explore_proc: "subprocess.Popen | None" = None  # 全局持有当前 node 子进程


def _kill_explore_proc(pid: Optional[int] = None) -> None:
    """Kill the node explore subprocess. Uses global _explore_proc or a PID from file."""
    global _explore_proc
    # 先 kill 内存中持有的进程对象
    if _explore_proc is not None:
        try:
            _explore_proc.kill()
            _explore_proc.wait(timeout=5)
        except Exception:
            pass
        _explore_proc = None
    # 再用 PID 文件兜底（服务重启后内存丢失）
    target_pid = pid
    if target_pid is None and _EXPLORE_PID_FILE.exists():
        try:
            target_pid = int(_EXPLORE_PID_FILE.read_text().strip())
        except Exception:
            pass
    if target_pid:
        try:
            os.kill(target_pid, 9)
        except ProcessLookupError:
            pass  # 已退出
        except Exception:
            pass
    try:
        _EXPLORE_PID_FILE.unlink(missing_ok=True)
    except Exception:
        pass


# 服务启动时清理上次残留的 node 进程
_kill_explore_proc()


_CARD_SCHEMES = {
    # keyword列表 → (accent色, 背景色, 文字色, oval描边色)
    "红": ("#CC0000", "#fff", "#1a1a1a", "#CC0000"),
    "蓝": ("#0057D9", "#fff", "#1a1a1a", "#0057D9"),
    "绿": ("#1E7D34", "#fff", "#1a1a1a", "#1E7D34"),
    "橙": ("#E85D04", "#fff", "#1a1a1a", "#E85D04"),
    "紫": ("#6B21A8", "#fff", "#1a1a1a", "#6B21A8"),
    "黑": ("#111", "#111", "#f0f0f0", "#f0f0f0"),
    "深色": ("#111", "#1a1a1a", "#f0f0f0", "#f5f5f5"),
    "极简": ("#333", "#fff", "#1a1a1a", "#333"),
    "暖": ("#C05621", "#FFF8F2", "#2D1B00", "#C05621"),
    "清新": ("#0D7A5F", "#F0FDF4", "#1a1a1a", "#0D7A5F"),
    "粉": ("#D63384", "#fff", "#1a1a1a", "#D63384"),
}
_SCHEME_DEFAULT = ("#CC0000", "#fff", "#1a1a1a", "#CC0000")

def _resolve_card_scheme(style_text: str) -> tuple:
    """根据图片风格文字匹配配色方案，返回 (accent, bg, text, oval)。"""
    for kw, scheme in _CARD_SCHEMES.items():
        if kw in style_text:
            return scheme
    return _SCHEME_DEFAULT


def _md_inline(raw: str) -> str:
    """HTML-escape a string then convert **bold** markers to <strong>."""
    escaped = _html.escape(raw)
    return re.sub(r'\*\*([^*]+)\*\*', r'<strong>\1</strong>', escaped)


def _gen_cards_for_draft(draft_path: Path, acc_name_display: str = "", force_template: str = "") -> list:
    """从单篇文案生成 info-card PNG，返回 URL 路径列表（/output/draft_cards/…）。"""
    capture_js = CARD_TOOL_DIR / "scripts" / "capture_card.js"
    if not capture_js.exists():
        raise FileNotFoundError(f"找不到 capture_card.js: {capture_js}")

    stem = draft_path.stem

    card_dir = DRAFT_DIR / stem
    card_dir.mkdir(parents=True, exist_ok=True)

    content = draft_path.read_text(encoding="utf-8")
    title_m = re.search(r'^# (.+)$', content, re.MULTILINE)
    hook_m = re.search(r'\*\*Hook\*\*：(.+)', content)
    cta_m = re.search(r'\n([^\n]{5,60}[💌📌👇])\s*\n', content)
    style_m = re.search(r'\*\*图片风格\*\*[：:]\s*([^\n]+)', content)

    _e = _html.escape  # 所有插入 HTML 的用户内容必须转义
    title_text = _strip_card_hashtags(title_m.group(1)) if title_m else stem
    hook_text = _strip_card_hashtags(hook_m.group(1).strip()) if hook_m else "刷到就是缘分！"
    cta_text = _strip_card_hashtags(cta_m.group(1).strip()) if cta_m else "收藏备用 💌"
    title = _e(title_text)
    hook = _e(hook_text)
    cta = _e(cta_text)
    style_text = style_m.group(1).strip() if style_m else ""
    # 优先使用文案中记录的人设配色，找不到时用当前激活人设，再退回风格关键词匹配
    persona_m = re.search(r'\*\*人设\*\*[：:]\s*(\S+)', content)
    persona_id = persona_m.group(1).strip() if persona_m else ""
    _persona = persona_manager.get_persona_by_id(persona_id) if persona_id else None
    if _persona is None:
        _persona = persona_manager.get_active_persona()
    if _persona:
        accent     = _persona.get("card_accent", "#CC0000")
        card_bg    = _persona.get("card_bg",     "#fff")
        text_color = _persona.get("card_text",   "#1a1a1a")
        oval_color = _persona.get("card_oval",   accent)
    else:
        accent, card_bg, text_color, oval_color = _resolve_card_scheme(style_text)
    card_template = force_template or (_persona or {}).get("card_template", "oval")
    # 深色方案文字反色
    info_bg = f"color-mix(in srgb, {accent} 8%, {card_bg})" if card_bg == "#fff" or card_bg == "#FFF8F2" or card_bg == "#F0FDF4" else f"rgba(255,255,255,0.08)"
    tip_bg = "#f9f9f9" if card_bg == "#fff" else f"rgba(255,255,255,0.08)"

    exclude_fields = {"账号参考","内容结构","Hook","情绪","图片风格","参考字数","封面参考图","主题","图文风格描述"}

    # 提取正文（--- 之后）
    body_start = content.find('\n---\n')
    body_text = content[body_start + 5:] if body_start >= 0 else content
    # 去掉标题行和话题标签行
    body_text = re.sub(r'^#+\s+[^\n]+', '', body_text.strip(), flags=re.MULTILINE)
    body_text = _strip_card_hashtags(body_text)

    # 方案1：数字/emoji 编号开头 + **粗体标题**（如 1️⃣ **xxx**、**xxx**）
    # 先去掉行首的 emoji 编号前缀再匹配标题
    body_norm = re.sub(r'^[\d️⃣1️⃣2️⃣3️⃣4️⃣5️⃣🔥✅⚠️💡📌👉\s]*\*\*', '**', body_text, flags=re.MULTILINE)
    sections = re.findall(r'\*\*([^*\n]{2,50})\*\*[：:]?\s*\n((?:(?!\*\*[^*\n]{2,50}\*\*)[^\n]+\n?)*)', body_norm)
    content_sections = [(h.strip(), b.strip()) for h, b in sections if h.strip() not in exclude_fields and len(h.strip()) > 2]

    # 方案2：冒号结尾的标题行（放宽到40字）
    if len(content_sections) < 2:
        colon_secs = re.findall(
            r'^([^*#\n]{2,40}[：:])\s*\n((?:[^\n]+\n?)+?)(?=\n\n|[^*#\n]{2,40}[：:]|$)',
            body_text, re.MULTILINE
        )
        content_sections = [(h.rstrip('：:').strip(), b.strip()) for h, b in colon_secs]

    # 方案3：按双换行拆段落
    if len(content_sections) < 2:
        paras = [p.strip() for p in re.split(r'\n{2,}', body_text)
                 if p.strip() and len(p.strip()) > 15 and not p.strip().startswith('#')]
        content_sections = [(p[:20].rstrip('，。！？ '), p) for p in paras]

    # 不写死张数：按实际段落数生成，最少1段最多6段（封面+内容+CTA = 3~8张）
    content_sections = content_sections[:6]

    words_m = re.search(r'\*\*参考字数\*\*：(\d+)', content)
    card_avg_words = int(words_m.group(1)) if words_m else 300
    card_sec_chars = max(60, card_avg_words // max(1, len(content_sections)))

    slug = re.sub(r'[^\w\-\u4e00-\u9fff]', '-', title)[:20]

    # ── 合并要点少的段落，避免一张卡只有 2 条内容大量留白 ────────────────
    def _sec_line_count(sb_text: str) -> int:
        return max(1, len([l for l in sb_text.split('\n') if l.strip()]))

    _CARD_MIN = 3   # 低于此行数时合并到上一组
    _CARD_MAX = 7   # 超过此行数时新起一组
    card_groups: list[list[tuple]] = []
    _cur: list[tuple] = []
    _cnt = 0
    for _sh, _sb in content_sections:
        _n = _sec_line_count(_sb)
        if _cnt + _n > _CARD_MAX and _cnt >= _CARD_MIN:
            card_groups.append(_cur)
            _cur, _cnt = [(_sh, _sb)], _n
        else:
            _cur.append((_sh, _sb))
            _cnt += _n
    if _cur:
        card_groups.append(_cur)

    card_defs = [("01-cover", "", "")]
    for ci, grp in enumerate(card_groups, 2):
        card_defs.append((f"{ci:02d}-content", "", ""))
    card_defs.append((f"{len(card_defs)+1:02d}-cta", "", ""))

    n_cards = len(card_defs)
    url_paths = []

    # 公共 CSS — 颜色由配色方案变量注入
    common_css = f"""
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #e8e8e8; display: flex; justify-content: center; padding: 40px; }}
.card-container {{
  width: 900px; height: 1200px; background: {card_bg};
  border-radius: 16px; font-family: 'PingFang SC', 'Hiragino Sans GB', 'Microsoft YaHei', -apple-system, sans-serif;
  position: relative; display: flex; flex-direction: column;
  box-shadow: 0 4px 20px rgba(0,0,0,0.10); overflow: hidden;
}}
.card-footer {{
  padding: 16px 84px 28px;
  display: flex; align-items: center; justify-content: flex-end;
  border-top: 1px solid rgba(128,128,128,0.15);
}}
.page-tag {{ font-size: 13px; color: {accent}; opacity: 0.5; letter-spacing: 0.04em; }}
.oval-zone {{ position: relative; height: 400px; padding: 60px 84px 44px; display: flex; align-items: center; }}
.oval-svg {{ position: absolute; top: 0; left: 0; pointer-events: none; }}
.oval-content {{ position: relative; z-index: 2; }}
.badge {{ font-size: 15px; font-weight: 900; color: {accent}; letter-spacing: 0.08em; text-transform: uppercase; margin-bottom: 10px; display: block; }}
h1 {{ font-size: 52px; font-weight: 900; color: {text_color}; line-height: 1.2; letter-spacing: -0.02em; }}
.subtitle {{ font-size: 18px; color: {text_color}; opacity: 0.55; margin-top: 10px; line-height: 1.5; }}
.body-section {{ flex: 1; padding: 36px 84px; display: flex; flex-direction: column; gap: 24px; }}
.info-box {{
  background: {info_bg}; border-left: 4px solid {accent};
  border-radius: 0 10px 10px 0; padding: 18px 22px;
}}
.info-label {{ font-size: 12px; font-weight: 900; color: {accent}; letter-spacing: 0.1em; margin-bottom: 8px; }}
.info-text {{ font-size: 16px; color: {text_color}; opacity: 0.75; line-height: 1.8; }}
.section-title {{
  font-size: 20px; font-weight: 900; color: {text_color};
  border-bottom: 2px solid rgba(128,128,128,0.15); padding-bottom: 14px;
}}
.point-list {{ display: flex; flex-direction: column; gap: 18px; }}
.point-row {{ display: flex; align-items: flex-start; gap: 16px; }}
.point-num {{ font-size: 26px; font-weight: 900; color: {accent}; min-width: 32px; line-height: 1.4; flex-shrink: 0; }}
.point-main {{ font-size: 17px; font-weight: 700; color: {text_color}; margin-bottom: 3px; line-height: 1.5;
  word-break: break-word; overflow-wrap: anywhere; white-space: normal; }}
.point-sub {{ font-size: 14px; color: {text_color}; opacity: 0.55; line-height: 1.6;
  word-break: break-word; overflow-wrap: anywhere; }}
.tip-box {{ margin-top: auto; background: {tip_bg}; border-radius: 10px; padding: 16px 20px; font-size: 14px; color: {text_color}; opacity: 0.7; line-height: 1.7; }}
.cta-zone {{ flex: 1; display: flex; flex-direction: column; align-items: center; justify-content: center; padding: 60px 84px; text-align: center; gap: 28px; }}
.cta-emoji {{ font-size: 72px; }}
.cta-text {{ font-size: 28px; font-weight: 900; color: {text_color}; line-height: 1.5; }}
.cta-sub {{ font-size: 16px; color: {text_color}; opacity: 0.55; line-height: 1.7; }}
.cta-tag {{ display: inline-block; background: {accent}; color: #fff; font-size: 16px; font-weight: 700; padding: 10px 32px; border-radius: 40px; margin-top: 8px; }}
"""

    # ── Memo 人设专用 CSS（iPhone Notes 备忘录风）───────────────────────────
    _line_h = 52   # 横线间距 px
    memo_css = f"""
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #C8C6BF; display: flex; justify-content: center; padding: 40px; }}
.card-container {{
  width: 900px; height: 1200px;
  background-color: #FEFEF6;
  background-image:
    repeating-linear-gradient(
      transparent 0px, transparent {_line_h - 1}px,
      #D8D5CD {_line_h - 1}px, #D8D5CD {_line_h}px
    ),
    linear-gradient(
      to right,
      transparent 96px, #F5B0B0 96px, #F5B0B0 98px, transparent 98px
    );
  border-radius: 12px;
  font-family: -apple-system, 'PingFang SC', 'Hiragino Sans GB', sans-serif;
  position: relative; display: flex; flex-direction: column;
  box-shadow: 0 4px 24px rgba(0,0,0,0.22); overflow: hidden;
}}
/* iOS Notes 顶栏 */
.notes-bar {{
  background: rgba(254,254,246,0.97);
  border-bottom: 0.5px solid #D6D3CB;
  padding: 0 40px;
  height: 76px;
  display: flex; align-items: center; justify-content: space-between;
  flex-shrink: 0; z-index: 5;
}}
.notes-back {{
  font-size: 28px; color: #E8A000;
  display: flex; align-items: center; gap: 4px;
}}
.notes-arrow {{ font-size: 36px; line-height: 1; margin-right: 2px; }}
.notes-icons {{ display: flex; gap: 28px; color: #E8A000; font-size: 28px; }}
/* 页码 */
.card-footer {{
  padding: 8px 56px 22px 128px;
  display: flex; justify-content: flex-end; flex-shrink: 0;
}}
.page-tag {{ font-size: 16px; color: #AAA; letter-spacing: 0.04em; }}
/* 封面 */
.cover-body {{
  flex: 1; padding: 36px 60px 32px 128px;
  display: flex; flex-direction: column; overflow: hidden;
}}
.cover-headline {{
  font-size: 64px; line-height: 1.35; color: #111;
  font-family: 'Kaiti SC', 'STKaiti', '楷体', 'Songti SC', serif;
  font-weight: 400; margin-bottom: {_line_h * 2}px;
  word-break: break-word;
}}
.cover-hook {{
  font-size: 52px; line-height: 1.5; color: #111;
  font-family: 'Kaiti SC', 'STKaiti', '楷体', 'Songti SC', serif;
  font-weight: 400; word-break: break-word;
}}
mark {{
  background: rgba(255, 230, 55, 0.58);
  border-radius: 3px; padding: 0 3px;
}}
/* 内容区 */
.content-body {{
  flex: 1; padding: 24px 60px 16px 128px;
  display: flex; flex-direction: column;
  overflow: hidden;
}}
.c-sec-head {{
  font-size: 27px; font-weight: 700; color: #111;
  margin: 16px 0 8px; line-height: 1.5;
}}
.c-sec-head:first-child {{ margin-top: 4px; }}
.c-para {{
  font-size: 24px; color: #222; line-height: 1.8;
  margin-bottom: 12px; word-break: break-word; overflow-wrap: anywhere;
}}
.c-para strong {{ font-weight: 700; color: #111; }}
.c-item {{
  font-size: 23px; color: #333; line-height: 1.7;
  margin-bottom: 8px; padding-left: 28px; position: relative;
  word-break: break-word; overflow-wrap: anywhere;
}}
.c-item strong {{ font-weight: 700; color: #111; }}
.c-wrong {{ color: #CC2200; }}
.c-right {{ color: #1A7A1A; font-weight: 700; }}
.c-tips-head {{
  font-size: 25px; font-weight: 700; color: #333;
  margin: 16px 0 8px; line-height: 1.5;
}}
.c-tips-item {{
  font-size: 22px; color: #444; line-height: 1.7;
  margin-bottom: 8px; padding-left: 28px; position: relative;
  word-break: break-word; overflow-wrap: anywhere;
}}
.c-tips-item::before {{ content: "•"; position: absolute; left: 6px; color: #888; }}
.c-tips-item strong {{ font-weight: 700; color: #222; }}
/* CTA */
.cta-zone {{
  flex: 1; display: flex; flex-direction: column;
  align-items: center; justify-content: center;
  padding: 60px 60px; text-align: center; gap: 28px;
}}
.cta-icon {{ font-size: 72px; }}
.cta-main {{ font-size: 36px; font-weight: 700; color: #111; line-height: 1.4; }}
.cta-sub {{ font-size: 22px; color: #666; line-height: 1.7; }}
.cta-btn {{
  background: {accent}; color: #fff; font-size: 22px; font-weight: 700;
  padding: 16px 52px; border-radius: 40px; display: inline-block;
}}
"""

    # ── 招聘帖专用 CSS（参考喜茶风：纯黄底 · 大标题居中 · 薪资对比色）────────
    job_css = f"""
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #D4C84A; display: flex; justify-content: center; padding: 40px; }}
.card-container {{
  width: 900px; height: 1200px;
  background: linear-gradient(180deg, #FFE84D 0%, #FFD700 60%, #FFC200 100%);
  border-radius: 20px;
  font-family: -apple-system, 'PingFang SC', 'Hiragino Sans GB', sans-serif;
  display: flex; flex-direction: column; align-items: center; justify-content: center;
  box-shadow: 0 8px 32px rgba(0,0,0,0.20); overflow: hidden;
  position: relative; text-align: center;
}}
/* 封面：大 logo 区 */
.job-logo-zone {{
  padding: 0 0 32px;
  display: flex; flex-direction: column; align-items: center; gap: 16px;
}}
.job-logo-emoji {{ font-size: 110px; line-height: 1; }}
.job-brand {{
  font-size: 72px; font-weight: 900; color: #111;
  letter-spacing: 0.03em; line-height: 1.2; word-break: break-word;
  text-align: center; width: 100%; padding: 0 48px;
}}
/* 关键信息块 */
.job-info-block {{
  padding: 0 60px; display: flex; flex-direction: column;
  align-items: center; gap: 20px;
}}
.job-location-line {{
  font-size: 36px; font-weight: 700; color: #111; line-height: 1.4;
}}
.job-lang-tag {{
  font-size: 32px; font-weight: 700; color: #111;
}}
.job-salary-line {{
  font-size: 38px; font-weight: 900; color: #111; margin-top: 8px;
}}
.job-salary-val {{
  color: #D4000A; font-size: 44px; font-weight: 900;
}}
/* 底部装饰 */
.job-deco {{
  padding: 32px 60px 0;
  display: flex; flex-direction: column; align-items: center; gap: 10px;
}}
.job-collab {{ font-size: 18px; color: rgba(0,0,0,0.45); letter-spacing: 0.08em; }}
.job-deco-emojis {{ font-size: 48px; letter-spacing: 4px; }}
/* 内容卡（列表信息） */
.job-content-wrap {{
  width: 100%; flex: 1; padding: 52px 64px 20px;
  display: flex; flex-direction: column; overflow: hidden; text-align: left;
}}
.job-content-header {{
  font-size: 28px; font-weight: 900; color: #111;
  border-bottom: 3px solid rgba(0,0,0,0.2);
  padding-bottom: 14px; margin-bottom: 24px;
  display: flex; align-items: center; gap: 10px;
}}
.job-item {{
  font-size: 26px; color: #111; line-height: 1.7;
  margin-bottom: 14px; padding-left: 32px; position: relative;
  word-break: break-word; overflow-wrap: anywhere; font-weight: 500;
}}
.job-item::before {{
  content: "▸"; position: absolute; left: 0; color: rgba(0,0,0,0.4); font-size: 22px;
}}
.job-item strong {{ font-weight: 800; color: #111; }}
.job-item .hi {{ color: #D4000A; font-weight: 800; }}
.job-para {{
  font-size: 25px; color: #222; line-height: 1.8;
  margin-bottom: 14px; word-break: break-word; font-weight: 500;
}}
/* CTA 卡 */
.job-cta-wrap {{
  flex: 1; display: flex; flex-direction: column;
  align-items: center; justify-content: center;
  padding: 60px 60px; gap: 28px; text-align: center;
}}
.job-cta-emoji {{ font-size: 80px; }}
.job-cta-title {{ font-size: 44px; font-weight: 900; color: #111; line-height: 1.3; }}
.job-cta-sub {{ font-size: 26px; color: rgba(0,0,0,0.55); line-height: 1.7; }}
.job-cta-btn {{
  background: #111; color: #FFE84D; font-size: 26px; font-weight: 900;
  padding: 18px 60px; border-radius: 50px; display: inline-block;
  letter-spacing: 0.06em; margin-top: 8px;
}}
.page-tag {{ font-size: 15px; color: rgba(0,0,0,0.3); letter-spacing: 0.04em; margin-top: auto; padding-bottom: 20px; }}
.card-footer {{ display: none; }}
"""

    # ── 坐标卡专用 CSS（参考京东帖：羊皮纸底 · 打字机日期 · 大坐标 · 可爱图标）
    from datetime import date as _date
    _today = _date.today().strftime("%B %d, %Y")
    spot_css = f"""
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #C8C2B5; display: flex; justify-content: center; padding: 40px; }}
.card-container {{
  width: 900px; height: 1200px;
  background-color: #EDE8DC;
  background-image: radial-gradient(circle, rgba(0,0,0,0.04) 1px, transparent 1px);
  background-size: 18px 18px;
  border-radius: 16px;
  font-family: -apple-system, 'PingFang SC', 'Hiragino Sans GB', sans-serif;
  display: flex; flex-direction: column;
  box-shadow: 0 6px 28px rgba(0,0,0,0.18); overflow: hidden;
  position: relative;
}}
/* 打字机日期 */
.spot-date {{
  position: absolute; top: 44px; left: 52px;
  font-family: 'Courier New', 'Menlo', monospace;
  font-size: 22px; color: #888; line-height: 1.4; letter-spacing: 0.01em;
}}
/* 核心内容区 — 垂直居中 */
.spot-center {{
  flex: 1; display: flex; flex-direction: column;
  align-items: flex-start; justify-content: center;
  padding: 0 72px 0 72px; gap: 0;
}}
.spot-pin {{ font-size: 80px; line-height: 1; margin-bottom: 20px; }}
.spot-company {{
  font-size: 60px; font-weight: 800; color: #111;
  line-height: 1.25; word-break: break-word; margin-bottom: 8px;
}}
.spot-sub {{
  font-size: 36px; color: #444; line-height: 1.4;
  margin-bottom: 48px; word-break: break-word;
}}
.spot-salary {{
  font-size: 44px; font-weight: 700; color: #111;
  line-height: 1.5; margin-bottom: 4px;
}}
.spot-timeline {{
  font-size: 40px; font-weight: 700; color: #111; line-height: 1.5;
}}
/* 右下角可爱图标 */
.spot-deco {{
  position: absolute; bottom: 36px; right: 52px;
  font-size: 88px; line-height: 1; opacity: 0.88;
}}
/* 内容卡 */
.spot-content {{
  flex: 1; padding: 100px 72px 40px; display: flex; flex-direction: column; overflow: hidden;
}}
.spot-content-head {{
  font-size: 28px; font-weight: 800; color: #555;
  border-bottom: 2px solid rgba(0,0,0,0.12);
  padding-bottom: 12px; margin-bottom: 22px; letter-spacing: 0.05em;
}}
.spot-item {{
  font-size: 26px; color: #222; line-height: 1.75;
  margin-bottom: 14px; padding-left: 30px; position: relative;
  word-break: break-word; overflow-wrap: anywhere; font-weight: 500;
}}
.spot-item::before {{ content: "▸"; position: absolute; left: 0; color: #999; }}
.spot-item .hi {{ color: #C0392B; font-weight: 700; }}
.spot-para {{
  font-size: 26px; color: #333; line-height: 1.8;
  margin-bottom: 14px; word-break: break-word;
}}
/* CTA */
.spot-cta {{
  flex: 1; display: flex; flex-direction: column;
  align-items: center; justify-content: center;
  padding: 60px; text-align: center; gap: 28px;
}}
.spot-cta-icon {{ font-size: 80px; }}
.spot-cta-title {{ font-size: 42px; font-weight: 900; color: #111; line-height: 1.35; }}
.spot-cta-sub {{ font-size: 26px; color: #666; line-height: 1.7; }}
.spot-cta-btn {{
  background: #111; color: #EDE8DC; font-size: 25px; font-weight: 800;
  padding: 18px 60px; border-radius: 50px; display: inline-block; letter-spacing: 0.05em;
}}
.page-tag {{ font-size: 15px; color: rgba(0,0,0,0.25); }}
.card-footer {{ display: none; }}
"""

    for ci, (cname, _bg, _body_html) in enumerate(card_defs):
        page_tag = f"{ci+1:02d} / {n_cards:02d}"
        is_cover = ci == 0
        is_cta   = ci == n_cards - 1

        if is_cover:
            if card_template == "spot":
                # ── 坐标卡封面：羊皮纸 + 打字机日期 + 大📍 + 薪资 + 可爱图标
                _salary_m = re.search(r'[£￡]\s*[\d,]+(?:\s*[-–]\s*[\d,]+)?(?:\s*[kK])?(?:/(?:month|yr|year|mo))?', body_text)
                _salary_str = _html.escape(_salary_m.group(0).strip()) if _salary_m else ""
                # 提取时间线关键词（X周下offer / X weeks）
                _tl_m = re.search(r'\d+\s*(?:周|weeks?)[^\n]{0,20}(?:下\s*offer|左右)?', body_text, re.IGNORECASE)
                _tl_str = _html.escape(_tl_m.group(0).strip()) if _tl_m else ""
                # 副标题：括号内容或 hook 前半句
                _sub = ""
                _paren = re.search(r'[（(]([^）)]{4,30})[）)]', body_text)
                if _paren:
                    _sub = _html.escape(_paren.group(1))
                elif hook_text and len(hook_text) > 4:
                    _sub = _html.escape(hook_text[:30])
                body = f"""
  <div class="spot-date">{_today}</div>
  <div class="spot-center">
    <div class="spot-pin">📍</div>
    <div class="spot-company">{title}</div>
    {f'<div class="spot-sub">（{_sub}）</div>' if _sub else ''}
    {f'<div class="spot-salary">£ {_salary_str.lstrip("£￡").strip()}</div>' if _salary_str else ''}
    {f'<div class="spot-timeline">{_tl_str}</div>' if _tl_str else ''}
  </div>
  <div class="spot-deco">🐻</div>"""
            elif card_template == "job":
                # ── 招聘封面：喜茶风 大 emoji + 品牌名 + 薪资红色高亮 ───────
                _salary_m = re.search(r'[£￡]\s*[\d,]+(?:\s*[-–]\s*[\d,]+)?(?:\s*[kK])?(?:/(?:month|yr|year|mo))?', body_text)
                _salary_str = _html.escape(_salary_m.group(0).strip()) if _salary_m else ""
                # 提取地点关键词
                _loc_m = re.search(r'(?:base\s+)?([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)|伦敦|London|英国', body_text)
                _loc_str = _loc_m.group(0).strip() if _loc_m else "London"
                body = f"""
  <div class="job-logo-zone">
    <div class="job-logo-emoji">📢</div>
    <div class="job-brand">{title}</div>
  </div>
  <div class="job-info-block">
    <div class="job-location-line">📍 base {_html.escape(_loc_str)}</div>
    <div class="job-lang-tag">🇨🇳 中文优势</div>
    {f'<div class="job-salary-line">salary：<span class="job-salary-val">{_salary_str}</span></div>' if _salary_str else f'<div class="job-salary-line">{hook}</div>'}
  </div>
  <div class="job-deco">
    <div class="job-collab">JOIN US · WE'RE HIRING</div>
    <div class="job-deco-emojis">🌟✨⭐🌟✨</div>
  </div>"""
            elif card_template == "memo":
                # 把 **bold** 转为黄色高亮，用 raw 文本操作
                def _hi(raw: str) -> str:
                    return re.sub(r'\*\*([^*]+)\*\*', r'<mark>\1</mark>', _html.escape(raw))
                # 从正文里取第一句有内容的话作为封面 hook
                _cover_hook_raw = ""
                for _bl in body_text.split('\n'):
                    _bl = _bl.strip()
                    if _bl and len(_bl) > 8 and not _bl.startswith(('✅','💯','#','**','---')):
                        _cover_hook_raw = _bl
                        break
                _cover_hook_html = _hi(_cover_hook_raw) if _cover_hook_raw else ""
                body = f"""
  <div class="notes-bar">
    <div class="notes-back"><span class="notes-arrow">‹</span> 备忘录</div>
    <div class="notes-icons"><span>⬆</span><span>···</span></div>
  </div>
  <div class="cover-body">
    <p class="cover-headline">{_hi(title_text)}</p>
    {f'<p class="cover-hook">{_cover_hook_html}</p>' if _cover_hook_html else ''}
  </div>"""
            else:
                # 封面：手绘椭圆 + 大标题 + hook
                title_lines = title.replace('，', '<br>').replace(',', '<br>') if len(title) > 10 else title
                oval_svg = f"""<svg class="oval-svg" width="900" height="400" viewBox="0 0 900 400">
  <path d="M 690,46 C 752,16 850,54 850,138 C 850,222 786,312 616,350
           C 446,388 234,386 98,366 C -38,346 -46,262 12,178
           C 70,94 168,32 334,6 C 500,-20 628,14 690,46
           C 712,57 726,64 720,58"
        stroke="{oval_color}" stroke-width="13" fill="none"
        stroke-linecap="round" stroke-linejoin="round"/>
</svg>"""
                cover_body = f"""
  <div class="oval-zone">
    {oval_svg}
    <div class="oval-content">
      <span class="badge">✨ 干货分享</span>
      <h1>{title_lines}</h1>
      <p class="subtitle">{hook}</p>
    </div>
  </div>
  <div class="body-section">
    <div class="info-box">
      <div class="info-label">本期内容</div>
      <div class="info-text">{"<br>".join(f"• {h}" for h, _ in content_sections[:4])}</div>
    </div>
    <div class="tip-box">💡 {hook if len(hook) > 5 else "收藏备用，用到的时候找得到"}</div>
  </div>"""
                body = cover_body

        elif is_cta:
            if card_template == "spot":
                body = f"""
  <div class="spot-date">{_today}</div>
  <div class="spot-cta">
    <div class="spot-cta-icon">💌</div>
    <div class="spot-cta-title">{cta}</div>
    <div class="spot-cta-sub">感兴趣私信获取完整 JD<br>评论区扣 <strong>1</strong> 都回 ✨</div>
    <span class="spot-cta-btn">私信了解</span>
    <div class="page-tag">{page_tag}</div>
  </div>
  <div class="spot-deco">🌟</div>"""
            elif card_template == "job":
                body = f"""
  <div class="job-cta-wrap">
    <div class="job-cta-emoji">💌</div>
    <div class="job-cta-title">{cta}</div>
    <div class="job-cta-sub">感兴趣私信获取完整 JD<br>或评论区扣 <strong>1</strong> ✨</div>
    <span class="job-cta-btn">私信了解</span>
    <div class="page-tag">{page_tag}</div>
  </div>"""
            elif card_template == "memo":
                body = f"""
  <div class="notes-bar">
    <div class="notes-back"><span class="notes-arrow">‹</span> 备忘录</div>
    <div class="notes-icons"><span>⬆</span><span>···</span></div>
  </div>
  <div class="cta-zone">
    <div class="cta-icon">💌</div>
    <div class="cta-main">{cta}</div>
    <div class="cta-sub">觉得有用就收藏，转发给需要的朋友 ✨</div>
    <span class="cta-btn">收藏备用</span>
  </div>"""
            else:
                body = f"""
  <div class="cta-zone">
    <div class="cta-emoji">💌</div>
    <div class="cta-text">{cta}</div>
    <div class="cta-sub">觉得有用就收藏，转发给需要的朋友 ✨</div>
    <span class="cta-tag">关注不迷路</span>
  </div>"""

        else:
            # 内容卡：支持单段落和多段落合并
            sec_idx = ci
            grp = card_groups[sec_idx - 1] if sec_idx - 1 < len(card_groups) else []

            if card_template == "spot":
                # ── 坐标卡内容卡 ──────────────────────────────────────────
                _hi_s = lambda s: re.sub(
                    r'([£￡]\s*[\d,]+(?:\s*[-–]\s*[\d,]+)?(?:\s*[kK])?(?:/\w+)?)',
                    r'<span class="hi">\1</span>', s)
                def _spot_line(raw: str) -> str:
                    l = raw.strip()
                    if not l: return ''
                    if re.match(r'^[-•·✅💯✓]\s', l) or re.match(r'^\d+[.、）)]\s', l):
                        clean = re.sub(r'^[-•·✅💯✓→]\s*', '', l)
                        clean = re.sub(r'^\d+[.、）)\s]\s*', '', clean)
                        return f'<p class="spot-item">{_hi_s(_md_inline(clean))}</p>'
                    return f'<p class="spot-para">{_hi_s(_md_inline(l))}</p>'

                inner = ''
                for _sh, _sb in grp:
                    if _sh:
                        inner += f'<div class="spot-content-head">{_html.escape(_sh).upper()}</div>'
                    for _l in _sb.split('\n'):
                        inner += _spot_line(_l)
                if not inner:
                    inner = '<div class="spot-content-head">岗位详情</div>'
                    for _l in body_text.split('\n'):
                        inner += _spot_line(_l)

                body = f"""
  <div class="spot-date">{_today}</div>
  <div class="spot-content">
    {inner}
    <div class="page-tag" style="margin-top:auto;text-align:right">{page_tag}</div>
  </div>
  <div class="spot-deco" style="font-size:60px">✨</div>"""
            elif card_template == "job":
                # ── 招聘内容卡：纯黄底 + 左对齐列表 ──────────────────────
                _hi_salary = lambda s: re.sub(
                    r'([£￡]\s*[\d,]+(?:\s*[-–]\s*[\d,]+)?(?:\s*[kK])?(?:/\w+)?)',
                    r'<span class="hi">\1</span>', s)
                def _job_line(raw: str) -> str:
                    l = raw.strip()
                    if not l: return ''
                    if re.match(r'^[-•·✅💯]\s', l) or re.match(r'^\d+[.、）)]\s', l):
                        clean = re.sub(r'^[-•·✅💯✓→]\s*', '', l)
                        clean = re.sub(r'^\d+[.、）)\s]\s*', '', clean)
                        return f'<p class="job-item">{_hi_salary(_md_inline(clean))}</p>'
                    return f'<p class="job-para">{_hi_salary(_md_inline(l))}</p>'

                inner_html = ''
                # 用 section header 作标题
                used_header = False
                for _sh, _sb in grp:
                    if _sh and not used_header:
                        inner_html += f'<div class="job-content-header">📌 {_html.escape(_sh)}</div>'
                        used_header = True
                    elif _sh:
                        inner_html += f'<div class="job-content-header" style="margin-top:24px">📌 {_html.escape(_sh)}</div>'
                    for _l in _sb.split('\n'):
                        inner_html += _job_line(_l)
                if not inner_html:
                    inner_html = f'<div class="job-content-header">📌 岗位详情</div>'
                    for _l in body_text.split('\n'):
                        inner_html += _job_line(_l)

                body = f"""
  <div class="job-content-wrap">
    {inner_html}
    <div class="page-tag" style="text-align:right">{page_tag}</div>
  </div>"""
            elif card_template == "memo":
                # ── Memo 人设：iPhone Notes 备忘录风格 ──────────────────────

                def _notes_line_html(raw_line: str) -> str:
                    """把一行原始 markdown 文本转为 Notes 风格 HTML 段落。"""
                    l = raw_line.strip()
                    if not l:
                        return ''
                    # ❌ 开头：红色错误示例
                    if l.startswith('❌'):
                        clean = re.sub(r'^❌\s*', '', l)
                        return f'<p class="c-item c-wrong">❌ {_md_inline(clean)}</p>'
                    # ✅ 开头（非 section 标题，而是内联 OK 答案）
                    if l.startswith('✅') and not re.match(r'^✅\s*\*\*', l):
                        clean = re.sub(r'^✅\s*', '', l)
                        return f'<p class="c-item c-right">✅ {_md_inline(clean)}</p>'
                    # 列表项
                    if re.match(r'^[-•·]\s', l):
                        clean = re.sub(r'^[-•·]\s*', '', l)
                        clean = re.sub(r'^\d+[.、）)\s]\s*', '', clean)
                        return f'<p class="c-item">• {_md_inline(clean)}</p>'
                    # 普通数字列表
                    m_num = re.match(r'^\d+[.、）)]\s*(.+)$', l)
                    if m_num:
                        return f'<p class="c-item">• {_md_inline(m_num.group(1))}</p>'
                    # 普通段落（去掉所有前导符号）
                    clean = re.sub(r'^[✅💯❌🔥💡📌👉▶️✓→•·]\s*', '', l)
                    clean = re.sub(r'^\*\*([^*]+)\*\*[：:]\s*', lambda m: '', clean)
                    if clean:
                        return f'<p class="c-para">{_md_inline(clean)}</p>'
                    return ''

                inner_html = ''
                for gi, (grp_sh, grp_sb) in enumerate(grp):
                    is_tips = any(kw in grp_sh.lower() for kw in ["tips", "checklist", "小tips", "注意", "清单"])
                    if is_tips:
                        lines_html = ''
                        for _l in grp_sb.split('\n'):
                            _l = _l.strip()
                            if not _l: continue
                            _l = re.sub(r'^[-•·✅💯]\s*', '', _l)
                            _l = re.sub(r'^\d+[.、）)\s]\s*', '', _l)
                            if _l:
                                lines_html += f'<p class="c-tips-item">{_md_inline(_l)}</p>'
                        inner_html += f'<p class="c-tips-head">💯 {_html.escape(grp_sh)}</p>{lines_html}'
                    else:
                        inner_html += f'<p class="c-sec-head">✅ <strong>{_html.escape(grp_sh)}</strong></p>'
                        for _l in grp_sb.split('\n'):
                            inner_html += _notes_line_html(_l)

                body = f"""
  <div class="notes-bar">
    <div class="notes-back"><span class="notes-arrow">‹</span> 备忘录</div>
    <div class="notes-icons"><span>⬆</span><span>···</span></div>
  </div>
  <div class="content-body">{inner_html}</div>"""

            else:
                # ── 椭圆人设（oval）：原有布局 ──────────────────────────────
                is_multi = len(grp) > 1

                def _clean_lines(sb_text: str) -> list:
                    raw = [l.strip() for l in sb_text.split('\n') if l.strip()]
                    out = []
                    for l in raw:
                        c = re.sub(r'^[\s•·✅⚠️❌🔥💡📌👉📅🌐📝🎥🏆💬🎯📧☕🤝🔹🔸▶️▸◆◇→\-—️⃣]+\s*', '', l).strip()
                        c = re.sub(r'^\d+[.、）)\s]\s*', '', c).strip()
                        c = re.sub(r'\*\*([^*]+)\*\*', r'\1', c)
                        if c and len(c) > 3:
                            out.append(c)
                    if len(out) < 2:
                        out = [s.strip() for s in re.split(r'[。！？；]', sb_text) if s.strip() and len(s.strip()) > 3]
                    return out

                # 收集所有行（用于字号决策）
                all_lines_count = sum(_sec_line_count(sb) for _, sb in grp)
                n = max(1, all_lines_count)

                # 字号/间距自适应
                if n <= 2:
                    h1_size, main_size, sub_size, pt_gap, oval_h, pt_num_size = 52, 34, 22, 60, 190, 44
                elif n <= 3:
                    h1_size, main_size, sub_size, pt_gap, oval_h, pt_num_size = 48, 28, 19, 44, 185, 38
                elif n <= 4:
                    h1_size, main_size, sub_size, pt_gap, oval_h, pt_num_size = 44, 24, 17, 30, 180, 32
                elif n <= 6:
                    h1_size, main_size, sub_size, pt_gap, oval_h, pt_num_size = 38, 20, 15, 18, 170, 26
                elif n <= 8:
                    h1_size, main_size, sub_size, pt_gap, oval_h, pt_num_size = 34, 18, 14, 13, 155, 22
                else:
                    h1_size, main_size, sub_size, pt_gap, oval_h, pt_num_size = 30, 16, 13, 10, 140, 20

                def _point_html(line: str, pi: int) -> str:
                    m = re.match(r'^([^：:，,]{2,18})[：:](.+)$', line)
                    if m:
                        mt = _e(m.group(1).strip())
                        st = _e(m.group(2).strip()[:120])
                    else:
                        mt, st = _e(line), ""
                    sub_div = f"<div class='point-sub' style='font-size:{sub_size}px'>{st}</div>" if st else ""
                    return f"""
      <div class="point-row" style="gap:14px">
        <span class="point-num" style="font-size:{pt_num_size}px;min-width:32px">{pi:02d}</span>
        <div>
          <div class="point-main" style="font-size:{main_size}px">{mt}</div>
          {sub_div}
        </div>
      </div>"""

                if not is_multi:
                    # ── 单段落：原有布局 ──
                    sh_raw, sb = grp[0]
                    sh = _e(sh_raw)
                    badge_text = f"POINT {sec_idx:02d} / 要点{sec_idx}"
                    lines = _clean_lines(sb)
                    points_html = "".join(_point_html(l, i) for i, l in enumerate(lines, 1))
                    body = f"""
  <div class="oval-zone" style="height:{oval_h}px;padding:32px 84px 20px">
    <div class="oval-content">
      <span class="badge">{badge_text}</span>
      <h1 style="font-size:{h1_size}px">{sh}</h1>
    </div>
  </div>
  <div class="body-section" style="padding-top:20px;gap:0">
    <div class="point-list" style="gap:{pt_gap}px;flex:1">{points_html}
    </div>
    <div class="tip-box" style="margin-top:24px">💡 {hook[:50] if len(hook) > 5 else "重点收藏，随时查阅"}</div>
  </div>"""
                else:
                    # ── 多段落合并：小标题区 + 每段 sub-heading + 要点 ──
                    first_sh = _e(grp[0][0])
                    badge_text = f"POINT {sec_idx:02d}"
                    sections_html = ""
                    pi_global = 1
                    for gi, (grp_sh, grp_sb) in enumerate(grp):
                        divider = f'<div style="height:1px;background:rgba(128,128,128,0.12);margin:{pt_gap//2}px 0 {pt_gap//2}px"></div>' if gi > 0 else ""
                        sh_html = f'<div style="font-size:{max(15,main_size-4)}px;font-weight:900;color:{text_color};margin-bottom:{max(8,pt_gap//3)}px">{_e(grp_sh)}</div>'
                        grp_lines = _clean_lines(grp_sb)
                        pts = "".join(_point_html(l, pi_global + i) for i, l in enumerate(grp_lines))
                        pi_global += len(grp_lines)
                        sections_html += f'{divider}<div>{sh_html}<div class="point-list" style="gap:{pt_gap}px">{pts}</div></div>'

                    oval_h_multi = max(130, oval_h - 50)
                    h1_multi = max(28, h1_size - 12)
                    body = f"""
  <div class="oval-zone" style="height:{oval_h_multi}px;padding:28px 84px 16px">
    <div class="oval-content">
      <span class="badge">{badge_text}</span>
      <h1 style="font-size:{h1_multi}px">{first_sh}</h1>
    </div>
  </div>
  <div class="body-section" style="padding-top:16px;gap:0">
    <div style="flex:1">{sections_html}</div>
    <div class="tip-box" style="margin-top:20px">💡 {hook[:50] if len(hook) > 5 else "重点收藏，随时查阅"}</div>
  </div>"""

        _css = (memo_css if card_template == "memo"
                else spot_css if card_template == "spot"
                else job_css if card_template == "job"
                else common_css)
        html = f"""<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<style>{_css}</style>
</head>
<body>
<div class="card-container">
{body}
  <div class="card-footer">
    <span class="page-tag">{page_tag}</span>
  </div>
</div>
</body>
</html>"""

        html_path = card_dir / f"{cname}-{slug}.html"
        png_path  = card_dir / f"{cname}-{slug}.png"
        html_path.write_text(html, encoding="utf-8")
        node_proc = subprocess.run(
            ["node", str(capture_js), str(html_path), str(png_path)],
            capture_output=True, text=True, timeout=60
        )
        html_path.unlink(missing_ok=True)
        if not png_path.exists() or node_proc.returncode != 0:
            err = (node_proc.stderr or "").strip() or (node_proc.stdout or "").strip()
            print(f"[card] ⚠️  卡片生成失败（exit={node_proc.returncode}）{': ' + err[:200] if err else ''}",
                  flush=True)
        if png_path.exists():
            url_paths.append(f"/output/draft_cards/{stem}/{png_path.name}")

    return url_paths


def _run_draft_step(keyword: str, copies: int, mode: str = "deep_report") -> list:
    """用 state['accounts'] 中的分析数据生成文案，写入 DRAFT_DIR，返回路径列表。"""
    DRAFT_DIR.mkdir(parents=True, exist_ok=True)
    draft_paths = []
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_kw = re.sub(r'[\\/:*?"<>|\s]', "_", keyword)[:20]
    draft_counter = [0]  # 全局序号，跨账号递增
    llm_key, llm_url, _ = _load_llm_key()
    llm_cfg = _load_model_config()
    mode = _normalize_mode(mode)

    active_persona = persona_manager.get_active_persona()
    add_log(f"当前人设：{active_persona.get('emoji','')} {active_persona['name']}", "info")
    add_log(f"分析模式：{_mode_label(mode)}", "info")

    for acc in state["accounts"]:
        name = acc.get("账号昵称", "未知账号")
        _aw = acc.get("平均文字字数")
        try:
            avg_words = int(_aw) if _aw and str(_aw) not in ("N/A", "", "None") else 300
        except (ValueError, TypeError):
            avg_words = 300
        style_desc  = str(acc.get("图文风格描述") or "")
        image_style = str(acc.get("图片风格") or "单图大字信息流")
        emotion     = str(acc.get("情绪标签") or "口语化")
        hooks_raw   = str(acc.get("Hook模板", "") or "")
        hook_examples = " | ".join(
            h.strip() for h in re.split(r"[|｜]", hooks_raw) if h.strip()
        )[:100]
        ctas_raw    = str(acc.get("CTA模板", "") or "")
        cta_examples = " | ".join(
            c.strip() for c in re.split(r"[|｜]", ctas_raw) if c.strip()
        )[:100]
        structure = str(acc.get("内容结构", "痛点引入 | 解决方案 | 行动号召") or "")
        style_params = {
            "avg_words": avg_words, "style_desc": style_desc,
            "image_style": image_style, "emotion": emotion,
            "structure": structure, "hook_examples": hook_examples,
            "cta_examples": cta_examples,
        }
        saved_style = _load_account_style(name)
        if saved_style:
            style_params = {k: saved_style.get(k, style_params[k]) for k in style_params}

        # 取该账号点赞/收藏最高那篇笔记的封面图作为视觉参考
        prefetched = (acc.get("explore_data") or {}).get("prefetched_notes", [])
        cover_url = ""
        if prefetched:
            best = max(prefetched, key=lambda p: (p.get("saves", 0) + p.get("likes", 0)), default={})
            cover_url = best.get("cover_url", "") or ""

        for i in range(copies):
            if llm_key:
                try:
                    draft_md = _generate_draft_llm(
                        keyword,
                        style_params,
                        llm_key,
                        llm_url,
                        llm_cfg["text_model"],
                        persona=active_persona,
                        mode=mode,
                    )
                except Exception as e:
                    import urllib.error as _ue
                    _detail = e.read().decode()[:300] if isinstance(e, _ue.HTTPError) else ""
                    add_log(f"  [{name}] LLM 生成失败: {e} | {_detail}", "warn")
                    draft_md = f"# {keyword} — 待完善\n\n（LLM 生成失败，请手动编写）\n"
            else:
                draft_md = f"# {keyword} — 待完善\n\n（未配置 API Key，请在 .env 中配置 DASHSCOPE_API_KEY）\n"

            metadata = (
                f"**主题**：{keyword}\n"
                f"**分析模式**：{_mode_label(mode)}\n"
                f"**人设**：{active_persona['id']}\n"
                f"**图片风格**：{style_params['image_style']}\n"
                f"**图文风格描述**：{style_params['style_desc']}\n"
                f"**情绪**：{style_params['emotion']}\n"
                f"**参考字数**：{style_params['avg_words']}字\n"
                + (f"**封面参考图**：{cover_url}\n" if cover_url else "")
                + "\n---\n\n"
            )
            draft_counter[0] += 1
            title_m2 = re.search(r'^#\s+(.+)$', draft_md, re.MULTILINE)
            if title_m2:
                safe_title = re.sub(r'[\\/:*?"<>|\s]', "_", title_m2.group(1).strip())[:30]
                fname = f"{safe_title}_{ts[:8]}.md"
            else:
                fname = f"文案{draft_counter[0]}_{safe_kw}_{ts[:8]}.md"
            fpath = DRAFT_DIR / fname
            fpath.write_text(metadata + draft_md, encoding="utf-8")
            draft_paths.append(fpath)
            title_match = re.search(r"^#\s+(.+)", draft_md, re.MULTILINE)
            add_log(f"  [{name}] 篇{i+1}：{(title_match.group(1)[:30] if title_match else keyword)}", "ok")

    state["drafts"] = [str(p) for p in draft_paths]
    return draft_paths


def run_pipeline(
    keyword: str,
    account_count: int = 5,
    copies: int = 2,
    retrieval_account: str = "",
    notes_per_account: int = 3,
    mode: str = "deep_report",
):
    global state, _explore_proc
    retrieval_account = (retrieval_account or "").strip()
    mode = _normalize_mode(mode)
    state.update({
        "status": "running", "keyword": keyword,
        "mode": mode,
        "mode_label": _mode_label(mode),
        "retrieval_account": retrieval_account,
        "step": 1, "logs": [], "accounts": [],
        "drafts": [], "cards": [], "stats": {}, "error": "",
    })

    try:
        # ── Step 1: 初始化 ────────────────────────────────────────────────────
        add_log(f"初始化 Pipeline，关键词：「{keyword}」", "info")
        add_log(f"分析模式：{_mode_label(mode)}", "info")
        if retrieval_account:
            add_log(f"指定检索账号：{retrieval_account}", "info")
        state["step"] = 1
        time.sleep(0.5)

        # ── Step 2: FlowLens Agent 完整检索 ──────────────────────────────────
        state["step"] = 2
        add_log(f"FlowLens Agent 启动，关键词「{keyword}」…", "info")

        import asyncio as _asyncio
        from flowlens_agent_pipeline import (
            _call_flowlens_agent,
            _extract_authors_from_site_results,
            _enrich_notes_via_screenshots,
            _build_excel_rows,
            _generate_analysis_html,
        )
        from analyze_accounts import write_excel

        agent_payload = _asyncio.run(_call_flowlens_agent(keyword, notes_per_account))
        run_dir = agent_payload["run_dir"]
        add_log(
            f"Agent 完成：{agent_payload['turns']} 轮，耗时 {agent_payload['total_duration_s']:.0f}s",
            "ok",
        )
        add_log(f"输出目录：{run_dir}", "info")

        # ── Step 3: 解析 site_results → 分析表 ──────────────────────────────
        state["step"] = 3
        add_log("解析 agent 输出，按账号聚合笔记并 AI 分析…", "info")

        authors_all = _extract_authors_from_site_results(run_dir)
        # 按深度笔记数降序，取前 account_count 个进入分析表
        authors_top = dict(
            sorted(authors_all.items(), key=lambda kv: len(kv[1]["notes"]), reverse=True)[:account_count]
        )
        total_notes = sum(len(g["notes"]) for g in authors_top.values())
        add_log(f"找到 {len(authors_top)} 个账号，共 {total_notes} 篇深度笔记", "ok")
        for grp in authors_top.values():
            add_log(f"  - {grp['name']}: {len(grp['notes'])} 篇", "info")

        if not authors_top:
            raise RuntimeError("未从 site_results 中提取到任何账号数据")

        add_log("截图视觉补全缺失数据（正文/互动数）…", "info")
        _enrich_notes_via_screenshots(authors_top, run_dir)

        ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)
        rows = _build_excel_rows(authors_top, keyword)
        from datetime import datetime as _dt
        xlsx_path = ANALYSIS_DIR / f"xhs_account_analysis_{_dt.now().strftime('%Y%m%d')}.xlsx"
        write_excel(rows, str(xlsx_path))
        add_log(f"分析表已保存：{xlsx_path.name}", "ok")

        # 生成 HTML 分析报告到同一目录
        import re as _re
        safe_kw = _re.sub(r'[\\/:*?"<>|\s]', "_", keyword)[:20]
        html_dest = ANALYSIS_DIR / f"xhs_report_{safe_kw}_{_dt.now().strftime('%Y%m%d')}.html"
        try:
            _generate_analysis_html(keyword, authors_top, rows, run_dir, html_dest)
            add_log(f"分析报告已保存：{html_dest.name}", "ok")
        except Exception as _html_e:
            import traceback
            add_log(f"HTML 生成失败：{_html_e}\n{traceback.format_exc()}", "warn")

        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        accounts_data = [
            d for row in ws.iter_rows(min_row=2, values_only=True)
            for d in [dict(zip(headers, row))]
            if d.get("账号昵称")
        ]

        if not accounts_data:
            raise RuntimeError("分析结果为空，请检查 FlowLens agent 输出")

        state["accounts"] = accounts_data[:account_count]
        add_log(f"分析完成，{len(state['accounts'])} 个账号写入报告", "ok")

        # ── Step 4: 生成文案 ──────────────────────────────────────────────────
        state["step"] = 4
        add_log(f"开始生成文案（{len(state['accounts'])} 账号 × {copies} 篇）...", "info")
        draft_paths = _run_draft_step(keyword, copies, mode=mode)
        add_log(f"文案生成完成，共 {len(draft_paths)} 篇", "ok")

        # ── Step 5: 生成图文卡片 ───────────────────────────────────────────────
        state["step"] = 5
        add_log(f"开始生成图文卡片...", "info")

        all_pngs = []

        for draft_path in draft_paths:
            url_paths = _gen_cards_for_draft(draft_path)
            all_pngs.extend(url_paths)
            add_log(f"  {len(url_paths)} 张卡片 → {draft_path.stem}/", "ok")

        state["cards"] = all_pngs
        add_log(f"图文卡片全部完成，共 {len(all_pngs)} 张", "ok")

        # ── 统计 ─────────────────────────────────────────────────────────────
        state["stats"] = {
            "accounts": len(state["accounts"]),
            "drafts": len(state["drafts"]),
            "cards": len(all_pngs),
        }
        state["status"] = "completed"
        state["step"] = 8
        add_log("🎉 Pipeline 全部完成！", "ok")

    except Exception as e:
        import traceback
        state["status"] = "error"
        state["error"] = str(e)
        add_log(f"Pipeline 异常: {e}", "error")
        traceback.print_exc()


# ── OpenAI Chat ───────────────────────────────────────────────────────────────

def _read_env(name: str) -> str:
    """按优先级读取环境变量：BASE_DIR/.env → ~/.baoyu-skills/.env → 系统环境
    优先读文件确保服务器运行期间 .env 热更新（无需重启）立即生效。"""
    for env_path in [BASE_DIR / ".env", Path.home() / ".baoyu-skills" / ".env"]:
        try:
            for line in env_path.read_text().splitlines():
                if line.startswith(name + "="):
                    return line.split("=", 1)[1].strip().strip('"')
        except Exception:
            pass
    return os.environ.get(name, "")


def _load_model_config() -> dict:
    """读取模型配置，全部有默认值，不写死业务逻辑"""
    _default_text = (
        _read_env("LLM_MODEL")
        or _read_env("MIDSCENE_TEXT_MODEL")
        or _read_env("MIDSCENE_MODEL_NAME")
        or "qwen-plus"
    )
    return {
        "text_model":  _default_text,
        "image_model": _read_env("DASHSCOPE_IMAGE_MODEL") or "wanx2.1-t2i-turbo",
        "image_size":  _read_env("DASHSCOPE_IMAGE_SIZE")  or "1024*1024",
    }


def _urlopen_with_retry(req, timeout: int, retries: int = 3):
    """urlopen with retry on SSL EOF errors (common with macOS LibreSSL)."""
    import ssl, time, urllib.request
    ctx = ssl.create_default_context()
    last_exc = None
    for attempt in range(retries):
        try:
            return urllib.request.urlopen(req, timeout=timeout, context=ctx)
        except Exception as e:
            msg = str(e)
            if "EOF occurred in violation of protocol" in msg or "UNEXPECTED_EOF_WHILE_READING" in msg:
                last_exc = e
                if attempt < retries - 1:
                    time.sleep(1.5 * (attempt + 1))
                    continue
            raise
    raise last_exc


def _generate_draft_llm(topic: str, style_params: dict,
                        key: str, base_url: str, model: str,
                        persona=None, mode: str = "deep_report") -> str:
    """调用 LLM，以分析数据为风格参考，创作小红书文案。返回 markdown 字符串。"""
    import urllib.request
    mode = _normalize_mode(mode)
    mode_cfg = _MODE_OPTIONS[mode]
    # 人设前缀优先；无人设时使用通用描述
    persona_prefix = (persona or {}).get("prompt_prefix", "")
    system = (
        persona_prefix + "\n\n你同时擅长依据风格参考创作有感染力的原创小红书笔记。"
        if persona_prefix
        else "你是专业的小红书内容创作者，擅长依据风格参考创作有感染力的原创笔记。"
    )
    # 人设写作风格附加说明
    persona_style_note = ""
    if persona:
        persona_style_note = (
            f"\n【创作人设风格（最高优先级）】\n"
            f"- 基调：{persona.get('tone', '')}\n"
            f"- 写作风格：{persona.get('writing_style', '')}\n"
            f"- 内容结构偏好：{persona.get('structure', '')}\n"
            f"- Hook 风格：{persona.get('hook_style', '')}\n"
            f"- CTA 风格：{persona.get('cta_style', '')}\n"
        )
    user = f"""请围绕主题「{topic}」创作一篇小红书笔记。
{persona_style_note}
{mode_cfg["instruction"]}

【写作风格参考（来自同类爆款账号分析，仅作补充参考）】
- 目标字数：约 {style_params['avg_words']} 字
- 语言风格：{style_params['style_desc'] or '口语化、亲切自然'}
- 情绪基调：{style_params['emotion'] or '积极正向'}
- 内容结构：{style_params['structure'] or '问题-解法-行动'}
- Hook 风格示例：{style_params['hook_examples'] or '无'}
- CTA 风格示例：{style_params['cta_examples'] or '无'}

【图片风格（供配图参考）】
{style_params['image_style'] or '单图大字信息流'}

【输出格式（严格遵守）】
用 Markdown 输出，包含：
1. 一级标题（笔记标题）
2. 正文（分段，可用 emoji，符合字数要求）
3. 话题标签（# 开头，3~5个，贴合主题）

只输出 Markdown 内容，不要任何额外说明。

补充要求：
- 当前分析模式：{mode_cfg["label"]}
- 模式目标：{mode_cfg["summary"]}
- 确保最终成稿明显符合当前模式，而不是泛泛而谈。"""

    payload = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
        "temperature": 0.85,
        "max_tokens": 2000,
        **({"enable_thinking": False} if "aliyuncs" in base_url or "dashscope" in base_url else {}),
    }).encode()
    req = urllib.request.Request(
        base_url, data=payload,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {key}"},
        method="POST",
    )
    with _urlopen_with_retry(req, timeout=150) as resp:
        data = json.loads(resp.read())
    raw = data["choices"][0]["message"]["content"].strip()
    # strip reasoning blocks from qwen3.5-plus / thinking models
    raw = re.sub(r"<think>[\s\S]*?</think>", "", raw).strip()
    return raw




def _load_llm_key() -> tuple:
    """返回 (api_key, base_url, model)，优先 DashScope，再回退 MIDSCENE"""
    # OPENAI_API_KEY 不使用（FlowLens 运行后会污染 os.environ，导致误走 OpenAI 端点）
    # openai_key = _read_env("OPENAI_API_KEY")
    # if openai_key:
    #     return openai_key, "https://api.openai.com/v1/chat/completions", "gpt-4o-mini"

    dashscope_key = _read_env("DASHSCOPE_API_KEY")
    if dashscope_key:
        return dashscope_key, "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions", "qwen-plus"

    midscene_key = _read_env("MIDSCENE_MODEL_API_KEY")
    if midscene_key:
        base = _read_env("MIDSCENE_MODEL_BASE_URL") or "https://dashscope.aliyuncs.com/compatible-mode/v1"
        if not base.endswith("/chat/completions"):
            base = base.rstrip("/") + "/chat/completions"
        model = _read_env("LLM_MODEL") or _read_env("MIDSCENE_MODEL_NAME") or "qwen-plus"
        return midscene_key, base, model

    return "", "", ""


def chat_with_gpt(draft_content: str, history: list, user_message: str) -> dict:
    """调用 LLM API 对文案进行对话修改，返回 {reply, updated_draft, intent}
    intent: 'image'（用户想调整卡片图片风格）| 'text'（文案修改/提问）
    """
    import urllib.request
    key, base_url, model = _load_llm_key()
    if not key:
        return {"error": "未找到 OPENAI_API_KEY 或 DASHSCOPE_API_KEY，请在 .env 中配置"}
    # 对话用轻量文字模型，避免推理模型的 <think> 延迟
    text_model = _load_model_config().get("text_model") or model

    system_prompt = """你是一位专业的小红书内容运营专家，同时负责判断用户意图。

每次回复必须以一行 JSON 开头，格式如下（单独占一行，之后才是正文）：
{"intent":"text"} 或 {"intent":"image","style":"xxx"}

intent 规则：
- "image"：用户明确要求调整卡片/图片的视觉风格、配色、排版（如"换成蓝色"、"极简风"、"重新生成卡片"）
- "text"：其他所有情况，包括修改文案、提问、闲聊

style 字段：仅 intent=image 时填写，用10字以内描述新风格，必须包含以下关键词之一：红、蓝、绿、橙、紫、黑、深色、极简、暖、清新、粉

文案修改规则：
1. 保持小红书风格：口语化、emoji 点缀、分点清晰
2. 保留原文案的核心信息和主题方向
3. 修改文案时先给出修改说明（1-3句话），再给出完整修改后的文案（用 ```markdown ... ``` 包裹）
4. 纯提问直接回答，不需要输出完整文案"""

    messages = [{"role": "system", "content": system_prompt}]
    if not history:
        messages.append({
            "role": "user",
            "content": f"以下是当前文案草稿：\n\n```\n{draft_content}\n```\n\n{user_message}"
        })
    else:
        for h in history:
            messages.append({"role": h["role"], "content": h["content"]})
        messages.append({"role": "user", "content": user_message})

    payload = json.dumps({
        "model": text_model,
        "messages": messages,
        "temperature": 0.7,
        "max_tokens": 2000,
        **({"enable_thinking": False} if "aliyuncs" in base_url or "dashscope" in base_url else {}),
    }).encode()

    req = urllib.request.Request(
        base_url,
        data=payload,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {key}"},
        method="POST",
    )
    try:
        with _urlopen_with_retry(req, timeout=60) as resp:
            data = json.loads(resp.read())
        raw = data["choices"][0]["message"]["content"]
        raw = re.sub(r"<think>[\s\S]*?</think>", "", raw).strip()

        # 解析第一行 JSON 意图
        intent = "text"
        style = ""
        lines = raw.split("\n", 1)
        reply = raw
        try:
            meta = json.loads(lines[0].strip())
            intent = meta.get("intent", "text")
            style = meta.get("style", "")
            reply = lines[1].strip() if len(lines) > 1 else ""
        except Exception:
            pass  # 没有 JSON 头，当 text 处理

        # 提取 ```markdown 块作为 updated_draft
        m = re.search(r"```(?:markdown)?\s*([\s\S]+?)```", reply)
        updated = m.group(1).strip() if m else None
        return {"reply": reply, "updated_draft": updated, "intent": intent, "style": style}
    except Exception as e:
        return {"error": str(e)}


# ── HTTP Handler ──────────────────────────────────────────────────────────────

class Handler(BaseHTTPRequestHandler):
    def log_message(self, *a): pass

    def do_GET(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        qs = parse_qs(parsed.query)

        if path == "/":
            self._html()
        elif path == "/api/state":
            self._json(state)
        elif path == "/api/run":
            keyword = qs.get("keyword", ["海外求职"])[0]
            mode = _normalize_mode(qs.get("mode", ["deep_report"])[0])
            retrieval_account = qs.get("retrieval_account", [""])[0].strip()
            try:
                accounts         = max(1, min(20, int(qs.get("accounts",          [5])[0])))
                notes_per_account = max(1, min(10, int(qs.get("notes_per_account", [3])[0])))
                copies           = max(1, min(5,  int(qs.get("copies",            [2])[0])))
            except (ValueError, TypeError):
                self._json({"error": "参数非法"}); return
            if state["status"] == "running":
                self._json({"started": False, "reason": "already_running"})
            else:
                Thread(
                    target=run_pipeline,
                    args=(keyword, accounts, copies, retrieval_account, notes_per_account, mode),
                    daemon=True,
                ).start()
                self._json({
                    "started": True,
                    "retrieval_account": retrieval_account,
                    "mode": mode,
                    "mode_label": _mode_label(mode),
                })
        elif path == "/api/open_output":
            target = ANALYSIS_DIR if ANALYSIS_DIR.exists() else OUTPUT_DIR
            subprocess.run(["open", str(target)], check=False)
            self._json({"ok": True, "target": str(target)})
        elif path == "/api/log_path":
            logs = sorted(LOG_DIR.glob("server_*.log"), key=lambda p: p.stat().st_mtime, reverse=True)
            self._json({
                "log_path": str(logs[0]) if logs else "",
                "log_dir": str(LOG_DIR),
            })
        elif path == "/api/draft":
            # 读取单个文案文件内容
            fpath = qs.get("file", [""])[0]
            try:
                resolved = Path(fpath).resolve()
                if not resolved.is_relative_to(DRAFT_DIR.resolve()):
                    self.send_response(403)
                    self.end_headers()
                    return
                content = resolved.read_text(encoding="utf-8")
                saved_images = _get_saved_images_for_draft(resolved)
                self._json({
                    "content": content,
                    "file": fpath,
                    "cards": saved_images,
                })
            except Exception as e:
                self._json({"error": str(e)})
        elif path == "/api/drafts":
            files = sorted(DRAFT_DIR.glob("**/*.md"), key=lambda p: p.stat().st_mtime, reverse=True)
            result = [_build_draft_summary(f) for f in files]
            self._json(result)
        elif path.startswith("/api/card_task/"):
            tid = path.split("/api/card_task/", 1)[1]
            self._json(_card_tasks.get(tid, {"status": "not_found"}))
        elif path == "/api/personas":
            personas = persona_manager.list_personas()
            active_id = persona_manager.get_active_persona_id()
            self._json({"personas": personas, "active": active_id})
        elif path == "/api/manage_accounts":
            self._json(_list_managed_accounts())
        elif path.startswith("/api/login_task/"):
            tid = path.split("/api/login_task/", 1)[1]
            self._json(_login_tasks.get(tid, {"status": "not_found"}))
        elif path == "/api/accounts":
            self._json(publish_manager.list_accounts())
        elif path == "/api/ready_drafts":
            self._json(publish_manager.list_ready_drafts())
        elif path == "/api/scheduled_tasks":
            self._json(publish_manager.get_scheduled_tasks())
        elif path.startswith("/output/draft_cards/"):
            # 服务 DRAFT_DIR 下的卡片 PNG
            # /output/draft_cards/{stem}/{file} → DRAFT_DIR/{stem}/{file}
            rel = path[len("/output/draft_cards/"):]
            card_path = (DRAFT_DIR / rel).resolve()
            if not card_path.is_relative_to(DRAFT_DIR.resolve()):
                self.send_response(403); self.end_headers()
            elif card_path.exists() and card_path.suffix.lower() == ".png":
                self.send_response(200)
                self.send_header("Content-Type", "image/png")
                self.send_header("Content-Length", str(card_path.stat().st_size))
                self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
                self.end_headers()
                self.wfile.write(card_path.read_bytes())
            else:
                self.send_response(404); self.end_headers()
        else:
            self.send_response(404); self.end_headers()

    def do_POST(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        length = int(self.headers.get("Content-Length", 0))
        try:
            body = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
        except Exception:
            self._json({"error": "请求格式错误（JSON解析失败）"}); return

        if path == "/api/account_login":
            name = body.get("name", "")
            mode = body.get("mode", "scan")
            try:
                self._json(_start_account_login_task(name, mode))
            except Exception as e:
                self._json({"error": str(e)})
        elif path == "/api/manage_accounts/add_cookie":
            name = body.get("name", "")
            cookie = body.get("cookie", "")
            overwrite = bool(body.get("overwrite", False))
            try:
                account = account_manager.save_cookie_account(name, cookie, overwrite=overwrite)
                self._json({"ok": True, "account": account})
            except Exception as e:
                self._json({"error": str(e)})
        elif path == "/api/manage_accounts/toggle":
            name = body.get("name", "")
            active = bool(body.get("active", True))
            try:
                account = account_manager.set_account_active(name, active)
                self._json({"ok": True, "account": account})
            except Exception as e:
                self._json({"error": str(e)})
        elif path == "/api/manage_accounts/nickname":
            name = body.get("name", "")
            nickname = (body.get("nickname", "") or "")[:100]
            try:
                account = account_manager.update_account_nickname(name, nickname)
                self._json({"ok": True, "account": account})
            except Exception as e:
                self._json({"error": str(e)})
        elif path == "/api/chat":
            draft_content = body.get("draft_content", "")
            history       = body.get("history", [])
            message       = body.get("message", "")
            draft_path    = body.get("draft_path", "")
            image_style   = body.get("image_style", "")
            topic         = body.get("topic", "")
            if not message:
                self._json({"error": "message 不能为空"})
                return

            # LLM 同时判断意图 + 处理文案
            result = chat_with_gpt(draft_content, history, message)
            if "error" in result:
                self._json(result); return

            intent = result.get("intent", "text")

            if intent == "image":
                if not draft_path:
                    self._json({"reply": result.get("reply", "请先选择一篇文案再生成卡片"), "intent": "card"}); return
                resolved = Path(draft_path).resolve()
                if not resolved.is_relative_to(DRAFT_DIR.resolve()):
                    self._json({"error": "非法路径", "intent": "card"}); return
                new_style = result.get("style", "")
                import uuid as _uuid
                task_id = _uuid.uuid4().hex
                _card_tasks[task_id] = {"status": "running"}

                def _bg_gen(resolved=resolved, new_style=new_style, task_id=task_id):
                    try:
                        with _get_draft_lock(resolved):
                            content = resolved.read_text(encoding="utf-8")
                            style_m = re.search(r'(\*\*图片风格\*\*[：:]\s*)([^\n]+)', content)
                            # 用 LLM 已给出的 style，没有再保留原值
                            final_style = new_style or (style_m.group(2).strip() if style_m else "简洁信息流")
                            if style_m:
                                updated = content[:style_m.start(2)] + final_style + content[style_m.end(2):]
                            else:
                                sep = content.find('\n---\n')
                                ins = sep if sep >= 0 else 0
                                updated = content[:ins] + f"\n**图片风格**：{final_style}" + content[ins:]
                            resolved.write_text(updated, encoding="utf-8")
                        url_paths = _gen_cards_for_draft(resolved)
                        _card_tasks[task_id] = {
                            "status": "done",
                            "cards": url_paths,
                            "reply": f"图片风格已更新为「{final_style}」，卡片已重新生成，共 {len(url_paths)} 张",
                        }
                    except Exception as e:
                        _card_tasks[task_id] = {"status": "error", "error": str(e)}

                Thread(target=_bg_gen, daemon=True).start()
                self._json({"intent": "card_task", "task_id": task_id, "reply": result.get("reply", "卡片生成中，请稍候…")})
            else:
                result["intent"] = "text"
                self._json(result)
        elif path == "/api/save_draft":
            fpath = body.get("file", "")
            content = body.get("content", "")
            image_path = body.get("image_path") if "image_path" in body else None
            image_paths = body.get("image_paths") if "image_paths" in body else None
            if not fpath:
                self._json({"error": "缺少 file 参数"}); return
            if not content.strip():
                self._json({"error": "文案内容为空"}); return
            try:
                resolved = Path(fpath).resolve()
                if not resolved.is_relative_to(DRAFT_DIR.resolve()):
                    self._json({"error": "路径非法"}); return
                # ── 如果新内容缺少人设字段，从原文件恢复元数据头 ──────────────
                # AI 编辑文案时只返回正文 markdown，元数据（**人设**等）会丢失
                if resolved.exists() and not re.search(r'\*\*人设\*\*[：:]', content):
                    try:
                        old_text = resolved.read_text(encoding="utf-8")
                        old_sep = old_text.find('\n---\n')
                        if old_sep >= 0:
                            old_meta = old_text[:old_sep]
                            if re.search(r'\*\*人设\*\*[：:]', old_meta):
                                new_sep = content.find('\n---\n')
                                if new_sep >= 0:
                                    # 有分隔符：用原元数据替换新内容的元数据块
                                    content = old_meta + content[new_sep:]
                                else:
                                    # 无分隔符：整个内容当正文，前置原元数据
                                    content = old_meta + "\n\n---\n\n" + content
                    except Exception:
                        pass
                old_resolved = resolved
                # 从内容提取标题，生成新文件名
                title_m = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
                if title_m:
                    safe_title = re.sub(r'[\\/:*?"<>|\s]', "_", title_m.group(1).strip())[:30]
                    date_str = datetime.now().strftime("%Y%m%d")
                    new_name = f"{safe_title}_{date_str}.md"
                    new_path = resolved.parent / new_name
                    if new_path != resolved and not new_path.exists():
                        resolved.rename(new_path)
                        resolved = new_path
                        publish_manager.move_pub_json(str(old_resolved), str(resolved))
                resolved.write_text(content, encoding="utf-8")
                if image_path is not None or image_paths is not None:
                    publish_manager.save_draft_metadata(
                        str(resolved),
                        image_path=image_path,
                        image_paths=image_paths,
                    )
                self._json({"ok": True, "new_path": str(resolved)})
            except Exception as e:
                self._json({"error": str(e)})
        elif path == "/api/mark_ready":
            md_path = body.get("md_path", "")
            title = body.get("title", "")
            desc = body.get("desc", "")
            image_path = body.get("image_path", "")
            image_paths = body.get("image_paths", [])
            account = body.get("account", "")
            image_style = body.get("image_style", "")
            writing_style = body.get("writing_style", "")
            if not md_path:
                self._json({"error": "md_path 不能为空"})
                return
            try:
                pub = publish_manager.mark_ready(
                    md_path=md_path,
                    title=title,
                    desc=desc,
                    image_path=image_path,
                    image_paths=image_paths,
                    account=account,
                    image_style=image_style,
                    writing_style=writing_style,
                )
                self._json(pub)
            except Exception as e:
                self._json({"error": str(e)})
        elif path == "/api/publish_batch":
            items = body.get("items", [])
            scheduled_at = body.get("scheduled_at", "")
            if not items:
                self._json({"error": "items 不能为空"})
                return
            if scheduled_at:
                ready_drafts = {d["md_path"]: d for d in publish_manager.list_ready_drafts()}
                task_items = []
                for item in items:
                    draft_path = item.get("draft_path", "")
                    draft = ready_drafts.get(draft_path)
                    if not draft:
                        # 文案不存在或不处于 ready 状态，拒绝该请求
                        self._json({"error": f"文案未找到或不处于 ready 状态: {draft_path}"})
                        return
                    task_items.append({
                        "draft_path": draft_path,
                        "account": item.get("account", ""),
                        "title": draft.get("title", ""),
                    })
                task = publish_manager.save_scheduled_task(
                    scheduled_at=scheduled_at,
                    items=task_items,
                )
                print(f"[publish_batch] 定时发布任务已保存: {len(task_items)} 项，"
                      f"计划于 {scheduled_at} 执行，任务ID: {task['id']}", flush=True)
                self._json({"ok": True, "task": task})
            else:
                print(f"[publish_batch] 立即发布 {len(items)} 项", flush=True)
                try:
                    result = publish_manager.do_publish_batch(items)
                except Exception as e:
                    import traceback; traceback.print_exc()
                    self._json({"ok": False, "error": str(e)}); return
                print(f"[publish_batch] 完成: {result}", flush=True)
                # 检查是否有 item 级别的错误
                errors = [f"{k}: {v['error']}" for k, v in result.items() if v.get("status") == "error"]
                if errors:
                    self._json({"ok": False, "error": "\n".join(errors)})
                else:
                    self._json({"ok": True, "result": result})
        elif path == "/api/redraft":
            # 从最新 xlsx 加载账号数据，直接跑文案生成，无需重新 explore/analyze
            keyword  = body.get("keyword", "") or state.get("keyword", "")
            try:
                copies = max(1, min(5, int(body.get("copies", 2))))
            except (ValueError, TypeError):
                self._json({"error": "copies 参数非法，必须为 1-5 的整数"}); return
            if state["status"] == "running":
                self._json({"error": "pipeline 正在运行，请等待完成"}); return
            xlsx_files = []
            if ANALYSIS_XLSX.exists():
                xlsx_files.append(ANALYSIS_XLSX)
            xlsx_files.extend(
                sorted(
                    list(OUTPUT_DIR.glob("xhs_account_analysis_*.xlsx")) +
                    list(ANALYSIS_DIR.glob("xhs_account_analysis_*.xlsx")),
                    key=lambda p: p.stat().st_mtime, reverse=True
                )
            )
            if not xlsx_files:
                self._json({"error": "未找到分析结果 xlsx"}); return
            import openpyxl as _opx
            wb = _opx.load_workbook(xlsx_files[0])
            ws = wb.active
            hdrs = [c.value for c in ws[1]]
            accs = [dict(zip(hdrs, r)) for r in ws.iter_rows(min_row=2, values_only=True)
                    if r and r[0]]
            if not accs:
                self._json({"error": "xlsx 无账号数据"}); return
            if not keyword:
                keyword = str(accs[0].get("搜索关键词") or "小红书内容")
            state["accounts"] = accs
            state["keyword"]  = keyword
            def _do_redraft():
                state["status"] = "running"; state["step"] = 4
                add_log(f"重新生成文案（{len(accs)} 账号 × {copies} 篇，关键词：{keyword}）", "info")
                try:
                    _run_draft_step(keyword, copies)
                except Exception as e:
                    add_log(f"文案生成失败: {e}", "error")
                finally:
                    state["status"] = "completed"; state["step"] = 8
            Thread(target=_do_redraft, daemon=True).start()
            self._json({"started": True, "accounts": len(accs), "keyword": keyword})
        elif path == "/api/gen_card":
            draft_path_str = body.get("draft_path", "")
            if not draft_path_str:
                self._json({"error": "需要 draft_path"}); return
            try:
                resolved = Path(draft_path_str).resolve()
                if not resolved.is_relative_to(DRAFT_DIR.resolve()):
                    self._json({"error": "路径非法"}); return
            except Exception:
                self._json({"error": "路径解析失败"}); return
            try:
                force_tpl = body.get("card_template", "")
                url_paths = _gen_cards_for_draft(resolved, force_template=force_tpl)
                self._json({"cards": url_paths})
            except Exception as e:
                self._json({"error": str(e)})
        elif path == "/api/personas/active":
            pid = (body or {}).get("id", "").strip()
            if not pid:
                self._json({"error": "缺少 id 字段"}); return
            if not persona_manager.get_persona_by_id(pid):
                self._json({"error": f"未知人设 id: {pid}"}); return
            persona_manager.set_active_persona(pid)
            p = persona_manager.get_active_persona()
            self._json({"ok": True, "active": pid, "name": p["name"]})
        else:
            self.send_response(404); self.end_headers()

    def do_DELETE(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        if path.startswith("/api/card_task/"):
            tid = path.split("/api/card_task/", 1)[1]
            _card_tasks.pop(tid, None)
            self._json({"ok": True})
        elif path.startswith("/api/login_task/"):
            tid = path.split("/api/login_task/", 1)[1]
            _login_tasks.pop(tid, None)
            self._json({"ok": True})
        elif path.startswith("/api/manage_accounts/"):
            name = path.split("/api/manage_accounts/", 1)[1]
            try:
                account_manager.delete_account(name, remove_storage=True)
                self._json({"ok": True})
            except Exception as e:
                self._json({"error": str(e)})
        elif path.startswith("/api/scheduled_tasks/"):
            task_id = path.split("/api/scheduled_tasks/", 1)[1]
            ok = publish_manager.cancel_task(task_id)
            self._json({"ok": ok})
        else:
            self.send_response(404); self.end_headers()

    def _json(self, data):
        body = json.dumps(data, ensure_ascii=False, default=str).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _html(self):
        body = HTML.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)


# ── HTML ──────────────────────────────────────────────────────────────────────

HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>小红书自动运营系统</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'PingFang SC','Segoe UI',sans-serif;
  background:linear-gradient(135deg,#ff6b6b 0%,#ff8e53 100%);min-height:100vh;padding:24px}
.wrap{max-width:1280px;margin:0 auto}
h1{color:#fff;font-size:2rem;font-weight:900;text-align:center;margin-bottom:6px;
  text-shadow:0 2px 12px rgba(0,0,0,.2)}
.sub{color:rgba(255,255,255,.85);text-align:center;margin-bottom:24px;font-size:1rem}

/* Persona switcher */
.persona-bar{background:#fff;border-radius:14px;padding:14px 20px;
  box-shadow:0 4px 16px rgba(0,0,0,.08);margin-bottom:16px;
  display:flex;align-items:center;gap:14px;flex-wrap:wrap}
.persona-label{font-size:.82rem;font-weight:700;color:#888;white-space:nowrap}
.persona-chips{display:flex;gap:10px;flex-wrap:wrap}
.persona-chip{padding:8px 16px;border-radius:20px;border:2px solid #eee;
  background:#fafafa;cursor:pointer;font-size:.85rem;font-weight:600;
  transition:all .18s;display:flex;align-items:center;gap:5px;line-height:1.3}
.persona-chip .p-desc{font-size:.72rem;font-weight:400;opacity:.7;margin-left:2px}
.persona-chip:hover{border-color:#ff8e53;color:#ff6b6b;background:#fff5f0}
.persona-chip.active{background:linear-gradient(135deg,#ff6b6b,#ff8e53);
  color:#fff;border-color:transparent;box-shadow:0 4px 14px rgba(255,107,107,.35)}
.persona-chip.active .p-desc{opacity:.85}
.tpl-chip{padding:5px 12px;border-radius:14px;border:1.5px solid #ddd;
  background:#fafafa;cursor:pointer;font-size:.78rem;font-weight:600;
  transition:all .15s;white-space:nowrap}
.tpl-chip:hover{border-color:#ff8e53;color:#ff6b6b;background:#fff5f0}
.tpl-chip.active{background:linear-gradient(135deg,#ff6b6b,#ff8e53);
  color:#fff;border-color:transparent;box-shadow:0 3px 10px rgba(255,107,107,.3)}

/* Control card */
.ctrl{background:#fff;border-radius:16px;padding:24px;box-shadow:0 8px 32px rgba(0,0,0,.12);margin-bottom:20px}
.ctrl-row{display:flex;gap:12px;align-items:center;flex-wrap:wrap}
.inp{flex:1;min-width:200px;padding:12px 16px;border:2px solid #eee;border-radius:10px;
  font-size:1rem;outline:none;transition:border-color .2s}
.inp:focus{border-color:#ff6b6b}
.num-inp{width:80px;padding:12px;border:2px solid #eee;border-radius:10px;
  font-size:1rem;text-align:center;outline:none}
.num-inp:focus{border-color:#ff6b6b}
.sel{padding:12px 14px;border:2px solid #eee;border-radius:10px;background:#fff;
  font-size:.96rem;outline:none;min-width:170px;color:#333}
.sel:focus{border-color:#ff6b6b}
.btn{padding:12px 28px;background:linear-gradient(135deg,#ff6b6b,#ff8e53);color:#fff;
  border:none;border-radius:10px;font-size:1rem;font-weight:700;cursor:pointer;
  transition:opacity .2s;white-space:nowrap}
.btn:disabled{opacity:.5;cursor:not-allowed}
.btn-outline{background:#fff;color:#ff6b6b;border:2px solid #ff6b6b}

/* Steps */
.steps{display:flex;gap:0;margin:20px 0 0;overflow-x:auto}
.step{flex:1;min-width:80px;text-align:center;position:relative}
.step-dot{width:32px;height:32px;border-radius:50%;background:#f0f0f0;color:#999;
  font-size:.75rem;font-weight:700;display:flex;align-items:center;justify-content:center;
  margin:0 auto 6px;transition:all .3s;position:relative;z-index:1}
.step-dot.done{background:#4caf50;color:#fff}
.step-dot.active{background:linear-gradient(135deg,#ff6b6b,#ff8e53);color:#fff;
  box-shadow:0 4px 12px rgba(255,107,107,.4)}
.step-label{font-size:.7rem;color:#999;line-height:1.3}
.step-label.active{color:#ff6b6b;font-weight:700}
.step:not(:last-child)::after{content:'';position:absolute;top:16px;left:50%;
  width:100%;height:2px;background:#f0f0f0;z-index:0}
.step:not(:last-child).done::after{background:#4caf50}

/* Grid */
.grid{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}
.card{background:#fff;border-radius:16px;padding:20px;box-shadow:0 4px 20px rgba(0,0,0,.08)}
.card-title{font-size:1rem;font-weight:700;color:#333;margin-bottom:14px;
  display:flex;align-items:center;gap:8px}

/* Logs */
.logs{height:300px;overflow-y:auto;font-family:'SF Mono','Fira Code',monospace;font-size:.8rem;
  background:#1a1a2e;border-radius:10px;padding:14px;display:flex;flex-direction:column;gap:4px}
.log{padding:3px 6px;border-radius:4px;line-height:1.5}
.log-info{color:#8892b0}
.log-ok{color:#64ffda;background:rgba(100,255,218,.06)}
.log-error{color:#ff6b6b;background:rgba(255,107,107,.08)}
.log-warn{color:#ffd700;background:rgba(255,215,0,.06)}
.log-debug{color:#5a6480}
.log-time{opacity:.5;margin-right:6px}

/* Stats */
.stat-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.stat{background:linear-gradient(135deg,#fff5f5,#fff);border:1.5px solid #ffe0e0;
  border-radius:12px;padding:16px;text-align:center}
.stat-val{font-size:2rem;font-weight:900;color:#ff6b6b;line-height:1}
.stat-lbl{font-size:.8rem;color:#999;margin-top:6px}

/* Accounts table */
.tbl{width:100%;border-collapse:collapse;font-size:.85rem}
.tbl th{background:#fff8f5;padding:10px 12px;text-align:left;border-bottom:2px solid #ffe0e0;
  color:#ff6b6b;font-weight:700;white-space:nowrap}
.tbl td{padding:10px 12px;border-bottom:1px solid #f5f5f5;color:#444;vertical-align:top}
.tbl tr:hover td{background:#fff8f5}
.badge{display:inline-block;padding:2px 8px;border-radius:20px;font-size:.75rem;font-weight:600}
.badge-a{background:#e8f5e9;color:#2e7d32}
.badge-b{background:#fff3e0;color:#e65100}
.tag{display:inline-block;background:#fff0f0;color:#ff6b6b;border-radius:4px;
  padding:1px 6px;font-size:.75rem;margin:1px}

/* Drafts */
.draft-list{display:flex;flex-direction:column;gap:10px;max-height:320px;overflow-y:auto}
.draft-item{background:#fff8f5;border:1.5px solid #ffe0e0;border-radius:10px;padding:14px 16px;cursor:pointer;transition:box-shadow .15s,border-color .15s}
.draft-item:hover{border-color:#ff6b6b;box-shadow:0 2px 10px rgba(255,107,107,.15)}
.draft-title{font-weight:700;color:#333;font-size:.9rem;margin-bottom:4px}
.draft-meta{font-size:.78rem;color:#999}

/* Status */
.status-bar{display:flex;align-items:center;gap:10px;padding:12px 16px;border-radius:10px;
  margin-bottom:16px;font-weight:600}
.status-idle{background:#f5f5f5;color:#999}
.status-running{background:#fff3e0;color:#e65100}
.status-completed{background:#e8f5e9;color:#2e7d32}
.status-error{background:#ffebee;color:#c62828}
.pulse{width:10px;height:10px;border-radius:50%;background:currentColor;
  animation:pulse 1s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.3}}

@media(max-width:768px){.grid{grid-template-columns:1fr}.ctrl-row{flex-direction:column}.stat-grid{grid-template-columns:repeat(2,1fr)}}
/* Publish section */
.pub-btn{padding:10px 24px;border:none;border-radius:10px;font-size:.95rem;font-weight:700;
  cursor:pointer;transition:opacity .2s}
.pub-btn-now{background:linear-gradient(135deg,#ff6b6b,#ff8e53);color:#fff}
.pub-btn-sched{background:linear-gradient(135deg,#667eea,#764ba2);color:#fff}
.pub-btn:disabled{opacity:.5;cursor:not-allowed}
.modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:100;
  align-items:center;justify-content:center}
.modal-overlay.open{display:flex}
.modal{background:#fff;border-radius:20px;padding:28px 32px;width:480px;max-width:95vw;
  max-height:90vh;overflow-y:auto;box-shadow:0 20px 60px rgba(0,0,0,.2)}
.modal h2{font-size:1.1rem;font-weight:900;color:#333;margin-bottom:20px}
.modal label{display:block;font-size:.85rem;color:#666;margin-bottom:4px;margin-top:14px}
.modal input,.modal select,.modal textarea{width:100%;padding:10px 12px;border:1.5px solid #eee;
  border-radius:8px;font-size:.9rem;outline:none;font-family:inherit}
.modal input:focus,.modal select:focus,.modal textarea:focus{border-color:#ff6b6b}
.modal textarea{resize:vertical;min-height:72px}
.char-count{font-size:.75rem;color:#aaa;text-align:right;margin-top:2px}
.char-count.over{color:#e53935}
.account-checkboxes{border:1.5px solid #eee;border-radius:8px;padding:10px 14px;
  max-height:140px;overflow-y:auto;display:flex;flex-direction:column;gap:6px}
.account-checkboxes label{font-size:.88rem;color:#333;display:flex;align-items:center;
  gap:8px;cursor:pointer;margin:0}
.acc-expired{color:#aaa}
.modal-footer{display:flex;justify-content:flex-end;gap:10px;margin-top:20px}
.btn-cancel-modal{padding:10px 22px;background:#f5f5f5;color:#666;border:none;
  border-radius:8px;font-size:.9rem;cursor:pointer}
.btn-confirm{padding:10px 22px;background:linear-gradient(135deg,#ff6b6b,#ff8e53);
  color:#fff;border:none;border-radius:8px;font-size:.9rem;font-weight:700;cursor:pointer}
.btn-confirm:disabled{opacity:.5;cursor:not-allowed}
.toast{position:fixed;top:20px;left:50%;transform:translateX(-50%);padding:12px 28px;
  border-radius:10px;font-weight:700;font-size:.95rem;z-index:200;
  box-shadow:0 4px 20px rgba(0,0,0,.15);display:none}
.toast.show{display:block}
.toast-ok{background:#e8f5e9;color:#2e7d32}
.toast-err{background:#ffebee;color:#c62828}
.sched-tbl{width:100%;border-collapse:collapse;font-size:.85rem;margin-top:12px}
.sched-tbl th{background:#f9f0ff;padding:9px 12px;text-align:left;
  border-bottom:2px solid #e0d0f0;color:#764ba2;font-weight:700;white-space:nowrap}
.sched-tbl td{padding:9px 12px;border-bottom:1px solid #f5f5f5;color:#444}
.sched-tbl tr:hover td{background:#faf5ff}
.badge-pending{background:#fff3e0;color:#e65100;padding:2px 8px;border-radius:20px;
  font-size:.75rem;font-weight:600}
.badge-done{background:#e8f5e9;color:#2e7d32;padding:2px 8px;border-radius:20px;
  font-size:.75rem;font-weight:600}
.badge-failed{background:#ffebee;color:#c62828;padding:2px 8px;border-radius:20px;
  font-size:.75rem;font-weight:600}
.badge-cancelled{background:#f5f5f5;color:#999;padding:2px 8px;border-radius:20px;
  font-size:.75rem;font-weight:600}
.acct-toolbar{display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;margin-bottom:14px}
.acct-toolbar .field{display:flex;flex-direction:column;gap:4px;min-width:220px;flex:1}
.acct-toolbar label{font-size:.8rem;color:#777}
.acct-inline-actions{display:flex;gap:8px;flex-wrap:wrap}
.acct-status{font-size:.82rem;color:#666;padding:8px 10px;border-radius:8px;background:#faf7f2;margin-bottom:12px}
.acct-status.running{background:#fff4e5;color:#b85c00}
.acct-status.done{background:#e8f5e9;color:#2e7d32}
.acct-status.error{background:#ffebee;color:#c62828}
.acct-table{width:100%;border-collapse:collapse;font-size:.85rem}
.acct-table th{background:#f7f1ea;padding:10px 12px;text-align:left;border-bottom:2px solid #eadfd3;color:#7a5c42}
.acct-table td{padding:10px 12px;border-bottom:1px solid #f2ede7;color:#444;vertical-align:top}
.acct-table tr:hover td{background:#fcfaf8}
.acct-pill{display:inline-block;padding:3px 8px;border-radius:999px;font-size:.74rem;font-weight:700}
.acct-pill-on{background:#e8f5e9;color:#2e7d32}
.acct-pill-off{background:#f5f5f5;color:#888}
.acct-pill-valid{background:#e3f2fd;color:#1565c0}
.acct-pill-invalid{background:#fff3e0;color:#e65100}
.acct-ops{display:flex;gap:6px;flex-wrap:wrap}
.acct-name-input{width:100%;padding:8px 10px;border:1.5px solid #e8dfd5;border-radius:8px;font-size:.84rem;outline:none}
.acct-name-input:focus{border-color:#ff8e53}
.acct-name-hint{font-size:.72rem;color:#999;margin-top:4px}
.pub-assign-row{display:grid;grid-template-columns:minmax(130px,160px) minmax(0,1fr) auto;gap:10px;align-items:start;padding:10px;border:1px solid #eee;border-radius:10px}
.pub-assign-row.disabled{opacity:.45;background:#fafafa}
.pub-assign-account{display:flex;align-items:flex-start;gap:8px;min-width:0;padding-top:2px}
.pub-assign-account input{margin-top:2px;flex:0 0 auto}
.pub-assign-name{line-height:1.45;word-break:break-word}
.pub-assign-chips{display:flex;flex-wrap:wrap;gap:8px;min-height:32px;align-content:flex-start}
.pub-assign-action{padding:6px 10px;font-size:.8rem;align-self:start}
@media(max-width:768px){
  .pub-assign-row{grid-template-columns:1fr}
  .pub-assign-action{width:100%}
}
</style>
</head>
<body>
<div class="wrap">
  <h1>🍠 小红书自动运营系统</h1>

  <!-- 人设切换 -->
  <div class="persona-bar">
    <span class="persona-label">创作人设</span>
    <div class="persona-chips" id="personaChips">
      <span style="color:#bbb;font-size:.83rem">加载中…</span>
    </div>
  </div>
  <p class="sub">关键词 → 爆款分析 → 文案生成 → 图文卡片</p>

  <!-- 控制面板 -->
  <div class="ctrl">
    <div class="ctrl-row">
      <input id="kw" class="inp" value="海外求职" placeholder="输入关键词...">
      <div style="display:flex;align-items:center;gap:8px;white-space:nowrap">
        <span style="font-size:.85rem;color:#666">输出账号数</span>
        <input id="accCount" class="num-inp" type="number" value="5" min="1" max="10">
      </div>
      <div style="display:flex;align-items:center;gap:8px;white-space:nowrap">
        <span style="font-size:.85rem;color:#666">文案篇数/账号</span>
        <input id="copies" class="num-inp" type="number" value="2" min="1" max="5">
      </div>
      <div style="display:flex;align-items:center;gap:8px;white-space:nowrap">
        <span style="font-size:.85rem;color:#666">分析模式</span>
        <select id="analysisMode" class="sel">
          <option value="deep_report">通用深度分析</option>
          <option value="topic_ideas">选题研究</option>
          <option value="ops_breakdown">运营拆解</option>
          <option value="experience_summary">经验总结</option>
        </select>
      </div>
      <!-- 检索账号选择器（暂停使用，默认不指定账号以加快检索速度，后期优化再启用）
      <div style="display:flex;align-items:center;gap:8px;white-space:nowrap">
        <span style="font-size:.85rem;color:#666">检索账号</span>
        <select id="retrievalAccountSel" style="padding:10px 12px;border:1.5px solid #eee;border-radius:12px;font-size:.92rem;min-width:170px;background:#fff">
          <option value="">自动轮询</option>
        </select>
      </div>
      -->
      <button id="runBtn" class="btn" onclick="startPipeline()">🚀 开始运营</button>
      <button class="btn btn-outline" onclick="openOutput()">📊 打开分析表格</button>
    </div>

    <!-- 步骤条 -->
    <div class="steps" id="stepsBar">
      <div class="step" id="s1"><div class="step-dot" id="sd1">1</div><div class="step-label" id="sl1">初始化</div></div>
      <div class="step" id="s2"><div class="step-dot" id="sd2">2</div><div class="step-label" id="sl2">Explore<br>抓取</div></div>
      <div class="step" id="s3"><div class="step-dot" id="sd3">3</div><div class="step-label" id="sl3">账号<br>分析</div></div>
      <div class="step" id="s4"><div class="step-dot" id="sd4">4</div><div class="step-label" id="sl4">文案<br>生成</div></div>
      <div class="step" id="s5"><div class="step-dot" id="sd5">5</div><div class="step-label" id="sl5">图文<br>卡片</div></div>
      <div class="step" id="s6"><div class="step-dot" id="sd6">✓</div><div class="step-label" id="sl6">完成</div></div>
    </div>
  </div>

  <!-- 状态条 -->
  <div id="statusBar" class="status-bar status-idle">
    <div class="pulse" id="pulse" style="display:none"></div>
    <span id="statusText">空闲，等待任务</span>
  </div>

  <!-- 主网格 -->
  <div class="grid">
    <!-- 日志 -->
    <div class="card">
      <div class="card-title">📋 实时日志</div>
      <div class="logs" id="logsBox"></div>
    </div>

    <!-- 统计 -->
    <div class="card">
      <div class="card-title">📊 统计</div>
      <div class="stat-grid">
        <div class="stat"><div class="stat-val" id="sv1">—</div><div class="stat-lbl">分析账号</div></div>
        <div class="stat"><div class="stat-val" id="sv2">—</div><div class="stat-lbl">生成文案</div></div>
        <div class="stat"><div class="stat-val" id="sv3">—</div><div class="stat-lbl">图文卡片</div></div>
      </div>
    </div>
  </div>

  <!-- 账号分析结果 -->
  <div class="card" id="accCard" style="display:none;margin-bottom:16px">
    <div class="card-title">🏆 账号分析结果</div>
    <div style="overflow-x:auto">
      <table class="tbl">
        <thead><tr>
          <th>#</th><th>账号昵称</th><th>粉丝数</th><th>权重评级</th>
          <th>内容偏好</th><th>Hook模板</th><th>CTA模板</th>
        </tr></thead>
        <tbody id="accTbody"></tbody>
      </table>
    </div>
  </div>

  <!-- 备选文案 -->
  <div class="card" id="draftCard" style="display:none;margin-bottom:16px">
    <div class="card-title">✍️ 备选文案</div>
    <div class="draft-list" id="draftList"></div>
  </div>

  <!-- 文案对话修改 -->
  <div class="card" style="margin-bottom:16px">
    <div class="card-title">💬 文案对话修改 <span style="font-size:.78rem;font-weight:400;color:#aaa">（接入 ChatGPT，对文案进行对话式优化）</span></div>
    <div style="display:flex;gap:16px;flex-wrap:wrap">
      <!-- 左：文案选择+预览 -->
      <div style="flex:1;min-width:280px">
        <div style="display:flex;gap:8px;margin-bottom:10px;align-items:center">
          <select id="draftSel" style="flex:1;padding:8px 12px;border:1.5px solid #eee;border-radius:8px;font-size:.9rem;outline:none" onchange="loadDraft()">
            <option value="">— 选择文案 —</option>
          </select>
          <button class="btn" style="padding:8px 14px;font-size:.85rem" onclick="refreshDraftList()">刷新</button>
        </div>
        <textarea id="draftPreview" oninput="onDraftEdit()" style="width:100%;height:340px;border:1.5px solid #eee;border-radius:10px;padding:12px;font-size:.82rem;line-height:1.7;resize:vertical;background:#fff;color:#444;font-family:'PingFang SC',sans-serif" placeholder="选择一篇文案后在此编辑..."></textarea>
        <div style="display:flex;gap:8px;margin-top:8px">
          <button class="btn" style="flex:1;padding:8px;font-size:.85rem;background:linear-gradient(135deg,#34c759,#28a745)" onclick="saveDraft()" id="applyBtn" disabled>💾 保存</button>
          <button class="btn" style="flex:1;padding:8px;font-size:.85rem;background:linear-gradient(135deg,#ff6b35,#ee5a24)" onclick="generateCards()" id="genImgBtn" disabled>🃏 生成卡片</button>
        </div>
        <div id="cardTemplateBar" style="display:none;margin-top:8px;gap:6px;flex-wrap:wrap;align-items:center">
          <span style="font-size:.78rem;color:#888;flex-shrink:0">卡片样式：</span>
          <div id="cardTemplateChips" style="display:flex;gap:6px;flex-wrap:wrap"></div>
        </div>
        <div style="margin-top:6px;font-size:.75rem;color:#aaa" id="imgStatus"></div>
        <div id="cardSlider" style="display:none;margin-top:10px;position:relative;background:#f5f0eb;border-radius:12px;overflow:hidden">
          <img id="cardSliderImg" style="width:100%;display:block;border-radius:12px" />
          <div style="position:absolute;bottom:8px;left:50%;transform:translateX(-50%);display:flex;gap:6px;align-items:center" id="sliderDots"></div>
          <button onclick="slideCard(-1)" style="position:absolute;left:6px;top:50%;transform:translateY(-50%);background:rgba(0,0,0,.35);color:#fff;border:none;border-radius:50%;width:32px;height:32px;font-size:16px;cursor:pointer;line-height:1">‹</button>
          <button onclick="slideCard(1)"  style="position:absolute;right:6px;top:50%;transform:translateY(-50%);background:rgba(0,0,0,.35);color:#fff;border:none;border-radius:50%;width:32px;height:32px;font-size:16px;cursor:pointer;line-height:1">›</button>
        </div>
        <div id="cardSelectBar" style="display:none;align-items:center;justify-content:space-between;gap:10px;margin-top:8px">
          <div id="cardSelectStatus" style="font-size:.78rem;color:#666">当前未选择保存图片</div>
          <button id="selectCardBtn" class="btn btn-outline" style="padding:6px 12px;font-size:.8rem" onclick="selectCurrentCard()">设为保存图片</button>
        </div>
        <div id="pubReadySection" style="display:none;margin-top:12px;border-top:1px solid #eee;padding-top:12px">
          <div style="font-size:.82rem;font-weight:600;color:#555;margin-bottom:8px">📤 发布准备</div>
          <label style="font-size:.78rem;color:#888">标题（≤20字）</label>
          <input id="pubTitleInput" type="text" maxlength="20"
            style="width:100%;margin-bottom:6px;padding:6px 8px;border:1px solid #ddd;border-radius:6px;font-size:.82rem">
          <label style="font-size:.78rem;color:#888">简介（≤100字）</label>
          <textarea id="pubDescInput" maxlength="100" rows="2"
            style="width:100%;padding:6px 8px;border:1px solid #ddd;border-radius:6px;font-size:.82rem;resize:vertical"
            oninput="updatePubDescCount()"></textarea>
          <div style="display:flex;justify-content:space-between;align-items:center;margin-top:6px">
            <span id="pubDescCount" style="font-size:.72rem;color:#bbb">0 / 100</span>
            <button id="markReadyBtn" onclick="markReady()"
              style="padding:6px 16px;background:linear-gradient(135deg,#007aff,#5856d6);color:#fff;border:none;border-radius:8px;font-size:.82rem;cursor:pointer">
              ✅ 标记为发布就绪
            </button>
          </div>
          <div id="readyStatusMsg" style="font-size:.75rem;color:#aaa;margin-top:4px"></div>
        </div>
      </div>
      <!-- 右：对话框 -->
      <div style="flex:1;min-width:280px;display:flex;flex-direction:column">
        <div id="chatHistory" style="flex:1;height:300px;overflow-y:auto;border:1.5px solid #eee;border-radius:10px;padding:12px;background:#fafafa;margin-bottom:10px;display:flex;flex-direction:column;gap:10px">
          <div style="text-align:center;color:#bbb;font-size:.82rem;padding-top:40px">选择文案后开始对话修改</div>
        </div>
        <div style="display:flex;gap:8px">
          <textarea id="chatInput" rows="3" style="flex:1;padding:10px 12px;border:1.5px solid #eee;border-radius:10px;font-size:.9rem;resize:none;outline:none;font-family:inherit" placeholder="输入修改建议，例如：把语气改得更亲切一点，加强紧迫感..." onkeydown="if(event.key==='Enter'&&(event.metaKey||event.ctrlKey))sendChat()"></textarea>
          <button class="btn" id="chatBtn" style="padding:10px 18px;font-size:.9rem;white-space:nowrap;align-self:flex-end" onclick="sendChat()">发送</button>
        </div>
        <div style="font-size:.75rem;color:#bbb;margin-top:4px">⌘Enter 发送 &nbsp;·&nbsp; 修改文案后点「生成卡片」可重新生成图片</div>
      </div>
    </div>
  </div>

  <!-- 登录与账号管理 -->
  <div class="card" style="margin-bottom:16px">
    <div class="card-title">🔐 登录入口与账号管理</div>
    <div class="acct-toolbar">
      <div class="acct-inline-actions">
        <button class="btn" type="button" onclick="startAccountLogin('scan')">扫码登录新账号</button>
      </div>
    </div>
    <div id="accountLoginStatus" class="acct-status">点击“扫码登录新账号”后，系统会自动识别原账号名并写入账号列表。</div>
    <div style="overflow-x:auto">
      <table class="acct-table">
        <thead>
          <tr>
            <th>显示名</th>
            <th>启用</th>
            <th>Session</th>
            <th>上次使用</th>
            <th>存储</th>
            <th>操作</th>
          </tr>
        </thead>
        <tbody id="accountManageTbody">
          <tr><td colspan="6" style="color:#bbb;text-align:center;padding:16px">暂无账号</td></tr>
        </tbody>
      </table>
    </div>
  </div>

  <!-- 发布管理 -->
  <div class="card" style="margin-bottom:16px">
    <div class="card-title">📤 发布管理</div>
    <div style="display:flex;gap:12px;margin-bottom:16px">
      <button class="pub-btn pub-btn-now" onclick="openPublishModal('now')">立即发布</button>
      <button class="pub-btn pub-btn-sched" onclick="openPublishModal('sched')">定时发布</button>
    </div>
    <div id="schedTasksWrap" style="overflow-x:auto">
      <table class="sched-tbl" id="schedTbl">
        <thead><tr>
          <th>发布时间</th><th>内容</th><th>账号</th><th>状态</th><th>操作</th>
        </tr></thead>
        <tbody id="schedTbody"><tr><td colspan="5" style="color:#bbb;text-align:center;padding:16px">暂无定时任务</td></tr></tbody>
      </table>
    </div>
  </div>

  <!-- 发布 Modal -->
  <div class="modal-overlay" id="pubModal">
    <div class="modal">
      <h2 id="modalTitle">立即发布</h2>
      <label>账号 & 内容分配</label>
      <div id="readyDraftHint" style="font-size:.78rem;color:#888;margin-bottom:8px"></div>
      <div id="publishAssignList" style="display:flex;flex-direction:column;gap:10px"></div>

      <div id="schedTimeRow" style="display:none">
        <label>发布时间</label>
        <div style="display:flex;gap:8px">
          <input id="schedDate" type="date" style="flex:1">
          <input id="schedTime" type="time" style="width:110px">
        </div>
      </div>

      <div class="modal-footer">
        <button class="btn-cancel-modal" onclick="closePublishModal()">取消</button>
        <button class="btn-confirm" id="confirmBtn" onclick="confirmPublish()">确认发布</button>
      </div>
    </div>
  </div>

  <div class="modal-overlay" id="draftPickModal">
    <div class="modal" style="max-width:420px">
      <h2 id="draftPickTitle">选择文案</h2>
      <div id="draftPickList" style="display:flex;flex-direction:column;gap:8px;max-height:320px;overflow:auto"></div>
      <div class="modal-footer">
        <button class="btn-cancel-modal" onclick="closeDraftPickModal()">取消</button>
        <button class="btn-confirm" onclick="confirmDraftPick()">确认</button>
      </div>
    </div>
  </div>

  <!-- Toast -->
  <div class="toast" id="toast"></div>
</div>

<script>
let pollId = null;

function startPipeline() {
  const kw = document.getElementById('kw').value.trim();
  const acc = document.getElementById('accCount').value;
  const cp = document.getElementById('copies').value;
  const mode = document.getElementById('analysisMode').value;
  const retrievalAccountSel = document.getElementById('retrievalAccountSel');
  const retrievalAccount = retrievalAccountSel ? retrievalAccountSel.value : '';
  if (!kw) return;
  fetch(`/api/run?keyword=${encodeURIComponent(kw)}&accounts=${acc}&copies=${cp}&mode=${encodeURIComponent(mode)}&retrieval_account=${encodeURIComponent(retrievalAccount)}`)
    .then(r => r.json()).then(d => {
      if (!d.started) {
        showToast('⚠️ 任务正在运行中，请等待完成后再试', 'err');
        return;
      }
      pollId = setInterval(fetchState, 800);
      document.getElementById('runBtn').disabled = true;
    });

}

function openOutput() { fetch('/api/open_output'); }

function fetchState() {
  fetch('/api/state').then(r => r.json()).then(updateUI);
}

function updateUI(s) {
  // Status bar
  const bar = document.getElementById('statusBar');
  const txt = document.getElementById('statusText');
  const pulse = document.getElementById('pulse');
  const modeLabel = s.mode_label ? ` · ${s.mode_label}` : '';
  const labels = {idle:'空闲，等待任务', running:`运行中… 步骤 ${s.step}/6${modeLabel}`, completed:`✅ 全部完成！${modeLabel}`, error:`❌ 出现错误${modeLabel}`};
  txt.textContent = labels[s.status] || s.status;
  bar.className = 'status-bar status-' + s.status;
  pulse.style.display = s.status === 'running' ? 'block' : 'none';
  if (s.mode) {
    const modeSel = document.getElementById('analysisMode');
    if (modeSel && document.activeElement !== modeSel) modeSel.value = s.mode;
  }

  // Steps
  for (let i = 1; i <= 6; i++) {
    const dot = document.getElementById('sd' + i);
    const lbl = document.getElementById('sl' + i);
    const stepNum = i;
    dot.className = 'step-dot';
    lbl.className = 'step-label';
    if (s.step > stepNum || s.status === 'completed') { dot.classList.add('done'); }
    else if (s.step === stepNum) { dot.classList.add('active'); lbl.classList.add('active'); }
  }

  // Logs
  const box = document.getElementById('logsBox');
  box.innerHTML = s.logs.map(l =>
    `<div class="log log-${l.level}"><span class="log-time">[${l.time}]</span>${escHtml(l.msg)}</div>`
  ).join('');
  box.scrollTop = box.scrollHeight;

  // Stats
  if (s.stats && s.stats.accounts) {
    document.getElementById('sv1').textContent = s.stats.accounts;
    document.getElementById('sv2').textContent = s.stats.drafts;
    document.getElementById('sv3').textContent = s.stats.cards;
  }

  // Accounts table
  if (s.accounts && s.accounts.length > 0) {
    document.getElementById('accCard').style.display = 'block';
    const tbody = document.getElementById('accTbody');
    tbody.innerHTML = s.accounts.map((a, i) => {
      const rating = a['账号权重评级'] || '';
      const badge = rating.startsWith('A') ? 'badge-a' : 'badge-b';
      const hooks = (a['Hook模板'] || '').split(/[|｜]/).filter(Boolean).slice(0,2).map(h => `<span class="tag">${escHtml(h.trim())}</span>`).join('');
      const ctas = (a['CTA模板'] || '').split(/[|｜]/).filter(Boolean).slice(0,1).map(c => `<span class="tag">${escHtml(c.trim())}</span>`).join('');
      return `<tr>
        <td>${i+1}</td>
        <td><strong>${escHtml(a['账号昵称'] || '')}</strong></td>
        <td>${fmtNum(a['粉丝数'])}</td>
        <td><span class="badge ${badge}">${escHtml(rating.split('｜')[0])}</span></td>
        <td style="max-width:160px;font-size:.78rem;color:#666">${escHtml((a['内容偏好'] || '').slice(0,50))}</td>
        <td>${hooks}</td>
        <td>${ctas}</td>
      </tr>`;
    }).join('');
  }

  // Drafts
  if (s.drafts && s.drafts.length > 0) {
    document.getElementById('draftCard').style.display = 'block';
    const dl = document.getElementById('draftList');
    dl.innerHTML = [...s.drafts].reverse().map(p => {
      const name = p.split('/').pop().replace('.md','');
      const parts = name.split('_');
      const accName = parts.slice(0,-2).join('_');
      const copy = parts[parts.length-2] || '';
      return `<div class="draft-item" data-path="${escHtml(p)}" onclick="selectDraft(this.dataset.path)">
        <div class="draft-title">📄 ${escHtml(name)}</div>
        <div class="draft-meta">账号：${escHtml(accName)} &nbsp;|&nbsp; ${escHtml(copy)}</div>
      </div>`;
    }).join('');
  }

  if (s.status !== 'running') {
    clearInterval(pollId);
    document.getElementById('runBtn').disabled = false;
  }
}

function fmtNum(n) {
  if (!n || isNaN(n)) return '—';
  if (n >= 10000) return (n/10000).toFixed(1) + 'w';
  return n.toString();
}

function escHtml(s) {
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

function escAttr(s) {
  return escHtml(s);
}

// ── 文案对话修改 ──────────────────────────────────────────────────────────────
let chatHistory = [];
let currentDraftFile = '';
let currentDraftContent = '';
let currentImageStyle = '';
let currentTopic = '';
let currentAccount = '';
let currentWritingStyle = '';
let _cardUrls = [];
let _cardIdx  = 0;
let _isGeneratingCards = false;
let _selectedImageUrls = [];

function currentCardUrl() {
  return _cardUrls[_cardIdx] || '';
}

function savedImagePaths() {
  return [..._selectedImageUrls];
}

function updateCardSelectionUI() {
  const bar = document.getElementById('cardSelectBar');
  const status = document.getElementById('cardSelectStatus');
  const btn = document.getElementById('selectCardBtn');
  if (!_cardUrls.length) {
    bar.style.display = 'none';
    return;
  }
  bar.style.display = 'flex';
  const currentNo = _cardIdx + 1;
  const selectedIndexes = _selectedImageUrls
    .map(url => _cardUrls.indexOf(url))
    .filter(idx => idx >= 0)
    .map(idx => idx + 1);
  if (selectedIndexes.length) {
    status.textContent = `已选 ${selectedIndexes.length} 张保存图（第 ${selectedIndexes.join('、')} 张），当前浏览第 ${currentNo} 张`;
  } else {
    status.textContent = `当前浏览第 ${currentNo} 张，尚未明确选择保存图`;
  }
  btn.textContent = _selectedImageUrls.includes(currentCardUrl()) ? '取消这张保存图' : '加入保存图片';
}

function selectCurrentCard() {
  if (!_cardUrls.length) return;
  const url = currentCardUrl();
  if (_selectedImageUrls.includes(url)) {
    _selectedImageUrls = _selectedImageUrls.filter(item => item !== url);
    showToast(`已取消第 ${_cardIdx + 1} 张保存图`, 'ok');
  } else {
    _selectedImageUrls = [..._selectedImageUrls, url];
    showToast(`已加入第 ${_cardIdx + 1} 张到保存图片`, 'ok');
  }
  updateCardSelectionUI();
  _renderSlider();
}

function refreshDraftList(selectPath, shouldLoad = true) {
  fetch('/api/drafts').then(r => r.json()).then(files => {
    const sel = document.getElementById('draftSel');
    const cur = selectPath || sel.value;
    sel.innerHTML = '<option value="">— 选择文案 —</option>' +
      files.map(f => {
        const suffix = (f.card_count || 0) > 0 ? ' [已保存图片]' : ' [仅文案]';
        return `<option value="${escHtml(f.path)}"${f.path===cur?' selected':''}>${escHtml(f.name + suffix)}</option>`;
      }).join('');
    if (selectPath) {
      sel.value = selectPath;
      if (shouldLoad) loadDraft();
    }
  });
}

function selectDraft(path) {
  refreshDraftList(path);
  document.getElementById('draftSel').closest('.card')
    .scrollIntoView({behavior:'smooth', block:'nearest'});
}

function loadDraft() {
  _isGeneratingCards = false;
  const file = document.getElementById('draftSel').value;
  if (!file) return;
  fetch(`/api/draft?file=${encodeURIComponent(file)}`).then(r => r.json()).then(d => {
    if (d.error) { alert(d.error); return; }
    currentDraftFile = file;
    currentDraftContent = d.content;
    document.getElementById('draftPreview').value = d.content;
    chatHistory = [];
    document.getElementById('applyBtn').disabled = false;
    document.getElementById('genImgBtn').disabled = _isGeneratingCards;
    _cardUrls = d.cards || [];
    _cardIdx = 0;
    _selectedImageUrls = [..._cardUrls];
    if (_cardUrls.length) {
      document.getElementById('cardSlider').style.display = 'block';
      _renderSlider();
      updateCardSelectionUI();
      document.getElementById('imgStatus').textContent = `🖼 已加载 ${_cardUrls.length} 张已保存确认图片，可继续调整选择或重新生成`;
    } else {
      updateCardSelectionUI();
      document.getElementById('imgStatus').textContent = '当前文案还没有保存确认的图片';
      document.getElementById('cardSlider').style.display = 'none';
    }
    document.getElementById('chatHistory').innerHTML =
      '<div style="text-align:center;color:#bbb;font-size:.82rem;padding-top:40px">文案已加载，可以开始对话修改，或点击「生成卡片」</div>';
    const meta = extractDraftMeta(d.content);
    currentImageStyle = meta.image_style || '';
    currentTopic = meta.topic || '';
    currentAccount = meta.account || '';
    currentWritingStyle = meta.writing_style || '';

    const titleMatch = d.content.match(/^#\\s+(.+)/m);
    document.getElementById('pubTitleInput').value = titleMatch ? titleMatch[1].trim().slice(0, 20) : '';
    const descMatch = d.content
      .replace(/^---[\\s\\S]*?---\\n/m, '')
      .replace(/^#+.+$/mg, '')
      .replace(/\\*\\*[^*]+\\*\\*/g, '')
      .trim()
      .slice(0, 100);
    document.getElementById('pubDescInput').value = descMatch;
    updatePubDescCount();
    document.getElementById('pubReadySection').style.display = 'block';
    document.getElementById('readyStatusMsg').textContent = '';
    document.getElementById('markReadyBtn').disabled = false;
    document.getElementById('markReadyBtn').textContent = '✅ 标记为发布就绪';
  });
}

function extractDraftMeta(content) {
  const get = (name) => {
    const m = content.match(new RegExp(`\\*\\*${name}\\*\\*[：:][ \\t]*(.+)`));
    return m ? m[1].trim() : '';
  };
  return {
    account: get('账号参考'),
    topic: get('主题'),
    image_style: get('图片风格'),
    writing_style: get('图文风格描述')
  };
}

function updatePubDescCount() {
  const desc = document.getElementById('pubDescInput').value || '';
  document.getElementById('pubDescCount').textContent = desc.length + ' / 100';
}

async function markReady() {
  const title = document.getElementById('pubTitleInput').value.trim();
  const desc = document.getElementById('pubDescInput').value.trim();
  if (!title) {
    document.getElementById('readyStatusMsg').textContent = '请填写标题';
    return;
  }
  document.getElementById('markReadyBtn').disabled = true;
  try {
    const imagePaths = savedImagePaths();
    const res = await fetch('/api/mark_ready', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        md_path: currentDraftFile,
        title,
        desc,
        image_path: imagePaths[0] || '',
        image_paths: imagePaths,
        account: currentAccount,
        image_style: currentImageStyle,
        writing_style: currentWritingStyle
      })
    }).then(r => r.json());
    if (res.error) {
      document.getElementById('readyStatusMsg').textContent = '❌ ' + res.error;
    } else {
      document.getElementById('readyStatusMsg').textContent = '✅ 已就绪，可在发布管理中选择发布';
      document.getElementById('markReadyBtn').textContent = '✅ 已就绪';
      loadReadyDrafts();
    }
  } catch (e) {
    document.getElementById('readyStatusMsg').textContent = '❌ ' + e.message;
  } finally {
    document.getElementById('markReadyBtn').disabled = false;
  }
}

function sendChat() {
  const input = document.getElementById('chatInput');
  const msg = input.value.trim();
  if (!msg) return;
  if (!currentDraftFile) { alert('请先选择一篇文案'); return; }

  appendChatMsg('user', msg);
  input.value = '';
  document.getElementById('chatBtn').disabled = true;
  document.getElementById('chatBtn').textContent = '思考中…';

  fetch('/api/chat', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      draft_content: currentDraftContent,
      draft_path:    currentDraftFile,
      image_style:   currentImageStyle,
      topic:         currentTopic,
      history:       chatHistory,
      message:       msg
    })
  }).then(r => r.json()).then(d => {
    document.getElementById('chatBtn').disabled = false;
    document.getElementById('chatBtn').textContent = '发送';
    if (d.error) {
      appendChatMsg('error', '❌ ' + d.error);
      return;
    }
    chatHistory.push({role:'user', content: msg});
    chatHistory.push({role:'assistant', content: d.reply});
    appendChatMsg('assistant', d.reply);
    if (d.intent === 'card_task' && d.task_id) {
      document.getElementById('imgStatus').textContent = '🔄 卡片生成中，请稍候…';
      _pollCardTask(d.task_id);
    } else if (d.intent === 'card' && d.cards && d.cards.length) {
      _cardUrls = d.cards; _cardIdx = 0;
      _selectedImageUrls = [];
      document.getElementById('cardSlider').style.display = 'block';
      _renderSlider();
      updateCardSelectionUI();
      document.getElementById('imgStatus').textContent = `✅ ${d.cards.length} 张卡片已重新生成，请选择要保存的图片`;
    } else if (d.updated_draft) {
      document.getElementById('draftPreview').value = d.updated_draft;
      document.getElementById('applyBtn').disabled = false;
    }
  }).catch(e => {
    document.getElementById('chatBtn').disabled = false;
    document.getElementById('chatBtn').textContent = '发送';
    appendChatMsg('error', '❌ 请求失败: ' + e.message);
  });
}

function appendChatMsg(role, text) {
  const box = document.getElementById('chatHistory');
  // 清除初始占位
  if (box.querySelector('div[style*="padding-top"]')) box.innerHTML = '';
  const div = document.createElement('div');
  div.style.cssText = `padding:10px 14px;border-radius:10px;font-size:.85rem;line-height:1.7;max-width:95%;word-break:break-word;white-space:pre-wrap;`;
  if (role === 'user') {
    div.style.background = '#ff6b6b22';
    div.style.alignSelf = 'flex-end';
    div.style.borderBottomRightRadius = '2px';
    div.textContent = text;
  } else if (role === 'assistant') {
    div.style.background = '#f0f0f0';
    div.style.alignSelf = 'flex-start';
    div.style.borderBottomLeftRadius = '2px';
    div.textContent = text;
  } else {
    div.style.background = '#fff3cd';
    div.style.color = '#856404';
    div.textContent = text;
  }
  box.appendChild(div);
  box.scrollTop = box.scrollHeight;
}

function onDraftEdit() {
  document.getElementById('applyBtn').disabled = false;
}

function saveDraft() {
  if (!currentDraftFile) {
    showToast('请先从左侧选择一篇文案', 'err'); return;
  }
  const content = document.getElementById('draftPreview').value;
  if (!content.trim()) { showToast('文案内容不能为空', 'err'); return; }
  const hasSelectedImages = _selectedImageUrls.length > 0;
  fetch('/api/save_draft', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({
      file: currentDraftFile,
      content,
      image_path: hasSelectedImages ? (_selectedImageUrls[0] || '') : null,
      image_paths: hasSelectedImages ? savedImagePaths() : null,
    })
  }).then(r => r.json()).then(d => {
    if (d.ok) {
      currentDraftContent = content;
      if (d.new_path && d.new_path !== currentDraftFile) {
        currentDraftFile = d.new_path;
        refreshDraftList(d.new_path, false);
      }
      showToast('已保存', 'ok');
    } else {
      showToast(d.error || '保存失败', 'err');
    }
  }).catch(e => showToast('保存失败：' + e.message, 'err'));
}

function applyUpdated() { saveDraft(); }

function _renderSlider() {
  if (!_cardUrls.length) return;
  const img = document.getElementById('cardSliderImg');
  img.src = _cardUrls[_cardIdx] + '?t=' + Date.now();
  const dots = document.getElementById('sliderDots');
  dots.innerHTML = _cardUrls.map((url, i) => {
    const isCurrent = i === _cardIdx;
    const isSelected = _selectedImageUrls.includes(url);
    let bg = 'rgba(255,255,255,.6)';
    let width = 8;
    if (isCurrent && isSelected) {
      bg = '#34c759';
      width = 22;
    } else if (isCurrent) {
      bg = '#ff6b35';
      width = 22;
    } else if (isSelected) {
      bg = '#0a84ff';
      width = 14;
    }
    return `<div style="width:${width}px;height:8px;border-radius:4px;background:${bg};transition:width .2s"></div>`;
  }).join('');
}

function slideCard(dir) {
  if (!_cardUrls.length) return;
  _cardIdx = (_cardIdx + dir + _cardUrls.length) % _cardUrls.length;
  _renderSlider();
  updateCardSelectionUI();
}

function _pollCardTask(taskId) {
  const timer = setInterval(async () => {
    try {
      const t = await fetch('/api/card_task/' + taskId).then(r => r.json());
      if (t.status === 'running') return;
      clearInterval(timer);
      fetch('/api/card_task/' + taskId, {method: 'DELETE'}).catch(() => {});
      if (t.status === 'done') {
        _cardUrls = t.cards || []; _cardIdx = 0;
        _selectedImageUrls = [];
        if (_cardUrls.length) {
          document.getElementById('cardSlider').style.display = 'block';
          _renderSlider();
          updateCardSelectionUI();
          document.getElementById('imgStatus').textContent = `✅ ${_cardUrls.length} 张卡片已重新生成，请选择要保存的图片`;
          appendChatMsg('assistant', t.reply || '卡片已更新');
        } else {
          updateCardSelectionUI();
          document.getElementById('imgStatus').textContent = '⚠️ 未生成任何卡片';
        }
      } else {
        document.getElementById('imgStatus').textContent = '❌ 生成失败: ' + (t.error || '未知错误');
      }
    } catch(e) {
      clearInterval(timer);
      document.getElementById('imgStatus').textContent = '❌ 轮询失败: ' + e.message;
    }
  }, 2000);
}

async function generateCards() {
  if (_isGeneratingCards) return;
  let draftPath = currentDraftFile || document.getElementById('draftSel').value;
  if (!draftPath) return;
  _isGeneratingCards = true;
  document.getElementById('imgStatus').textContent = '卡片生成中，请稍候（约15秒）…';
  document.getElementById('genImgBtn').disabled = true;
  document.getElementById('cardSlider').style.display = 'none';
  // 先保存当前文案内容，确保生成的卡片与编辑器内容一致
  const content = document.getElementById('draftPreview').value;
  if (content) {
    try {
      const sr = await fetch('/api/save_draft', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({file: draftPath, content})
      }).then(r => r.json());
      if (sr.new_path) {
        currentDraftFile = sr.new_path;
        draftPath = sr.new_path;
        refreshDraftList(sr.new_path, false);
      }
    } catch(e) {
      document.getElementById('imgStatus').textContent = '❌ 保存失败: ' + e.message;
      _isGeneratingCards = false;
      document.getElementById('genImgBtn').disabled = false;
      return;
    }
  }
  try {
    const res = await fetch('/api/gen_card', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({draft_path: draftPath, card_template: _selectedCardTemplate || ''})
    }).then(r => r.json());
    if (res.error) {
      document.getElementById('imgStatus').textContent = '❌ ' + res.error;
      _isGeneratingCards = false;
      document.getElementById('genImgBtn').disabled = false;
      return;
    }
    _cardUrls = res.cards || [];
    _cardIdx  = 0;
    _selectedImageUrls = [];
    if (_cardUrls.length) {
      document.getElementById('cardSlider').style.display = 'block';
      _renderSlider();
      updateCardSelectionUI();
      document.getElementById('imgStatus').textContent = `✅ ${_cardUrls.length} 张卡片生成完成，可多选后再保存`;
    } else {
      updateCardSelectionUI();
      document.getElementById('imgStatus').textContent = '⚠️ 未生成任何卡片，请检查文案内容';
    }
  } catch(e) {
    document.getElementById('imgStatus').textContent = '❌ 请求失败: ' + e.message;
  }
  _isGeneratingCards = false;
  document.getElementById('genImgBtn').disabled = false;
}

// 初始加载文案列表
refreshDraftList();

// 初始拉取状态
fetchState();

// ── 人设切换 ─────────────────────────────────────────────────────────────────

let _personasData = [];
let _activePersonaId = '';
let _selectedCardTemplate = '';

function loadPersonas() {
  fetch('/api/personas').then(r => r.json()).then(data => {
    _personasData = data.personas || [];
    _activePersonaId = data.active || '';
    const chips = document.getElementById('personaChips');
    if (!chips) return;
    chips.innerHTML = _personasData.map(p => `
      <div class="persona-chip${p.id === _activePersonaId ? ' active' : ''}"
           onclick="switchPersona('${p.id}')">
        ${p.emoji || ''} ${p.name}
        <span class="p-desc">${p.description || ''}</span>
      </div>
    `).join('');
    _renderCardTemplateChips();
  }).catch(() => {});
}

function _renderCardTemplateChips() {
  const bar = document.getElementById('cardTemplateBar');
  const chipsEl = document.getElementById('cardTemplateChips');
  if (!bar || !chipsEl) return;
  const persona = _personasData.find(p => p.id === _activePersonaId);
  const templates = (persona && persona.card_templates) || [{id:'oval', name:'⭕ 椭圆卡片'}];
  // 重置选中：选默认模板（第一个）
  if (!_selectedCardTemplate || !templates.find(t => t.id === _selectedCardTemplate)) {
    _selectedCardTemplate = templates[0].id;
  }
  chipsEl.innerHTML = templates.map(t => `
    <div class="tpl-chip${t.id === _selectedCardTemplate ? ' active' : ''}"
         onclick="selectCardTemplate('${t.id}')">${t.name}</div>
  `).join('');
  bar.style.display = templates.length > 1 ? 'flex' : 'none';
}

function selectCardTemplate(id) {
  _selectedCardTemplate = id;
  _renderCardTemplateChips();
}

function switchPersona(id) {
  fetch('/api/personas/active', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({id})
  }).then(r => r.json()).then(d => {
    if (d.ok) {
      _selectedCardTemplate = '';  // 重置，让新人设默认模板生效
      loadPersonas();
      showToast(`人设已切换：${d.name}`);
    } else {
      showToast('切换失败：' + (d.error || ''), 'err');
    }
  });
}

loadPersonas();

// ── 账号登录与管理 ────────────────────────────────────────────────────────────

let _managedAccounts = [];
let _accountLoginTaskId = '';

function accountDisplayName(acc) {
  return (acc && (acc.display_name || acc.nickname || acc.name)) || '';
}

function accountNicknameInputId(name) {
  return 'nicknameInput_' + encodeURIComponent(String(name || ''));
}

function fmtIsoTime(s) {
  if (!s) return '—';
  const d = new Date(s);
  if (Number.isNaN(d.getTime())) return String(s).replace('T', ' ').slice(0, 16);
  return d.toLocaleString('zh-CN', {
    month: '2-digit',
    day: '2-digit',
    hour: '2-digit',
    minute: '2-digit',
    hour12: false
  });
}

function updateAccountLoginStatus(text, status) {
  const el = document.getElementById('accountLoginStatus');
  el.textContent = text;
  el.className = 'acct-status' + (status ? ' ' + status : '');
}

function loadManagedAccounts() {
  fetch('/api/manage_accounts').then(r => r.json()).then(accounts => {
    _managedAccounts = Array.isArray(accounts) ? accounts : [];
    renderManagedAccounts();
    renderRetrievalAccountOptions();
  });
}

function renderRetrievalAccountOptions() {
  const sel = document.getElementById('retrievalAccountSel');
  if (!sel) return;
  const current = sel.value;
  const options = ['<option value="">自动轮询</option>'].concat(
    _managedAccounts
      .filter(acc => acc.active)
      .map(acc => `<option value="${escHtml(acc.name)}">${escHtml(accountDisplayName(acc))}${acc.valid ? '' : '（待刷新）'}</option>`)
  );
  sel.innerHTML = options.join('');
  if (_managedAccounts.some(acc => acc.name === current && acc.active)) {
    sel.value = current;
  }
}

function renderManagedAccounts() {
  const tbody = document.getElementById('accountManageTbody');
  if (!_managedAccounts.length) {
    tbody.innerHTML = '<tr><td colspan="6" style="color:#bbb;text-align:center;padding:16px">暂无账号，先扫码登录新账号</td></tr>';
    return;
  }
  tbody.innerHTML = _managedAccounts.map(acc => {
    const enableBadge = acc.active
      ? '<span class="acct-pill acct-pill-on">启用中</span>'
      : '<span class="acct-pill acct-pill-off">已停用</span>';
    const validBadge = acc.valid
      ? '<span class="acct-pill acct-pill-valid">有效</span>'
      : '<span class="acct-pill acct-pill-invalid">待刷新</span>';
    const storageState = acc.storage_exists ? 'storage 已就绪' : 'storage 缺失';
    const displayName = accountDisplayName(acc);
    const inputValue = acc.nickname ? acc.nickname : displayName;
    return `<tr>
      <td>
        <input class="acct-name-input" id="${accountNicknameInputId(acc.name)}" value="${escAttr(inputValue)}" placeholder="${escAttr(acc.name || '')}">
        <div class="acct-name-hint">未设置昵称时默认显示原账号名</div>
      </td>
      <td>${enableBadge}</td>
      <td>${validBadge}<div style="font-size:.74rem;color:#999;margin-top:4px">${escHtml(fmtIsoTime(acc.session_expires_at))}</div></td>
      <td>${escHtml(fmtIsoTime(acc.last_used_at))}</td>
      <td><div style="font-size:.74rem;color:#666">${escHtml(storageState)}</div><div style="font-size:.72rem;color:#aaa;margin-top:4px">${escHtml(acc.storage_path || '')}</div></td>
      <td>
        <div class="acct-ops">
          <button class="btn btn-outline" type="button" style="padding:6px 10px;font-size:.8rem" onclick="saveManagedNickname('${escJs(acc.name)}')">保存昵称</button>
          <button class="btn btn-outline" type="button" style="padding:6px 10px;font-size:.8rem" onclick="refreshAccountLogin('${escJs(acc.name)}')">刷新登录</button>
          <button class="btn btn-outline" type="button" style="padding:6px 10px;font-size:.8rem" onclick="toggleManagedAccount('${escJs(acc.name)}', ${acc.active ? 'false' : 'true'})">${acc.active ? '停用' : '启用'}</button>
          <button class="btn btn-outline" type="button" style="padding:6px 10px;font-size:.8rem;color:#c62828" onclick="deleteManagedAccount('${escJs(acc.name)}')">删除</button>
        </div>
      </td>
    </tr>`;
  }).join('');
}

function startAccountLogin(mode) {
  const name = arguments.length > 1 ? String(arguments[1] || '').trim() : '';
  if (mode === 'refresh' && !name) {
    showToast('缺少要刷新的账号', 'err');
    return;
  }
  updateAccountLoginStatus('正在启动浏览器，请在 120 秒内完成扫码…', 'running');
  fetch('/api/account_login', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({name, mode})
  }).then(r => r.json()).then(task => {
    if (task.error) {
      updateAccountLoginStatus('启动失败：' + task.error, 'error');
      return;
    }
    _accountLoginTaskId = task.id;
    pollAccountLoginTask(task.id);
  }).catch(e => updateAccountLoginStatus('启动失败：' + e.message, 'error'));
}

function refreshAccountLogin(name) {
  startAccountLogin('refresh', name);
}

function saveManagedNickname(name) {
  const input = document.getElementById(accountNicknameInputId(name));
  if (!input) return;
  const nickname = input.value.trim();
  fetch('/api/manage_accounts/nickname', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({name, nickname})
  }).then(r => r.json()).then(data => {
    if (data.error) {
      showToast(data.error, 'err');
      return;
    }
    loadManagedAccounts();
    showToast('昵称已保存', 'ok');
  }).catch(e => showToast('保存昵称失败：' + e.message, 'err'));
}

function pollAccountLoginTask(taskId) {
  const timer = setInterval(async () => {
    try {
      const task = await fetch('/api/login_task/' + taskId).then(r => r.json());
      if (task.status === 'running') {
        updateAccountLoginStatus(`${task.name}：${task.message}`, 'running');
        return;
      }
      clearInterval(timer);
      fetch('/api/login_task/' + taskId, {method: 'DELETE'}).catch(() => {});
      if (task.status === 'done') {
        updateAccountLoginStatus(`${task.name}：${task.message}`, 'done');
        showToast('登录成功，账号已更新', 'ok');
        loadManagedAccounts();
      } else {
        updateAccountLoginStatus(`${task.name}：${task.message || '登录失败'}`, 'error');
        showToast('登录失败，请查看状态提示', 'err');
      }
    } catch (e) {
      clearInterval(timer);
      updateAccountLoginStatus('登录状态轮询失败：' + e.message, 'error');
    }
  }, 2000);
}

function toggleManagedAccount(name, active) {
  fetch('/api/manage_accounts/toggle', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({name, active})
  }).then(r => r.json()).then(data => {
    if (data.error) {
      showToast(data.error, 'err');
      return;
    }
    loadManagedAccounts();
    showToast(active ? '账号已启用' : '账号已停用', 'ok');
  }).catch(e => showToast('更新失败：' + e.message, 'err'));
}

function deleteManagedAccount(name) {
  if (!confirm(`确认删除账号「${name}」及本地登录态？`)) return;
  fetch('/api/manage_accounts/' + encodeURIComponent(name), {method: 'DELETE'})
    .then(r => r.json()).then(data => {
      if (data.error) {
        showToast(data.error, 'err');
        return;
      }
      loadManagedAccounts();
      showToast('账号已删除', 'ok');
    }).catch(e => showToast('删除失败：' + e.message, 'err'));
}

// ── 发布管理 ────────────────────────────────────────────────────────────────

let _pubMode = 'now';
let _allAccounts = [];
let _readyDrafts = [];
let _allDrafts = [];
let _pubAssignments = {};
let _draftPickAccount = '';

function showToast(msg, type) {
  const t = document.getElementById('toast');
  t.textContent = msg;
  t.className = 'toast show toast-' + type;
  setTimeout(() => t.className = 'toast', 3000);
}

function _assignedPaths() {
  return new Set(Object.values(_pubAssignments).flat());
}

function loadReadyDrafts() {
  return Promise.all([
    fetch('/api/ready_drafts').then(r => r.json()),
    fetch('/api/drafts').then(r => r.json()),
  ]).then(([ready, all]) => {
    _readyDrafts = ready;
    _allDrafts = all;
    const hint = document.getElementById('readyDraftHint');
    if (hint) {
      const withCards = all.filter(d => (d.card_count || 0) > 0).length;
      hint.textContent = all.length
        ? `共 ${all.length} 篇文案可选，其中 ${withCards} 篇已保存确认图片`
        : '暂无文案，请先生成';
    }
    return ready;
  });
}

function renderPublishAssignments() {
  const wrap = document.getElementById('publishAssignList');
  if (!wrap) return;
  if (!_allAccounts.length) {
    wrap.innerHTML = '<div style="color:#bbb;font-size:.85rem">暂无可用账号</div>';
    return;
  }
  wrap.innerHTML = _allAccounts.map(acc => {
    const disabled = !acc.valid;
    const checked = !!_pubAssignments[acc.name];
    const paths = _pubAssignments[acc.name] || [];
    const displayName = acc.display_name || acc.nickname || acc.name;
    const chips = paths.map(path => {
      const draft = _readyDrafts.find(d => d.md_path === path);
      const title = draft ? draft.title : PathFallback(path);
      return `<span style="display:inline-flex;align-items:center;gap:6px;padding:4px 8px;background:#f3f4f6;border-radius:999px;font-size:.78rem">
        ${escHtml(title)}
        <button type="button" onclick="removeAssignedDraft('${escJs(acc.name)}','${escJs(path)}')" style="border:none;background:none;color:#999;cursor:pointer;padding:0">×</button>
      </span>`;
    }).join('');
    return `<div class="pub-assign-row${disabled?' disabled':''}">
      <label class="pub-assign-account">
        <input type="checkbox" ${checked?'checked':''} ${disabled?'disabled':''} onchange="toggleAccountAssignment('${escJs(acc.name)}', this.checked)">
        <span class="pub-assign-name">${escHtml(displayName)}${disabled?' ⚠️ 已过期':''}</span>
      </label>
      <div class="pub-assign-chips">${chips || '<span style="font-size:.78rem;color:#bbb">尚未选择文案</span>'}</div>
      <button type="button" class="btn btn-outline pub-assign-action" ${disabled?'disabled':''}
        onclick="openDraftPickModal('${escJs(acc.name)}')">＋ 选择文案</button>
    </div>`;
  }).join('');
}

function PathFallback(path) {
  const parts = String(path || '').split('/');
  return parts[parts.length - 1] || path;
}

function escJs(s) {
  return String(s || '').replace(/\\\\/g, '\\\\\\\\').replace(/'/g, "\\\\'");
}

function openPublishModal(mode) {
  _pubMode = mode;
  document.getElementById('modalTitle').textContent = mode === 'now' ? '立即发布' : '定时发布';
  document.getElementById('schedTimeRow').style.display = mode === 'sched' ? 'block' : 'none';
  document.getElementById('confirmBtn').disabled = false;
  _pubAssignments = {};

  if (mode === 'sched') {
    const tomorrow = new Date(); tomorrow.setDate(tomorrow.getDate() + 1);
    document.getElementById('schedDate').value = tomorrow.toISOString().slice(0, 10);
    document.getElementById('schedTime').value = '09:00';
  }

  fetch('/api/accounts').then(r => r.json()).then(accounts => {
    _allAccounts = accounts;
    loadReadyDrafts().then(() => renderPublishAssignments());
  });

  document.getElementById('pubModal').classList.add('open');
}

function closePublishModal() {
  document.getElementById('pubModal').classList.remove('open');
}

function toggleAccountAssignment(account, checked) {
  if (checked) _pubAssignments[account] = _pubAssignments[account] || [];
  else delete _pubAssignments[account];
  renderPublishAssignments();
}

function removeAssignedDraft(account, draftPath) {
  const next = (_pubAssignments[account] || []).filter(p => p !== draftPath);
  if (next.length) _pubAssignments[account] = next;
  else delete _pubAssignments[account];
  renderPublishAssignments();
}

function openDraftPickModal(account) {
  if (!_pubAssignments[account]) _pubAssignments[account] = [];
  _draftPickAccount = account;
  const assigned = _assignedPaths();
  const selfAssigned = new Set(_pubAssignments[account] || []);
  const readyPaths = new Set(_readyDrafts.map(d => d.md_path));

  // 用全量草稿，过滤掉已被其他账号占用的
  const candidates = _allDrafts.filter(d => !assigned.has(d.path) || selfAssigned.has(d.path));

  document.getElementById('draftPickTitle').textContent = `选择文案（${account}）`;
  const list = document.getElementById('draftPickList');
  if (!candidates.length) {
    list.innerHTML = '<div style="color:#999;font-size:.85rem;padding:8px 0">暂无可选文案</div>';
  } else {
    list.innerHTML = candidates.map(d => {
      const checked = selfAssigned.has(d.path) ? 'checked' : '';
      const isReady = readyPaths.has(d.path);
      const badge = isReady ? '<span style="font-size:.7rem;background:#e8f5e9;color:#2e7d32;padding:1px 6px;border-radius:99px;margin-left:6px">已就绪</span>' : '';
      const imageBadge = (d.card_count || 0) > 0
        ? '<span style="font-size:.7rem;background:#eef6ff;color:#1565c0;padding:1px 6px;border-radius:99px;margin-left:6px">已保存图片</span>'
        : '<span style="font-size:.7rem;background:#f5f5f5;color:#999;padding:1px 6px;border-radius:99px;margin-left:6px">仅文案</span>';
      const preview = d.preview_image
        ? `<img src="${escHtml(d.preview_image)}" style="width:58px;height:78px;object-fit:cover;border-radius:8px;border:1px solid #eee;flex:0 0 auto">`
        : '';
      return `<label style="display:flex;align-items:flex-start;gap:8px;padding:8px 0;border-bottom:1px solid #f3f3f3;cursor:pointer">
        <input type="checkbox" value="${escHtml(d.path)}" ${checked} style="margin-top:3px">
        ${preview}
        <div>
          <div style="font-size:.85rem;color:#333">${escHtml(d.title || d.name)}${badge}${imageBadge}</div>
          <div style="font-size:.75rem;color:#bbb">${escHtml(d.name)}</div>
        </div>
      </label>`;
    }).join('');
  }
  document.getElementById('draftPickModal').classList.add('open');
}

function closeDraftPickModal() {
  document.getElementById('draftPickModal').classList.remove('open');
}

async function confirmDraftPick() {
  const paths = [];
  document.querySelectorAll('#draftPickList input:checked').forEach(cb => paths.push(cb.value));

  // 自动标记未就绪的草稿
  const readyPaths = new Set(_readyDrafts.map(d => d.md_path));
  for (const path of paths) {
    if (readyPaths.has(path)) continue;
    const draft = _allDrafts.find(d => d.path === path);
    const title = (draft?.title || draft?.name || path.split('/').pop().replace('.md','')).slice(0, 20);
    try {
      const r = await fetch('/api/mark_ready', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({md_path: path, title, desc: '', image_path: '', account: _draftPickAccount, image_style: '', writing_style: ''})
      });
      const d = await r.json();
      if (d.error) showToast('标记就绪失败: ' + d.error, 'err');
    } catch(e) { showToast('标记就绪失败: ' + e.message, 'err'); }
  }

  if (paths.length) _pubAssignments[_draftPickAccount] = paths;
  else delete _pubAssignments[_draftPickAccount];
  await loadReadyDrafts();
  closeDraftPickModal();
  renderPublishAssignments();
}

function confirmPublish() {
  const items = [];
  Object.entries(_pubAssignments).forEach(([account, paths]) => {
    paths.forEach(draft_path => items.push({draft_path, account}));
  });
  if (!items.length) {
    showToast('请先为账号分配就绪文案', 'err');
    return;
  }
  const btn = document.getElementById('confirmBtn');
  btn.disabled = true;
  btn.textContent = '发布中...';

  const payload = {items};
  if (_pubMode === 'sched') {
    const date = document.getElementById('schedDate').value;
    const time = document.getElementById('schedTime').value;
    payload.scheduled_at = date + 'T' + time + ':00';
  }

  fetch('/api/publish_batch', {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify(payload)
  }).then(r => r.json()).then(data => {
    btn.disabled = false;
    btn.textContent = '确认发布';
    if (!data.ok) {
      showToast((_pubMode === 'sched' ? '提交失败：' : '发布失败：') + (data.error || ''), 'err');
      return;
    }
    closePublishModal();
    _pubAssignments = {};
    renderPublishAssignments();
    loadReadyDrafts();
    loadScheduledTasks();
    showToast(_pubMode === 'sched' ? '已加入定时队列！' : '批量发布已提交', 'ok');
  }).catch(() => {
    btn.disabled = false;
    btn.textContent = '确认发布';
    showToast('请求失败', 'err');
  });
}

function cancelScheduledTask(id) {
  if (!confirm('确认取消该定时任务？')) return;
  fetch('/api/scheduled_tasks/' + id, {method: 'DELETE'})
    .then(r => r.json()).then(() => loadScheduledTasks());
}

function loadScheduledTasks() {
  fetch('/api/scheduled_tasks').then(r => r.json()).then(tasks => {
    const tbody = document.getElementById('schedTbody');
    if (!tasks.length) {
      tbody.innerHTML = '<tr><td colspan="5" style="color:#bbb;text-align:center;padding:16px">暂无定时任务</td></tr>';
      return;
    }
    const statusLabel = {pending:'<span class="badge-pending">待发</span>',
      done:'<span class="badge-done">已发</span>',
      failed:'<span class="badge-failed">失败</span>',
      cancelled:'<span class="badge-cancelled">已取消</span>'};
    tbody.innerHTML = tasks.map(t => {
      const accText = t.items && t.items.length
        ? [...new Set(t.items.map(i => i.account).filter(Boolean))].join(', ')
        : (t.accounts && t.accounts.length ? t.accounts.join(', ') : '全部');
      const titleText = t.items && t.items.length
        ? [...new Set(t.items.map(i => i.title || PathFallback(i.draft_path)).filter(Boolean))].join(' / ')
        : ((t.title || '').slice(0, 20) || PathFallback(t.folder || ''));
      const cancelBtn = t.status === 'pending'
        ? `<button class="btn btn-outline" style="padding:4px 12px;font-size:.78rem" onclick="cancelScheduledTask('${t.id}')">取消</button>`
        : '—';
      return '<tr><td>' + t.scheduled_at.replace('T', ' ').slice(0, 16) + '</td>' +
        '<td>' + escHtml(titleText).slice(0, 40) + '</td>' +
        '<td>' + accText + '</td>' +
        '<td>' + (statusLabel[t.status] || t.status) + '</td>' +
        '<td>' + cancelBtn + '</td></tr>';
    }).join('');
  });
}

window.addEventListener('load', () => {
  loadScheduledTasks();
  refreshDraftList();
  loadReadyDrafts();
  loadManagedAccounts();
});
</script>
</body>
</html>
"""


def main():
    port = 8888
    httpd = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"\n🍠 小红书自动运营系统")
    print(f"{'='*50}")
    print(f"📍 http://localhost:{port}")
    print(f"{'='*50}\n")
    if "--no-browser" not in sys.argv:
        webbrowser.open(f"http://localhost:{port}")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n👋 已关闭")


if __name__ == "__main__":
    setup_file_logging()
    main()
