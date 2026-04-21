#!/usr/bin/env python3
"""
小红书爆款账号发现 + 分析 Pipeline

步骤：
1. 调用真实 explore.js（有头浏览器）抓取关键词爆款账号
2. 去重：跳过 output/style_models/ 中已分析的账号
3. 取前 5 个新账号，调用 analyze_accounts.py 分析
4. 打印两张表格：Explore 结果 + 分析结果

用法：
    python scripts/explore_and_analyze.py "英国留学求职" [--count 5]
"""

import json
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from flowlens_runtime import get_flowlens_python
from flowlens_xhs_backend import discover_accounts_via_flowlens

BASE_DIR = Path(__file__).parent.parent
STYLE_MODELS_DIR = BASE_DIR / "output" / "style_models"
ANALYSIS_XLSX = BASE_DIR / "output" / "爆款分析" / "xhs_account_analysis.xlsx"
SCRIPTS_DIR = BASE_DIR / "scripts"


def get_analyzed_names() -> set:
    """已分析账号昵称集合（来自 output/style_models/*.json）"""
    if not STYLE_MODELS_DIR.exists():
        return set()
    names = set()
    for p in STYLE_MODELS_DIR.glob("*.json"):
        # 文件名就是昵称（经过 sanitize），直接用于模糊去重
        names.add(p.stem)
    return names


def run_explore(keyword: str, search_limit: int = 30,
                viral_threshold: int = 50, author_limit: int = 20) -> dict | None:
    """调用 FlowLens 检索接口获取真实搜索结果。"""
    print(f"\n{'='*60}")
    print(f"[EXPLORE] 关键词：{keyword}")
    print(f"[EXPLORE] searchLimit={search_limit}  viralThreshold={viral_threshold}  authorLimit={author_limit}")
    print(f"[EXPLORE] 使用 FlowLens XHS 接口")
    print(f"{'='*60}\n")
    try:
        return discover_accounts_via_flowlens(
            keyword,
            search_limit=search_limit,
            viral_threshold=viral_threshold,
            author_limit=author_limit,
        )
    except Exception as exc:
        print(f"[ERROR] FlowLens explore 失败: {exc}")
        return None


def deduplicate(authors: list, analyzed_names: set, target: int = 5) -> list:
    """
    去重：跳过昵称（或其 sanitize 版）已在 analyzed_names 中的账号。
    返回最多 target 个新账号。
    """
    def sanitize(name: str) -> str:
        import re
        safe = re.sub(r'[/\\:*?"<>|]', '', name)
        safe = safe.replace(' ', '_').strip()
        return safe or 'unknown_account'

    result = []
    skipped = []
    for a in authors:
        nick = a.get('author') or a.get('nickname', '')
        safe = sanitize(nick)
        if nick in analyzed_names or safe in analyzed_names:
            skipped.append(nick)
            continue
        result.append(a)
        if len(result) >= target:
            break

    if skipped:
        print(f"[去重] 跳过已分析账号：{skipped}")
    print(f"[去重] 新账号 {len(result)} 个（目标 {target} 个）")
    return result


def print_explore_table(authors: list, keyword: str):
    """打印 Explore 检索结果表格"""
    print(f"\n{'='*70}")
    print(f"📊 EXPLORE 检索结果 — 关键词「{keyword}」")
    print(f"{'='*70}")
    header = f"{'#':<4} {'账号昵称':<22} {'爆款数':<7} {'最高爆款分':<11} {'代表标题'}"
    print(header)
    print('-' * 70)
    for i, a in enumerate(authors, 1):
        nick   = (a.get('author') or a.get('nickname', ''))[:20]
        viral  = a.get('viral_count', 0)
        score  = a.get('max_viral_score', 0)
        title  = (a.get('top_post_title') or '')[:24]
        print(f"{i:<4} {nick:<22} {viral:<7} {score:<11.1f} {title}")
    print()


def build_accounts_json(authors: list, keyword: str) -> list:
    """将 explore 结果转为 analyze_accounts.py 所需格式"""
    accounts = []
    for a in authors:
        nick = a.get('author') or a.get('nickname', '')
        # 从 viral_notes 或 all_notes 取代表笔记
        viral = a.get('viral_notes') or a.get('all_notes') or []
        top_post = viral[0] if viral else {}
        top_title = top_post.get('title') or a.get('top_post_title', '')
        top_url   = top_post.get('url') or top_post.get('link') or a.get('top_post_url', '')
        internal_id = a.get('user_id', '')
        accounts.append({
            "name": nick,
            "xhs_id": "",
            "user_id": internal_id,   # 直接传内部 ID，跳过 search_user_id 搜索
            "author_hex_id": a.get("author_hex_id", ""),
            "profile_url": a.get("profile_url", ""),
            "status": "爬取中",
            "explore_data": {
                "keyword": keyword,
                "viral_count": a.get('viral_count', 0),
                "max_viral_score": a.get('max_viral_score', 0),
                "top_post_title": top_title,
                "top_post_url": top_url
            }
        })
    return accounts


def run_analyze(accounts: list) -> list:
    """调用 analyze_accounts.py，返回分析结果行"""
    with tempfile.NamedTemporaryFile(
        mode='w', suffix='.json', delete=False, encoding='utf-8'
    ) as f:
        json.dump(accounts, f, ensure_ascii=False, indent=2)
        tmp_path = f.name

    print(f"\n{'='*60}")
    print(f"[ANALYZE] 分析 {len(accounts)} 个账号（有头浏览器）...")
    print(f"[ANALYZE] accounts-file: {tmp_path}")
    print(f"{'='*60}\n")

    try:
        proc = subprocess.run(
            [get_flowlens_python(), str(SCRIPTS_DIR / 'analyze_accounts.py'),
             '--accounts-file', tmp_path],
            cwd=str(BASE_DIR),
            timeout=600
        )
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    if proc.returncode != 0:
        print(f"[WARN] analyze_accounts.py 返回码 {proc.returncode}")

    # 读取最新生成的 Excel 结果，但更重要的是直接从 style_models 读
    return []


def print_analyze_table(accounts: list):
    """打印 analyze 汇总表（从 style_models 读取刚生成的数据）"""
    from openpyxl import load_workbook
    if not ANALYSIS_XLSX.exists():
        print("[WARN] 未找到分析 Excel 文件")
        return

    latest = str(ANALYSIS_XLSX)
    print(f"\n{'='*90}")
    print(f"📋 ANALYZE 分析结果 — 来源：{Path(latest).name}")
    print(f"{'='*90}")

    try:
        wb = load_workbook(latest, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("[WARN] Excel 无数据")
            return

        # 找我们要的列
        headers = list(rows[0])
        want_cols = ['账号昵称', '粉丝数', '笔记总数', '平均点赞',
                     '平均收藏', '账号权重评级', '爆款笔记数', '最高爆款评分']
        idx = {}
        for col in want_cols:
            try:
                idx[col] = headers.index(col)
            except ValueError:
                pass

        # 只打印本次分析的账号
        target_names = {a['name'] for a in accounts}
        print_rows = [r for r in rows[1:] if r[idx.get('账号昵称', 0)] in target_names]

        if not print_rows:
            # 如果找不到精确匹配，打最后 N 行
            print_rows = rows[max(1, len(rows) - len(accounts)):]

        col_w = {'账号昵称': 22, '粉丝数': 10, '笔记总数': 8,
                 '平均点赞': 8, '平均收藏': 8, '账号权重评级': 14,
                 '爆款笔记数': 8, '最高爆款评分': 10}

        # 表头
        header_line = ''
        for col in want_cols:
            w = col_w.get(col, 10)
            if col in idx:
                header_line += f"{col:{w}}"
        print(header_line)
        print('-' * 90)

        for row in print_rows:
            line = ''
            for col in want_cols:
                w = col_w.get(col, 10)
                if col in idx:
                    val = str(row[idx[col]] or '')[:w-1]
                    line += f"{val:{w}}"
            print(line)
        print()

    except Exception as e:
        print(f"[WARN] 读取 Excel 失败：{e}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="小红书爆款账号发现 + 分析")
    parser.add_argument('keyword', nargs='?', default='英国留学求职')
    parser.add_argument('--count', type=int, default=5, help='目标新账号数（默认5）')
    parser.add_argument('--search-limit', type=int, default=30)
    parser.add_argument('--viral-threshold', type=int, default=50)
    parser.add_argument('--author-limit', type=int, default=25)
    args = parser.parse_args()

    print(f"\n🚀 小红书爆款账号发现+分析 Pipeline")
    print(f"   关键词: {args.keyword}")
    print(f"   目标账号: {args.count} 个（已排除历史分析账号）")
    print(f"   时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Step 1: 已分析账号集合
    analyzed = get_analyzed_names()
    print(f"\n[去重] 已分析账号数：{len(analyzed)}")

    # Step 2: explore
    explore_result = run_explore(
        keyword=args.keyword,
        search_limit=args.search_limit,
        viral_threshold=args.viral_threshold,
        author_limit=args.author_limit
    )
    if not explore_result:
        print("❌ explore 失败，程序退出")
        sys.exit(1)

    all_authors = explore_result.get('authors', [])
    print(f"\n[EXPLORE] 总抓取笔记: {explore_result.get('total_fetched', 0)}")
    print(f"[EXPLORE] 爆款通过: {explore_result.get('viral_passed', 0)}")
    print(f"[EXPLORE] 发现作者: {len(all_authors)}")

    # 打印 explore 表（全部）
    print_explore_table(all_authors, args.keyword)

    # Step 3: 去重，取前 N 个
    new_authors = deduplicate(all_authors, analyzed, target=args.count)
    if not new_authors:
        print("❌ 所有发现账号均已分析过，无新账号可处理")
        sys.exit(0)

    # Step 4: 构建 accounts 格式
    accounts = build_accounts_json(new_authors, args.keyword)

    print(f"\n[分析] 即将分析以下 {len(accounts)} 个新账号：")
    for i, a in enumerate(accounts, 1):
        print(f"  {i}. {a['name']}")

    # Step 5: analyze
    run_analyze(accounts)

    # Step 6: 打印分析结果表
    print_analyze_table(accounts)

    print("✅ Pipeline 完成！")


if __name__ == '__main__':
    main()
