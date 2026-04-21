#!/usr/bin/env python3
"""
XHS 账号分析报表生成脚本

运行方式：
    python scripts/analyze_accounts.py

依赖：
    pip install xhs==0.2.13 openpyxl anthropic python-dotenv

输出：
    output/爆款分析/xhs_account_analysis.xlsx
    output/style_models/*.json
"""

import json
import os
import re
import sys
import time
import shutil
import random
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict
from urllib.parse import urlparse

sys.path.insert(0, str(Path(__file__).parent))

from flowlens_xhs_backend import FlowLensXhsFetcher


class _PublishManagerStub:
    """publish_manager 导入失败时的降级替身，避免分析流程被发布依赖阻断。"""

    def __init__(self, import_error: str = ""):
        self.import_error = import_error

    def save_account_style(self, *args, **kwargs):
        raise RuntimeError(self.import_error or "publish_manager 不可用")


try:
    import publish_manager as _publish_manager
    publish_manager = _publish_manager
except BaseException as e:
    publish_manager = _PublishManagerStub(str(e))

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*args, **kwargs):
        pass

# ── 账号清单 ──────────────────────────────────────────────────────────────────
# 账号列表已改为动态传入，通过 --accounts-json 参数从 xhs-explore 结果获取
# 不再使用硬编码账号，保留空列表作为兜底（直接运行脚本时无账号可分析）
ACCOUNTS = []

NOTES_LIMIT = 6  # 6 篇足够识别内容模式，比 10 篇节省 ~40% 抓取时间
BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / "output"
ANALYSIS_DIR = OUTPUT_DIR / "爆款分析"
ANALYSIS_XLSX_PATH = ANALYSIS_DIR / "xhs_account_analysis.xlsx"


# ── 纯函数层（可单元测试）────────────────────────────────────────────────────

def sanitize_filename(name: str) -> str:
    """去掉非法字符，返回安全的文件名"""
    try:
        illegal_chars = r'[/\\:*?"<>|]'
        safe_name = re.sub(illegal_chars, '', name)
        safe_name = safe_name.replace(' ', '_')
        safe_name = safe_name.strip()
        if not safe_name:
            return "unknown_account"
        return safe_name
    except Exception:
        return "unknown_account"


def filter_notes_by_recent_month(notes: List[Dict], days: int = 30) -> List[Dict]:
    """
    仅保留近 N 天内的笔记（基于系统时间）
    
    Args:
        notes: 笔记列表
        days: 保留天数（默认 30 天 = 1 个月）
    
    Returns:
        已过滤的笔记列表
    """
    if not notes:
        return []
    
    try:
        now = datetime.now()
        cutoff = now - timedelta(days=days)
        
        filtered = []
        for note in notes:
            # 尝试从多个可能的时间字段获取
            create_time = (
                note.get("create_time") or
                note.get("created_at") or
                note.get("publish_time") or
                note.get("time") or
                None
            )
            
            if not create_time:
                # 保守策略：缺失时间时默认保留，避免静默丢失可用样本；同时打标供调用方识别
                flagged = dict(note)
                flagged["create_time_missing"] = True
                filtered.append(flagged)
                continue
            
            # 解析时间戳
            try:
                if isinstance(create_time, str):
                    # ISO 格式或时间戳字符串
                    if 'T' in create_time or '-' in create_time:
                        note_dt = datetime.fromisoformat(create_time.replace('Z', '+00:00'))
                    else:
                        # 假设是秒级时间戳
                        note_dt = datetime.fromtimestamp(int(create_time))
                else:
                    # 假设是毫秒级或秒级时间戳
                    ts = int(create_time)
                    if ts > 10**12:  # 毫秒时间戳
                        note_dt = datetime.fromtimestamp(ts / 1000)
                    else:  # 秒级时间戳
                        note_dt = datetime.fromtimestamp(ts)
                
                # 移除时区信息以便比较
                if note_dt.tzinfo:
                    note_dt = note_dt.replace(tzinfo=None)
                
                if note_dt >= cutoff:
                    filtered.append(note)
                else:
                    print(f"    ⏭️  跳过旧笔记 (ID: {note.get('note_id', '')[:8]}, 时间: {note_dt.strftime('%Y-%m-%d')})")
            except Exception as e:
                print(f"    ⚠️  时间解析异常 ({create_time}): {e}, 保留该笔记")
                flagged = dict(note)
                flagged["create_time_missing"] = True
                filtered.append(flagged)
        
        print(f"  📅 时间过滤: {len(notes)} → {len(filtered)} 篇 (近 {days} 天)")
        return filtered
    
    except Exception as e:
        print(f"  ❌ 时间过滤异常: {e}，保留全部笔记")
        return notes


def filter_high_performance_notes(notes: List[Dict]) -> List[Dict]:
    """过滤出高表现的笔记（点赞数 > 平均值 * 1.2）"""
    try:
        if not notes:
            return []
        total_likes = sum(note.get("liked_count", 0) for note in notes)
        avg_like = total_likes / len(notes) if notes else 0
        threshold = avg_like * 1.2
        filtered = [note for note in notes if note.get("liked_count", 0) > threshold]
        if not filtered:
            return notes
        return filtered
    except Exception:
        return notes


def build_style_model(notes: List[Dict]) -> Dict:
    """从笔记列表构建风格模型"""
    try:
        result = {
            "title_templates": [],
            "hook_templates": [],
            "content_structure": ["岗位亮点", "要求门槛", "投递方式"],
            "cta_templates": [],
            "tone": "紧迫+口语化",
            "image_style": "单图大字信息流"
        }
        
        if not notes:
            return result
        
        titles = [note.get("title", "").strip() for note in notes if note.get("title")]
        titles = [t for t in titles if t]
        titles = list(dict.fromkeys(titles))
        result["title_templates"] = titles[:3]
        
        hooks = []
        for note in notes:
            content = (note.get("desc") or note.get("content") or "").strip()
            if content:
                first_line = content.split('\n')[0][:30]
                if first_line:
                    hooks.append(first_line)
        hooks = list(dict.fromkeys(hooks))
        result["hook_templates"] = hooks[:3]
        
        ctas = []
        for note in notes:
            content = (note.get("desc") or note.get("content") or "").strip()
            if content:
                lines = content.split('\n')
                last_line = lines[-1].strip() if lines else ""
                if last_line:
                    ctas.append(last_line)
        ctas = list(dict.fromkeys(ctas))
        result["cta_templates"] = ctas[:3]
        
        return result
    
    except Exception as e:
        print(f"[Error] build_style_model 异常: {e}")
        return {
            "title_templates": [],
            "hook_templates": [],
            "content_structure": ["岗位亮点", "要求门槛", "投递方式"],
            "cta_templates": [],
            "tone": "紧迫+口语化",
            "image_style": "单图大字信息流"
        }


def save_style_model(account_name: str, style_model: Dict) -> str:
    """保存风格模型为 JSON 文件"""
    try:
        output_dir = Path("output/style_models")
        output_dir.mkdir(parents=True, exist_ok=True)
        safe_name = sanitize_filename(account_name)
        file_path = output_dir / f"{safe_name}.json"
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(style_model, f, ensure_ascii=False, indent=2)
        return str(file_path)
    except Exception as e:
        print(f"[Error] save_style_model 异常: {e}")
        return ""


def strip_json_fences(text: str) -> str:
    """去除 Claude 返回的 Markdown 代码块包装"""
    text = text.strip()
    match = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
    if match:
        return match.group(1).strip()
    return text


def compute_stats(notes: list) -> dict:
    """从笔记列表计算统计指标"""
    if not notes:
        return {
            "avg_images": 0, "avg_words": 0,
            "avg_liked": 0, "avg_collected": 0, "avg_comments": 0,
        }
    n = len(notes)
    return {
        "avg_images":    round(sum((x.get("image_count") or 0) for x in notes) / n, 1),
        "avg_words":     sum((x.get("word_count") or 0) for x in notes) // n,
        "avg_liked":     sum((x.get("liked_count") or 0) for x in notes) // n,
        "avg_collected": sum((x.get("collected_count") or 0) for x in notes) // n,
        "avg_comments":  sum((x.get("comment_count") or 0) for x in notes) // n,
    }


def format_image_preference(notes: list) -> str:
    """根据图片数量分布生成偏好描述"""
    if not notes:
        return "无数据"
    counts = [x.get("image_count", 0) for x in notes]
    single = sum(1 for c in counts if c <= 1)
    multi  = len(counts) - single
    avg    = round(sum(counts) / len(counts), 1)
    if single == 0:
        return f"多图为主（均{avg}张）"
    if multi == 0:
        return "单图为主"
    return f"单图/多图混合（均{avg}张）"


def compute_interaction_rate(avg_liked: int, avg_collected: int, fans_count: int) -> str:
    """计算互动率"""
    if fans_count == 0:
        return "N/A"
    rate = (avg_liked + avg_collected) / fans_count * 100
    return f"{rate:.1f}%"


def compute_stickiness(avg_collected: int, avg_liked: int) -> str:
    """计算用户粘度"""
    if avg_liked == 0:
        return "N/A"
    rate = avg_collected / avg_liked * 100
    return f"{rate:.1f}%"


def is_valid_detailed_note(note: dict) -> bool:
    """只有详情页真正抓取完成的笔记才参与后续统计与分析。"""
    return bool(note.get("detail_fetched"))


def is_complete_analysis_row(row: dict) -> bool:
    """判断该行是否属于完整可用的账号分析结果。"""
    skip_ratings = {
        "低活跃度-无法获取详情",
        "未找到账号",
        "无笔记数据",
        "无近期笔记",
        "低活跃度",
        "分析失败",
    }
    rating = str(row.get("账号权重评级") or "")
    return bool(rating) and rating not in skip_ratings and rating != "N/A"


def _extract_cover_url(note: dict) -> str:
    """从 user_posted note 结构中尽量提取可用封面图 URL。"""
    if not isinstance(note, dict):
        return ""
    cover = note.get("cover") or {}
    candidates = []
    if isinstance(cover, dict):
        candidates.extend([
            cover.get("url_default"),
            cover.get("url_pre"),
            cover.get("url"),
        ])
        url_list = cover.get("url_list") or []
        if isinstance(url_list, list) and url_list:
            candidates.append(url_list[0])
    image_list = note.get("image_list") or []
    if isinstance(image_list, list) and image_list:
        first = image_list[0] or {}
        if isinstance(first, dict):
            candidates.extend([
                first.get("url_default"),
                first.get("url_pre"),
                first.get("url"),
            ])
    for c in candidates:
        if isinstance(c, str) and c.strip():
            return c.strip()
    return ""


def _is_safe_http_url(url: str) -> bool:
    """仅允许 http/https 且非本地回环地址，避免无效/风险 image_url。"""
    if not isinstance(url, str) or not url.strip():
        return False
    try:
        u = urlparse(url.strip())
    except Exception:
        return False
    if u.scheme not in {"http", "https"}:
        return False
    host = (u.hostname or "").lower()
    if not host:
        return False
    if host in {"localhost", "127.0.0.1", "0.0.0.0", "::1"}:
        return False
    if host.endswith(".local"):
        return False
    return True


def _parse_analysis_output(raw: str) -> dict:
    """解析模型输出，兼容 pipe/冒号/全角分隔以及 JSON 返回。"""
    if not isinstance(raw, str):
        return {}
    text = raw.strip()
    if not text:
        return {}

    # 1) 先尝试 JSON（含 markdown fenced json）
    try:
        maybe_json = strip_json_fences(text)
        obj = json.loads(maybe_json)
        if isinstance(obj, dict):
            parsed = {}
            for k, v in obj.items():
                if k in DEFAULT_ANALYSIS and isinstance(v, (str, int, float)):
                    parsed[k] = str(v).strip()
            if parsed:
                return parsed
    except Exception:
        pass

    # 2) 行解析：支持 "｜"、":"、"：" 分隔
    parsed = {}
    normalized = text.replace("｜", "|")
    for raw_line in normalized.splitlines():
        line = raw_line.strip().lstrip("-•*").strip()
        if not line:
            continue

        key = ""
        val = ""
        if "|" in line:
            key, _, val = line.partition("|")
        elif "：" in line:
            key, _, val = line.partition("：")
        elif ":" in line:
            key, _, val = line.partition(":")
        else:
            continue
        key = key.strip().replace(" ", "")
        val = val.strip()
        if key in DEFAULT_ANALYSIS and val:
            parsed[key] = val
    return parsed


# ── XHS 抓取层（FlowLens）────────────────────────────────────────────────────

def _load_cookie(preferred_account: str = "") -> tuple:
    """兼容旧测试/调用方的占位接口；实际抓取已不再依赖 Cookie。"""
    del preferred_account
    raise RuntimeError("XHS Playwright/Cookie 抓取路径已移除，请使用 FlowLensXhsFetcher")


# 保留旧名字仅用于兼容导入；实际实现已完全切换到 FlowLens。
XhsFetcher = FlowLensXhsFetcher


# ── Claude 分析层 ─────────────────────────────────────────────────────────────

CLAUDE_PROMPT_TEMPLATE = """你是一名小红书运营分析师。以下是账号「{nickname}」最近{n}篇笔记内容：

{content}

请分析该账号，用以下固定格式回答，每行一个字段，字段名和内容用"|"分隔，不要有多余说明：
图文风格描述|不超过100字，描述视觉风格、配色偏好、排版习惯、文案语气
内容偏好|不超过30字，高频话题方向
受众分析|不超过30字，主要受众群体特征
目标对象关联点|不超过30字，内容与目标用户的痛点/需求连接
账号权重评级|A/B/C加不超过20字理由
Hook模板|提炼2-3个开头钩子句式，用" | "分隔，如：刷到就是缘分！ | 我踩过的坑你别再踩
内容结构|该账号笔记的通用内容框架，3-5个环节，用" | "分隔，如：痛点引入 | 解决方案 | 经验总结
CTA模板|提炼2-3个结尾行动引导句式，用" | "分隔，如：评论区告诉我你的情况 | 点收藏备用"""

DEFAULT_ANALYSIS = {
    "图文风格描述":   "分析失败",
    "内容偏好":       "分析失败",
    "受众分析":       "分析失败",
    "目标对象关联点": "分析失败",
    "账号权重评级":   "分析失败",
    "Hook模板":       "分析失败",
    "内容结构":       "分析失败",
    "CTA模板":        "分析失败",
}


def _load_dashscope_key() -> tuple:
    """返回 (api_key, base_url, model)，优先 DASHSCOPE_API_KEY，其次 MIDSCENE_MODEL_API_KEY"""
    for env_path in [Path.cwd() / ".env", Path(__file__).parent.parent / ".env"]:
        if env_path.exists():
            load_dotenv(env_path)
            break
    key = os.getenv("DASHSCOPE_API_KEY", "")
    if key:
        vision_model = os.getenv("VISION_MODEL", "") or "qwen-vl-max"
        return key, "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions", vision_model
    # 备选：MIDSCENE_MODEL_API_KEY
    key = os.getenv("MIDSCENE_MODEL_API_KEY", "")
    base = os.getenv("MIDSCENE_MODEL_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1")
    model = os.getenv("VISION_MODEL", "") or os.getenv("MIDSCENE_MODEL_NAME", "qwen-vl-max")
    if base and not base.endswith("/chat/completions"):
        base = base.rstrip("/") + "/chat/completions"
    return key, base, model


def analyze_with_claude(notes: list, nickname: str) -> dict:
    """通过 DashScope qwen-vl-max 视觉模型分析笔记风格（含封面图片）"""
    import urllib.request, urllib.error, json as _json

    api_key, base_url, model_name = _load_dashscope_key()
    if not api_key:
        print("  ⚠️  未配置 DASHSCOPE_API_KEY 或 MIDSCENE_MODEL_API_KEY，跳过分析")
        return DEFAULT_ANALYSIS.copy()

    # 构建多模态消息：每篇笔记包含封面图 + 文字
    # 支持图片的模型白名单（vl系列 + qwen3.5-plus + kimi-k2.5 + MiniMax-M2.5）
    _vision_keywords = ("vl", "vision", "qwen3.5-plus", "kimi-k2", "minimax-m2")
    supports_vision = any(k in model_name.lower() for k in _vision_keywords)
    note_blocks = []
    image_count = 0
    max_cover_images = 1
    for i, note in enumerate(notes, 1):
        title    = note.get("title", "")
        desc     = note.get("desc", "")
        cover    = note.get("cover_url", "")
        header   = f"[{i}] 标题：{title}\n正文：{desc[:500]}"
        note_blocks.append({"type": "text", "text": header})
        if supports_vision and _is_safe_http_url(cover):
            if image_count < max_cover_images:
                note_blocks.append({"type": "image_url", "image_url": {"url": cover}})
                image_count += 1
            else:
                print(f"  ℹ️  封面图超出上限({max_cover_images})，跳过: {cover[:120]}")

    print(f"  🧠 AI分析({model_name}): notes={len(notes)}, images={image_count}")

    instruction = (
        f"你是一名小红书运营分析师。以上是账号「{nickname}」最近{len(notes)}篇笔记内容（含封面图）。\n\n"
        "请分析该账号，用以下固定格式回答，每行一个字段，字段名和内容用\"|\"分隔，不要有多余说明：\n"
        "图文风格描述|不超过100字，描述视觉风格、配色偏好、排版习惯、文案语气\n"
        "内容偏好|不超过30字，高频话题方向\n"
        "受众分析|不超过30字，主要受众群体特征\n"
        "目标对象关联点|不超过30字，内容与目标用户的痛点/需求连接\n"
        "账号权重评级|A/B/C加不超过20字理由\n"
        "Hook模板|提炼2-3个开头钩子句式，用\" | \"分隔，如：刷到就是缘分！ | 我踩过的坑你别再踩\n"
        "内容结构|该账号笔记的通用内容框架，3-5个环节，用\" | \"分隔，如：痛点引入 | 解决方案 | 经验总结\n"
        "CTA模板|提炼2-3个结尾行动引导句式，用\" | \"分隔，如：评论区告诉我你的情况 | 点收藏备用"
    )

    messages = [
        {
            "role": "user",
            "content": note_blocks + [{"type": "text", "text": instruction}],
        }
    ]

    payload = _json.dumps({
        "model": model_name,
        "messages": messages,
        "temperature": 0.2,
    }).encode("utf-8")

    req = urllib.request.Request(
        base_url,
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    for attempt in range(2):
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = _json.loads(resp.read().decode("utf-8"))
            raw = (((data or {}).get("choices") or [{}])[0].get("message") or {}).get("content", "")
            if isinstance(raw, list):
                raw = " ".join(b.get("text", "") for b in raw if isinstance(b, dict))
            # 去掉思维链 <think>...</think> 块（推理模型如 qwen3.5-plus 会输出）
            import re as _re
            raw_clean = _re.sub(r"<think>[\s\S]*?</think>", "", str(raw)).strip()
            parsed = _parse_analysis_output(raw_clean)
            if parsed:
                return {k: parsed.get(k, "分析失败") for k in DEFAULT_ANALYSIS}
            print("  ⚠️  DashScope 返回可解析字段为空，回退默认分析")
            return DEFAULT_ANALYSIS.copy()
        except urllib.error.HTTPError as e:
            body = ""
            try:
                body = e.read().decode("utf-8", errors="ignore")[:300]
            except Exception:
                body = ""
            print(f"  ⚠️  DashScope HTTP错误({e.code}) attempt={attempt+1}/2: {body}")
            if attempt == 0:
                time.sleep(1.2)
                continue
            return DEFAULT_ANALYSIS.copy()
        except Exception as e:
            print(f"  ⚠️  DashScope 分析失败 attempt={attempt+1}/2: {e}")
            if attempt == 0:
                time.sleep(1.2)
                continue
            return DEFAULT_ANALYSIS.copy()

    return DEFAULT_ANALYSIS.copy()


# ── Excel 写入层 ──────────────────────────────────────────────────────────────

HEADERS = [
    "账号昵称", "状态",
    "搜索关键词", "爆款笔记数", "最高爆款评分", "代表爆款标题", "代表爆款链接",
    "粉丝数", "笔记总数", "分析笔记篇数",
    "平均图片张数", "图片偏好", "平均文字字数",
    "图文风格描述", "内容偏好",
    "平均点赞数", "平均收藏数", "平均评论数",
    "互动率", "用户粘度",
    "受众分析", "目标对象关联点", "账号权重评级",
    "标题模板", "Hook模板", "内容结构", "CTA模板", "情绪标签", "图片风格",
]


def write_excel(rows: list, output_path: str) -> None:
    """将分析结果写入 Excel 文件"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "账号分析"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            val  = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 列宽按列名定义，与 HEADERS 顺序无关，新增/调整列不会错位
    col_width_map = {
        "账号昵称": 20, "状态": 10,
        "搜索关键词": 14, "爆款笔记数": 10, "最高爆款评分": 12,
        "代表爆款标题": 30, "代表爆款链接": 50,
        "粉丝数": 10, "笔记总数": 10, "分析笔记篇数": 12,
        "平均图片张数": 12, "图片偏好": 12, "平均文字字数": 12,
        "图文风格描述": 45, "内容偏好": 25,
        "平均点赞数": 12, "平均收藏数": 12, "平均评论数": 12,
        "互动率": 10, "用户粘度": 10,
        "受众分析": 25, "目标对象关联点": 25, "账号权重评级": 25,
        "标题模板": 25, "Hook模板": 25, "内容结构": 25,
        "CTA模板": 25, "情绪标签": 15, "图片风格": 20,
    }
    for col_idx, header in enumerate(HEADERS, 1):
        width = col_width_map.get(header, 15)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A2"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel 已保存：{output_path}")


def get_analysis_workbook_path() -> Path:
    """统一的累计分析工作簿路径。"""
    return ANALYSIS_XLSX_PATH


# ── 主编排函数 ─────────────────────────────────────────────────────────────────

def process_account(fetcher, account: dict, explore_data: dict = None) -> dict:
    """处理单个账号，返回 Excel 行 dict"""
    nickname = account["name"]
    xhs_id   = account.get("xhs_id", "")
    status   = account["status"]
    print(f"\n  处理账号：{nickname}...")

    row = {h: "N/A" for h in HEADERS}
    row["账号昵称"] = nickname
    row["状态"]     = status

    # 填充 xhs-explore 已采集的爆款数据
    if explore_data:
        row["搜索关键词"]   = explore_data.get("keyword", "")
        row["爆款笔记数"]   = explore_data.get("viral_count", "")
        row["最高爆款评分"] = explore_data.get("max_viral_score", "")
        row["代表爆款标题"] = explore_data.get("top_post_title", "")
        row["代表爆款链接"] = explore_data.get("top_post_url", "")

    profile_url = str(account.get("profile_url") or "").strip()
    _raw_hex = account.get("author_hex_id") or account.get("user_id")
    hex_id = _raw_hex if (_raw_hex and _raw_hex != "profile") else None
    user_ref = profile_url or hex_id or fetcher.search_user_id(nickname, xhs_id)
    if not user_ref:
        row["账号权重评级"] = "未找到账号"
        print(f"  ⚠️  未找到账号，跳过")
        return row
    if profile_url:
        print(f"  🔍 profile_url（explore直传）: {profile_url}")
    elif hex_id:
        print(f"  🔍 user_ref（explore直传）: {hex_id}")
    else:
        print(f"  🔍 user_ref（搜索获取）: {user_ref}")

    # 解析 prefetched_notes 作为兜底数据源，不再提前返回
    _prefetched_fallback = []
    _prefetched = (explore_data or {}).get("prefetched_notes")
    if _prefetched:
        _prefetched_fallback = [
            {
                "title": p.get("title", ""), "desc": p.get("content", ""),
                "word_count": len(p.get("content", "")),
                "image_count": None,
                "liked_count": p.get("likes", 0), "collected_count": p.get("saves", 0),
                "comment_count": None,
                "detail_fetched": False, "cover_url": "",
            }
            for p in _prefetched if p.get("content")
        ]
        if _prefetched_fallback:
            print(f"  ℹ️   explore 已预抓 {len(_prefetched_fallback)} 篇笔记，继续补全粉丝/互动数据")

    profile, notes = fetcher.fetch_profile_and_notes(user_ref, limit=NOTES_LIMIT)
    row["粉丝数"]   = profile["fans_count"]
    row["笔记总数"] = profile["notes_count"]
    print(f"  粉丝: {profile['fans_count']}  笔记: {len(notes)} 篇")

    if not notes:
        row["分析笔记篇数"] = 0
        row["账号权重评级"] = "无笔记数据"
        print(f"  ⚠️  无笔记数据，跳过")
        return row

    # 先过滤时间（仅保留近 1 个月的笔记）
    notes = filter_notes_by_recent_month(notes, days=30)
    if not notes:
        row["分析笔记篇数"] = 0
        row["账号权重评级"] = "无近期笔记"
        print(f"  ⚠️  无近 1 个月内的笔记，跳过分析")
        return row

    detailed = fetcher.fetch_note_details_concurrent(notes[:NOTES_LIMIT])

    # 过滤掉无法获取正文的笔记
    detailed = [n for n in detailed if is_valid_detailed_note(n)]
    print(f"  ✅ 获取了 {len(detailed)} 篇笔记详情")

    if not detailed:
        if _prefetched_fallback:
            print(f"  ⚡ 详情抓取失败，回退使用 explore 预抓数据（{len(_prefetched_fallback)} 篇）")
            detailed = _prefetched_fallback
        else:
            row["分析笔记篇数"] = 0
            row["账号权重评级"] = "低活跃度-无法获取详情"
            print(f"  ⚠️  全部笔记均无正文，跳过分析")
            return row

    _from_prefetched = detailed and all(n.get("detail_fetched") is False for n in detailed)
    stats = compute_stats(detailed)
    row["分析笔记篇数"] = len(detailed)

    if _from_prefetched:
        row["平均图片张数"] = "N/A"
        row["图片偏好"]     = "N/A"
        row["平均评论数"]   = "N/A"
    else:
        row["平均图片张数"] = stats["avg_images"]
        row["图片偏好"]     = format_image_preference(detailed)
        row["平均评论数"]   = stats["avg_comments"]

    row["平均文字字数"] = stats["avg_words"]
    row["平均点赞数"]   = stats["avg_liked"]
    row["平均收藏数"]   = stats["avg_collected"]
    row["互动率"]       = compute_interaction_rate(
        stats["avg_liked"], stats["avg_collected"], profile["fans_count"]
    )
    row["用户粘度"] = compute_stickiness(stats["avg_collected"], stats["avg_liked"])

    analysis = analyze_with_claude(detailed, nickname)
    row["图文风格描述"]    = analysis.get("图文风格描述", "分析失败")
    row["内容偏好"]        = analysis.get("内容偏好", "分析失败")
    row["受众分析"]        = analysis.get("受众分析", "分析失败")
    row["目标对象关联点"]  = analysis.get("目标对象关联点", "分析失败")
    row["账号权重评级"]    = analysis.get("账号权重评级", "分析失败")
    # Hook/内容结构/CTA 由 Claude 语义提炼，比首末行启发式更准确
    row["Hook模板"]  = analysis.get("Hook模板", "分析失败")
    row["内容结构"]  = analysis.get("内容结构", "分析失败")
    row["CTA模板"]   = analysis.get("CTA模板", "分析失败")

    # 过滤高表现笔记，构建基于标题/tone/image_style 的风格模型，保存 JSON
    filtered_notes = filter_high_performance_notes(detailed)
    style_model = build_style_model(filtered_notes)
    save_style_model(nickname, style_model)

    # 标题模板取自笔记真实标题（原始数据），情绪标签/图片风格取 build_style_model 的规则结果
    row["标题模板"] = " | ".join(style_model.get("title_templates", []))
    row["情绪标签"] = style_model.get("tone", "")
    row["图片风格"] = style_model.get("image_style", "")

    return row


def load_existing_excel(path: str) -> list:
    """读取已有 Excel，返回 rows list（每行为 dict）"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows
    except Exception:
        return []


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--names", nargs="+", help="只处理指定账号（支持多个，空格分隔）")
    parser.add_argument("--accounts-json", dest="accounts_json",
                        help="JSON格式账号列表字符串（注意 shell 引号转义风险，推荐用 --accounts-file）")
    parser.add_argument("--accounts-file", dest="accounts_file",
                        help="JSON账号列表文件路径（推荐，避免 shell 引号问题），"
                             "格式同 --accounts-json")
    parser.add_argument("--login-account", dest="login_account", default="",
                        help="指定用于检索/分析的小红书登录账号名")
    parser.add_argument("--target-count", dest="target_count", type=int, default=0,
                        help="目标有效账号数量。设置后将从候选池持续取账号，"
                             "直到获得指定数量的完整数据行（跳过低活跃/无数据账号），"
                             "候选池不足时输出已获取的有效行。0=不启用（默认）")
    args = parser.parse_args()

    # 优先使用 --accounts-file（文件方式，避免 shell 引号转义问题）
    if args.accounts_file:
        try:
            with open(args.accounts_file, encoding="utf-8") as f:
                accounts_to_run = json.load(f)
            print(f"✅ 从 --accounts-file 载入 {len(accounts_to_run)} 个账号")
        except Exception as e:
            print(f"❌ --accounts-file 读取失败: {e}")
            return
    elif args.accounts_json:
        try:
            accounts_to_run = json.loads(args.accounts_json)
            print(f"✅ 从 --accounts-json 载入 {len(accounts_to_run)} 个账号")
        except Exception as e:
            print(f"❌ --accounts-json 解析失败: {e}")
            return
    elif args.names:
        accounts_to_run = [a for a in ACCOUNTS if a["name"] in args.names]
        if not accounts_to_run:
            print(f"❌ 未找到匹配账号：{args.names}")
            return
    else:
        accounts_to_run = ACCOUNTS
        if not accounts_to_run:
            print("❌ 无账号可分析。请通过 --accounts-json 或 --names 传入账号。")
            return

    target_count = args.target_count

    print("=" * 60)
    print("XHS 账号分析报表生成")
    if target_count:
        print(f"候选账号数：{len(accounts_to_run)}，目标有效账号：{target_count}，每账号分析 {NOTES_LIMIT} 篇")
    else:
        print(f"处理账号数：{len(accounts_to_run)}，每账号分析 {NOTES_LIMIT} 篇")
    print("=" * 60)

    if args.login_account:
        print("ℹ️  --login-account 已忽略：XHS 抓取已切换为 FlowLens 接口，不再使用 Playwright Cookie 账号池")

    fetcher = FlowLensXhsFetcher()
    print("🌐 启动 FlowLens XHS 会话...")
    fetcher.start()

    accounts_ordered = list(reversed(accounts_to_run))
    new_rows: list = []
    skipped_rows: list = []  # 保留跳过的行，候选不足时填充输出

    try:
        for i, account in enumerate(accounts_ordered, 1):
            if target_count and len(new_rows) >= target_count:
                break
            label = (f"[{len(new_rows)+1}/{target_count} | 候选{i}/{len(accounts_ordered)}]"
                     if target_count else f"[{i}/{len(accounts_ordered)}]")
            print(f"\n{label} {account['name']}")
            explore_data = account.get("explore_data")
            row = process_account(fetcher, account, explore_data)
            if target_count:
                if is_complete_analysis_row(row):
                    new_rows.append(row)
                    print(f"  ✅ 有效账号 {len(new_rows)}/{target_count}")
                else:
                    skipped_rows.append(row)
                    print(f"  ⏭  跳过（数据不足），已有 {len(new_rows)}/{target_count} 个有效账号")
            else:
                new_rows.append(row)
                if not is_complete_analysis_row(row):
                    print(f"  ⏭  跳过输出（{row.get('账号权重评级', '无数据')}）")
            if i < len(accounts_ordered):
                delay = random.uniform(2, 4)
                print(f"  ⏳ 等待 {delay:.1f}s...")
                time.sleep(delay)

        # 候选不足时，用跳过的行补位（保证输出不少于有效行数）
        if target_count and len(new_rows) < target_count:
            gap = target_count - len(new_rows)
            print(f"\n⚠️  候选池耗尽，有效账号 {len(new_rows)}/{target_count}，"
                  f"用 {min(gap, len(skipped_rows))} 个不完整行补位")
            new_rows.extend(skipped_rows[:gap])
    finally:
        fetcher.stop()

    output_path = str(get_analysis_workbook_path())
    existing_rows = load_existing_excel(output_path) if Path(output_path).exists() else []
    existing_names = {r.get("账号昵称", "") for r in existing_rows if r.get("账号昵称")}
    deduped_new = [r for r in new_rows if r.get("账号昵称") not in existing_names]
    merged_rows = existing_rows + deduped_new
    write_excel(merged_rows, output_path)
    if existing_rows:
        print(
            f"\n✅ 已追加 {len(deduped_new)} 条分析记录（跳过 {len(new_rows) - len(deduped_new)} 条重复），总计 {len(merged_rows)} 条"
        )
    else:
        print(f"\n✅ 已创建累计分析 Excel，共写入 {len(deduped_new)} 条记录")

    if isinstance(publish_manager, _PublishManagerStub):
        print(f"⚠️  publish_manager 不可用，跳过保存账号风格偏好: {publish_manager.import_error}", flush=True)
        return

    for acc in new_rows:
        account_name = acc.get("账号昵称", "").strip()
        if not account_name:
            print(f"⚠️  跳过无效账号记录（昵称为空）", flush=True)
            continue
        if not is_complete_analysis_row(acc):
            print(f"⚠️  跳过不完整分析结果，不写入账号风格偏好: {account_name}", flush=True)
            continue
        try:
            publish_manager.save_account_style(
                account_name,
                {
                    "avg_words": acc.get("平均文字字数", 300),
                    "style_desc": acc.get("图文风格描述", ""),
                    "image_style": acc.get("图片风格", ""),
                    "emotion": acc.get("情绪标签", ""),
                    "structure": acc.get("内容结构", ""),
                    "hook_examples": acc.get("Hook模板", ""),
                    "cta_examples": acc.get("CTA模板", ""),
                },
                source="analysis",
            )
            print(f"  ✅ 已保存账号风格偏好: {account_name}", flush=True)
        except Exception as e:
            print(f"  ⚠️  保存风格偏好失败 ({account_name}): {e}", flush=True)


if __name__ == "__main__":
    main()
